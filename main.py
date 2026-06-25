# === Imports ===
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed — rely on real environment variables

import os, logging, urllib.parse, threading
from datetime import datetime, date, timedelta
from typing import Tuple
from functools import wraps
from collections import defaultdict
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, abort, g, Response, current_app, send_file, jsonify
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import UniqueConstraint, or_, func, text
from typing import Optional
from datetime import timezone
from zoneinfo import ZoneInfo  # بايثون 3.9+ موجودة افتراضيًا

# === Paths (عرّف BASE_DIR أولاً) ===
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# === Flask app ===
app = Flask(__name__, template_folder="templates", static_folder="static")
import secrets as _secrets
_secret = os.environ.get("SECRET_KEY")
if not _secret:
    _secret = _secrets.token_hex(32)
    import logging as _logging
    _logging.getLogger(__name__).warning(
        "SECRET_KEY not set in environment -- using a random key. "
        "Sessions will not survive restarts. Set SECRET_KEY in your .env file."
    )
app.config["SECRET_KEY"] = _secret
app.config["SQLALCHEMY_ECHO"] = False  # تم تعطيله في production
application = app 

# (اختياري) logging بعد إنشاء app
try:
    os.makedirs(os.path.join(BASE_DIR, "tmp"), exist_ok=True)
    fh = logging.FileHandler(os.path.join(BASE_DIR, "tmp", "app.log"))
    fh.setLevel(logging.INFO)
    app.logger.setLevel(logging.INFO)
    app.logger.addHandler(fh)
    app.logger.info("App booted")
except Exception as e:
    print("log init error:", e)

# ---- Database (FORCE MySQL ONLY) ----

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_NAME = os.getenv("DB_NAME", "zappimiw_nsh")
DB_USER = os.getenv("DB_USER", "zappimiw_nsh")
DB_PASS = os.getenv("DB_PASS", "")  # لا تضع كلمة المرور هنا — استخدم environment variable

# تأكد أن كل القيم موجودة
if not all([DB_HOST, DB_NAME, DB_USER, DB_PASS]):
    raise Exception("❌ Database environment variables are missing")

# ترميز كلمة المرور
safe_pass = urllib.parse.quote_plus(DB_PASS)

# ربط MySQL فقط (بدون SQLite نهائياً)
app.config["SQLALCHEMY_DATABASE_URI"] = (
    f"mysql+pymysql://{DB_USER}:{safe_pass}@{DB_HOST}/{DB_NAME}?charset=utf8mb4"
)

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 280,
}

db = SQLAlchemy(app)

# ── CSRF Protection ──────────────────────────────────────────────
# API routes (Bearer token) are skipped; only web form POSTs are protected.
try:
    from flask_wtf.csrf import CSRFProtect, generate_csrf
    app.config["WTF_CSRF_CHECK_DEFAULT"] = False  # نتحكم يدوياً عبر before_request
    csrf = CSRFProtect(app)

    @app.before_request
    def _csrf_protect():
        if request.method in ("GET", "HEAD", "OPTIONS", "TRACE"):
            return
        if request.path.startswith("/api/"):
            return   # API routes use Bearer token — no CSRF needed
        csrf.protect()

    @app.after_request
    def _set_csrf_cookie(response):
        response.headers["X-CSRF-Token"] = generate_csrf()
        return response

except ImportError:
    pass  # Flask-WTF not installed yet — run: pip install Flask-WTF

# Ensure csrf_token is always available in templates even if Flask-WTF failed
if 'csrf_token' not in app.jinja_env.globals:
    app.jinja_env.globals['csrf_token'] = lambda: ''

# فلتر enumerate لـ Jinja2
app.jinja_env.filters['enumerate'] = enumerate

# ── Rate Limiter ─────────────────────────────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        app,
        key_func=get_remote_address,
        default_limits=[],
        storage_uri="memory://",
    )
except ImportError:
    limiter = None  # flask-limiter not installed — run: pip install flask-limiter

# decorator helper: applies rate limit when limiter is available
_login_limit = (limiter.limit("5 per minute") if limiter is not None else lambda f: f)

# ── Swagger / OpenAPI ────────────────────────────────────────────────
try:
    from flasgger import Swagger
    app.config["SWAGGER"] = {
        "title":   "NSH SafeTrack API",
        "version": "1.0",
        "uiversion": 3,
        "specs_route": "/api/docs/",
        "description": "REST API for NSH SafeTrack — iOS + Web SaaS platform",
        "securityDefinitions": {
            "Bearer": {
                "type": "apiKey",
                "name": "Authorization",
                "in":   "header",
                "description": "JWT token: Bearer <token>",
            }
        },
        "security": [{"Bearer": []}],
        "specs": [{"endpoint": "apispec", "route": "/api/docs/apispec.json"}],
    }
    swagger = Swagger(app)
except ImportError:
    pass  # flasgger not installed — run: pip install flasgger

# طباعة للتأكد
print(f"USING DATABASE: {DB_USER}@{DB_HOST}/{DB_NAME}")
from werkzeug.exceptions import NotFound

@app.before_request
def load_logged_in_user():
    if request.endpoint and request.endpoint.startswith('static'):
        return
    user_id = session.get("user_id")
    g.user = db.session.get(User, user_id) if user_id else None
    g.role = g.user.role if g.user else None
    g.company_id = session.get("company_id")
    g.company = db.session.get(Company, g.company_id) if g.company_id else None




# ===================== Constants =====================
WEIGHTS = {
    "targets_rows": 4,
    "target_per_row": 10,
    "perf_items": {
        "punctuality": 10,
        "quality": 10,
        "productivity": 10,
        "communication": 10,
        "problemsolving": 10,
        "compliance": 10,
    },
    "perf_scale_max": 5,
    "bands": {"excellent": 90, "good": 80, "satisfactory": 70},
}

SE_WEIGHTS = {
    "perf_items": {
        "leadership": 1,
        "communication": 1,
        "scheduling": 1,
        "compliance": 1,
        "team_support": 1,
        "reporting": 1,
    },
    "perf_scale_max": 5,
    "bands": {"excellent": 90, "good": 80, "satisfactory": 70},
}

# Week definition: Sunday → Thursday (Python weekday: Mon=0 ... Sun=6)
WEEK_START_WEEKDAY = 6  # Sunday
WEEK_END_WEEKDAY = 3    # Thursday


RIYADH_TZ = ZoneInfo("Asia/Riyadh")

def to_local(dt):
    if dt is None:
        return None
    # لو التاريخ naive (بدون tzinfo)، اعتبره UTC لأنه محفوظ بـ utcnow()
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(RIYADH_TZ)
    
@app.template_filter("local_dt")
def local_dt(dt, fmt="%Y-%m-%d %H:%M"):
    dt_local = to_local(dt)
    return dt_local.strftime(fmt) if dt_local else "—"


@app.template_filter("from_json")
def from_json_filter(value):
    import json as _json
    try:
        return _json.loads(value) if value else {}
    except Exception:
        return {}


@app.template_global("enumerate")
def jinja_enumerate(iterable, start=0):
    return enumerate(iterable, start)
    

class WarningReason(db.Model):
    """أسباب الإنذار — قابلة للإضافة من المشرف"""
    __tablename__ = "warning_reason"
    id         = db.Column(db.Integer, primary_key=True)
    text       = db.Column(db.String(255), nullable=False, unique=True)
    is_default = db.Column(db.Boolean, default=False)   # الأسباب الافتراضية
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class WarningSignature(db.Model):
    """توقيع الموظف على نموذج الإنذار"""
    __tablename__ = "warning_signature"
    id           = db.Column(db.Integer, primary_key=True)
    request_id   = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False, unique=True)
    employee_name= db.Column(db.String(255), nullable=True)
    signed_at    = db.Column(db.DateTime, default=datetime.utcnow)
    signature    = db.Column(db.Text, nullable=False)  # base64 PNG

    request = db.relationship("Request", foreign_keys=[request_id],
                              backref=db.backref("warning_sig", uselist=False))

class LeaveSignature(db.Model):
    """توقيع الموظف على نموذج الإجازة"""
    __tablename__ = "leave_signature"
    id           = db.Column(db.Integer, primary_key=True)
    request_id   = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False, unique=True)
    employee_name= db.Column(db.String(255), nullable=True)
    signed_at    = db.Column(db.DateTime, default=datetime.utcnow)
    signature    = db.Column(db.Text, nullable=False)  # base64 PNG
    request      = db.relationship("Request", foreign_keys=[request_id],
                                   backref=db.backref("leave_sig", uselist=False))

class PermissionSignature(db.Model):
    """توقيع الموظف على نموذج الاستئذان"""
    __tablename__ = "permission_signature"
    id           = db.Column(db.Integer, primary_key=True)
    request_id   = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False, unique=True)
    employee_name= db.Column(db.String(255), nullable=True)
    signed_at    = db.Column(db.DateTime, default=datetime.utcnow)
    signature    = db.Column(db.Text, nullable=False)  # base64 PNG
    request      = db.relationship("Request", foreign_keys=[request_id],
                                   backref=db.backref("permission_sig", uselist=False))

# ===================== Models =====================
# ========== Model: Request ==========
class Request(db.Model):
    __tablename__ = "requests"
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    supervisor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)      # كان users.id
    supervisor = db.relationship("User", foreign_keys=[supervisor_id])

    employee_id = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)    # كان employees.id
    employee = db.relationship("Employee", foreign_keys=[employee_id])

    type = db.Column(db.String(32), nullable=False, default="leave")
    start_date = db.Column(db.Date, nullable=True)
    end_date   = db.Column(db.Date, nullable=True)
    reason = db.Column(db.Text, nullable=True)

    status = db.Column(db.String(16), nullable=False, default="pending")
    decided_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    decided_at = db.Column(db.DateTime, nullable=True)
    admin_comment = db.Column(db.Text, nullable=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)

    __table_args__ = (
        db.Index("ix_requests_emp_created", "employee_id", "created_at"),
        db.Index("ix_requests_status", "status"),
    )

class HRTask(db.Model):
    __tablename__ = "hr_task"
    id          = db.Column(db.Integer, primary_key=True)
    request_id  = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False, unique=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)
    type        = db.Column(db.String(32), nullable=False)  # مطابق ل requests.type
    status      = db.Column(db.String(16), nullable=False, default="pending")  # pending/applied/cancelled
    created_at  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    applied_at  = db.Column(db.DateTime, nullable=True)
    applied_by      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    official_reason = db.Column(db.Text, nullable=True)
    warning_pdf     = db.Column(db.String(255), nullable=True)

    request  = db.relationship("Request",  foreign_keys=[request_id])
    employee = db.relationship("Employee", foreign_keys=[employee_id])
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)


# ===================== Models =====================

class DailyEvaluation(db.Model):
    __tablename__ = "daily_evaluation"
    id = db.Column(db.Integer, primary_key=True)

    # FK صحيحة: employee (مفرد) و user
    employee_id  = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)
    evaluator_id = db.Column(db.Integer, db.ForeignKey("user.id"),     nullable=False)

    eval_date = db.Column(db.Date, nullable=False)

    # Targets
    t1_text = db.Column(db.Text); t1_percent = db.Column(db.Float); t1_remarks = db.Column(db.Text)
    t2_text = db.Column(db.Text); t2_percent = db.Column(db.Float); t2_remarks = db.Column(db.Text)
    t3_text = db.Column(db.Text); t3_percent = db.Column(db.Float); t3_remarks = db.Column(db.Text)
    t4_text = db.Column(db.Text); t4_percent = db.Column(db.Float); t4_remarks = db.Column(db.Text)

    # Performance + comments
    p_punctuality = db.Column(db.Integer);    c_punctuality = db.Column(db.Text)
    p_quality = db.Column(db.Integer);        c_quality = db.Column(db.Text)
    p_productivity = db.Column(db.Integer);   c_productivity = db.Column(db.Text)
    p_communication = db.Column(db.Integer);  c_communication = db.Column(db.Text)
    p_problemsolving = db.Column(db.Integer); c_problemsolving = db.Column(db.Text)
    p_compliance = db.Column(db.Integer);     c_compliance = db.Column(db.Text)

    strengths = db.Column(db.Text)
    improvements = db.Column(db.Text)
    training_needed = db.Column(db.Text)

    targets_score = db.Column(db.Float)
    performance_score = db.Column(db.Float)
    total_score = db.Column(db.Float)
    overall_band = db.Column(db.String(30))

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)

    __table_args__ = (
        db.UniqueConstraint("employee_id", "eval_date", name="uq_daily_emp_date"),
    )

    employee  = db.relationship("Employee", backref=db.backref("daily_evaluations", lazy="dynamic"))
    evaluator = db.relationship("User",     backref=db.backref("daily_evaluations_given", lazy="dynamic"))




# compute_daily_scores: النسخة الصحيحة (targets 40% + perf 60%) موجودة أسفل بعد Models


class Company(db.Model):
    __tablename__ = "company"
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(255), nullable=False)
    slug        = db.Column(db.String(100), unique=True, nullable=False)
    plan        = db.Column(db.Enum("free", "pro"), default="free")
    max_users   = db.Column(db.Integer, default=50)
    is_active   = db.Column(db.Boolean, default=False)  # pending until super admin approves
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    owner_email = db.Column(db.String(255), nullable=True)


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    supervisor_code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(120), default="")
    role = db.Column(db.String(20), default="supervisor")  # supervisor | admin | site_supervisor | safety_officer | super_admin
    is_active = db.Column(db.Boolean, default=True)
    is_hidden = db.Column(db.Boolean, default=False)
    # SaaS additions
    company_id    = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    email         = db.Column(db.String(255), nullable=True)
    password_hash = db.Column(db.String(255), nullable=True)
    company       = db.relationship("Company", foreign_keys=[company_id])

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, pw)

    @property
    def is_supervisor(self):
        return self.role in ("supervisor", "site_supervisor")

class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)   # nullable للموظف غير المعيّن
    emp_number = db.Column(db.String(50), nullable=False, unique=True)
    name = db.Column(db.String(120), nullable=False)
    department = db.Column(db.String(120), default="")
    site = db.Column(db.String(120), default="")
    is_active = db.Column(db.Boolean, default=True)
    status = db.Column(db.String(20), nullable=False, default="active")   # active | resigned | unassigned
    resigned_at = db.Column(db.Date, nullable=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    

class Evaluation(db.Model):
    __table_args__ = (UniqueConstraint("employee_id", "week_start", "week_end", name="uq_emp_week"),)
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)
    evaluator_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    week_start = db.Column(db.Date, nullable=False)
    week_end   = db.Column(db.Date, nullable=False)

    # Targets (4 rows)
    t1_text = db.Column(db.String(255), default=""); t1_percent = db.Column(db.Float, default=0); t1_remarks = db.Column(db.String(255), default="")
    t2_text = db.Column(db.String(255), default=""); t2_percent = db.Column(db.Float, default=0); t2_remarks = db.Column(db.String(255), default="")
    t3_text = db.Column(db.String(255), default=""); t3_percent = db.Column(db.Float, default=0); t3_remarks = db.Column(db.String(255), default="")
    t4_text = db.Column(db.String(255), default=""); t4_percent = db.Column(db.Float, default=0); t4_remarks = db.Column(db.String(255), default="")

    # Performance (ratings 1..5 + comments)
    p_punctuality = db.Column(db.Integer, default=0); c_punctuality = db.Column(db.String(255), default="")
    p_quality = db.Column(db.Integer, default=0); c_quality = db.Column(db.String(255), default="")
    p_productivity = db.Column(db.Integer, default=0); c_productivity = db.Column(db.String(255), default="")
    p_communication = db.Column(db.Integer, default=0); c_communication = db.Column(db.String(255), default="")
    p_problemsolving = db.Column(db.Integer, default=0); c_problemsolving = db.Column(db.String(255), default="")
    p_compliance = db.Column(db.Integer, default=0); c_compliance = db.Column(db.String(255), default="")

    strengths = db.Column(db.Text, default="")
    improvements = db.Column(db.Text, default="")
    training_needed = db.Column(db.Text, default="")

    targets_score = db.Column(db.Float, default=0)
    perf_score    = db.Column(db.Float, default=0)
    total_score   = db.Column(db.Float, default=0)
    overall_band  = db.Column(db.String(30), default="")

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)

class SiteSupervisorMap(db.Model):
    __table_args__ = (UniqueConstraint("site_sup_id", "supervisor_id", name="uq_site_sup_pair"),)
    id = db.Column(db.Integer, primary_key=True)
    site_sup_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    supervisor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)

class SupervisorEvaluation(db.Model):
    __table_args__ = (UniqueConstraint("supervisor_id", "week_start", "week_end", name="uq_sup_week"),)
    id = db.Column(db.Integer, primary_key=True)
    supervisor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    evaluator_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    week_start = db.Column(db.Date, nullable=False)
    week_end   = db.Column(db.Date, nullable=False)

    # Targets (4 rows)
    t1_text = db.Column(db.String(255), default=""); t1_percent = db.Column(db.Float, default=0); t1_remarks = db.Column(db.String(255), default="")
    t2_text = db.Column(db.String(255), default=""); t2_percent = db.Column(db.Float, default=0); t2_remarks = db.Column(db.String(255), default="")
    t3_text = db.Column(db.String(255), default=""); t3_percent = db.Column(db.Float, default=0); t3_remarks = db.Column(db.String(255), default="")
    t4_text = db.Column(db.String(255), default=""); t4_percent = db.Column(db.Float, default=0); t4_remarks = db.Column(db.String(255), default="")

    # Performance (ratings 1..5 + comments)
    p_punctuality = db.Column(db.Integer, default=0); c_punctuality = db.Column(db.String(255), default="")
    p_quality = db.Column(db.Integer, default=0); c_quality = db.Column(db.String(255), default="")
    p_productivity = db.Column(db.Integer, default=0); c_productivity = db.Column(db.String(255), default="")
    p_communication = db.Column(db.Integer, default=0); c_communication = db.Column(db.String(255), default="")
    p_problemsolving = db.Column(db.Integer, default=0); c_problemsolving = db.Column(db.String(255), default="")
    p_compliance = db.Column(db.Integer, default=0); c_compliance = db.Column(db.String(255), default="")

    strengths = db.Column(db.Text, default="")
    improvements = db.Column(db.Text, default="")
    training_needed = db.Column(db.Text, default="")

    targets_score = db.Column(db.Float, default=0)
    perf_score    = db.Column(db.Float, default=0)
    total_score   = db.Column(db.Float, default=0)
    overall_band  = db.Column(db.String(30), default="")

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)

class Attendance(db.Model):
    __table_args__ = (UniqueConstraint("employee_id", "date", name="uq_emp_date"),)
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey("employee.id"), nullable=False)
    supervisor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), nullable=False)  # present / absent / leave
    remarks = db.Column(db.String(255), default="")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    employee = db.relationship("Employee", backref="attendances")
    supervisor = db.relationship("User", backref="attendances")


# ===================== Helpers =====================
def _apply_leave_attendance_for_request(req: Request):
    """
    يعلّم حضور الموظّف كـ Leave لكل يوم بين start_date و end_date (شاملاً).
    يتعامل مع التكرار (يحدّث السجل إن وُجد أو ينشئه إن لم يوجد).
    نستخدم supervisor_id من الطلب لأنه مطلوب في Attendance.
    نضع وسم صغير في remarks لتمييز السجلات المرتبطة بهذا الطلب (للتراجع إن لزم).
    """
    if not (req and req.employee_id and req.start_date and req.end_date):
        return

    marker = f"auto_leave_req_{req.id}"
    cur = req.start_date
    while cur <= req.end_date:
        rec = Attendance.query.filter_by(employee_id=req.employee_id, date=cur).first()
        if rec:
            # غيّر الحالة إلى إجازة
            rec.status = "leave"
            # أضف الوسم إذا مو موجود
            if marker not in (rec.remarks or ""):
                rec.remarks = (rec.remarks + " " + marker).strip() if rec.remarks else marker
        else:
            # أنشئ سجل جديد (supervisor_id مطلوب)
            rec = Attendance(
                employee_id=req.employee_id,
                supervisor_id=req.supervisor_id,  # صاحب الطلب
                date=cur,
                status="leave",
                remarks=marker,
            )
            db.session.add(rec)
        cur += timedelta(days=1)

def _apply_sick_attendance_for_request(req: Request):
    """يسجّل غياب (absent) للسكليف تلقائياً لكل يوم بين start_date و end_date."""
    if not (req and req.employee_id and req.start_date and req.end_date):
        return
    marker = f"auto_sick_req_{req.id}"
    cur = req.start_date
    while cur <= req.end_date:
        rec = Attendance.query.filter_by(employee_id=req.employee_id, date=cur).first()
        if rec:
            rec.status  = "absent"
            if marker not in (rec.remarks or ""):
                rec.remarks = (rec.remarks + " " + marker).strip() if rec.remarks else marker
        else:
            db.session.add(Attendance(
                employee_id=req.employee_id,
                supervisor_id=req.supervisor_id,
                date=cur,
                status="absent",
                remarks=marker,
            ))
        cur += timedelta(days=1)

def get_employee_summary(emp_id: int):
    from calendar import monthrange
    today = datetime.now(RIYADH_TZ).date()

    # ── الحضور: الشهر الحالي ──
    month_start = today.replace(day=1)
    _, days_in_month = monthrange(today.year, today.month)
    q_att = Attendance.query.filter(
        Attendance.employee_id == emp_id,
        Attendance.date >= month_start,
        Attendance.date <= today
    ).all()
    present = sum(1 for a in q_att if (a.status or "").lower() == "present")
    absent  = sum(1 for a in q_att if (a.status or "").lower() == "absent")
    leave   = sum(1 for a in q_att if (a.status or "").lower() == "leave")
    recorded = present + absent + leave
    att_rate = round(present / days_in_month * 100, 1) if present > 0 else 0.0

    # ── التقييمات: آخر 8 أسابيع ──
    weeks_since = today - timedelta(days=7 * 8)
    evals = (Evaluation.query
             .filter(Evaluation.employee_id == emp_id,
                     Evaluation.week_start >= weeks_since)
             .order_by(Evaluation.week_start.desc())
             .all())
    avg_score = round(sum((e.total_score or 0) for e in evals) / len(evals), 1) if evals else None

    # ── آخر تقييم أسبوعي ──
    last_eval = evals[0] if evals else None
    last_eval_data = None
    if last_eval:
        last_eval_data = {
            "week_start":  str(last_eval.week_start),
            "week_end":    str(last_eval.week_end),
            "total_score": last_eval.total_score,
            "band":        last_eval.overall_band,
            "targets":     last_eval.targets_score,
            "perf":        last_eval.perf_score,
        }

    # ── آخر تقييم يومي ──
    last_daily = (DailyEvaluation.query
                  .filter(DailyEvaluation.employee_id == emp_id)
                  .order_by(DailyEvaluation.eval_date.desc())
                  .first())

    return {
        # حضور الشهر الحالي
        "month_year":    f"{today.year}-{today.month:02d}",
        "present":       present,
        "absent":        absent,
        "leave":         leave,
        "recorded":      recorded,
        "days_in_month": days_in_month,
        "att_rate":      att_rate,
        # تقييمات آخر 8 أسابيع
        "avg_score_8w":  avg_score,
        "eval_count_8w": len(evals),
        "last_eval":     last_eval_data,
        # آخر تقييم يومي
        "last_daily_date":  str(last_daily.eval_date) if last_daily else None,
        "last_daily_total": last_daily.total_score if last_daily else None,
        "last_daily_band":  last_daily.overall_band if last_daily else None,
    }
    
# ✅ استبدل الدالة بالكامل بهذا الإصدار
def aggregate_week_from_dailies(emp_id: int, ws: date, we: date) -> Optional[Evaluation]:
    # اجلب الأيام (أحد..خميس)
    dailies = (DailyEvaluation.query
               .filter(DailyEvaluation.employee_id == emp_id,
                       DailyEvaluation.eval_date >= ws,
                       DailyEvaluation.eval_date <= we)
               .all())
    if not dailies:
        return None

    n = len(dailies)
    # متوسط الأرقام من الحقول الصحيحة
    avg_targets = sum((d.targets_score or 0) for d in dailies) / n
    avg_perf    = sum((d.performance_score or 0) for d in dailies) / n
    # نجمع المتوسطين بدل متوسط total اليومي
    avg_total   = round(avg_targets + avg_perf, 2)

    # حدد Band
    if   avg_total >= WEIGHTS["bands"]["excellent"]:
        band = "Excellent"
    elif avg_total >= WEIGHTS["bands"]["good"]:
        band = "Good"
    elif avg_total >= WEIGHTS["bands"]["satisfactory"]:
        band = "Satisfactory"
    else:
        band = "Needs Improvement"

    # أنشئ/حدّث Evaluation الأسبوعي
    ev = Evaluation.query.filter_by(employee_id=emp_id, week_start=ws, week_end=we).first()
    if not ev:
        # نأخذ آخر مُقيّم يومي كـ evaluator (أو مدير الموظف إذا تحب)
        last_evaluator = max(dailies, key=lambda d: d.created_at).evaluator_id
        ev = Evaluation(employee_id=emp_id, evaluator_id=last_evaluator,
                        week_start=ws, week_end=we)
        db.session.add(ev)

    ev.targets_score = round(avg_targets, 2)
    ev.perf_score    = round(avg_perf, 2)
    ev.total_score   = avg_total
    ev.overall_band  = band

    db.session.commit()
    return ev


def previous_week_range(ws: date) -> Tuple[date, date]:
    prev_ws = ws - timedelta(days=7)
    prev_we = prev_ws + timedelta(days=4)
    return prev_ws, prev_we


def compute_supervisor_scores(se: SupervisorEvaluation) -> None:
    items = {
        "leadership": getattr(se, "p_leadership", 0),
        "communication": se.p_communication,
        "scheduling": getattr(se, "p_scheduling", 0),
        "compliance": se.p_compliance,
        "team_support": getattr(se, "p_team_support", 0),
        "reporting": getattr(se, "p_reporting", 0),
    }
    per_item_weight = 100.0 / len(items)
    p_score = 0.0
    for _, rating in items.items():
        rating = int(rating or 0)
        p_score += (rating / SE_WEIGHTS["perf_scale_max"]) * per_item_weight

    total = round(p_score, 2)
    if total >= SE_WEIGHTS["bands"]["excellent"]:
        band = "Excellent"
    elif total >= SE_WEIGHTS["bands"]["good"]:
        band = "Good"
    elif total >= SE_WEIGHTS["bands"]["satisfactory"]:
        band = "Satisfactory"
    else:
        band = "Needs Improvement"

    se.perf_score = total
    se.total_score = total
    se.overall_band = band

def default_week_today() -> Tuple[date, date]:
    today = datetime.now(RIYADH_TZ).date()
    days_since_sun = (today.weekday() - WEEK_START_WEEKDAY) % 7
    start = today - timedelta(days=days_since_sun)
    end = start + timedelta(days=4)
    return start, end

def ensure_db_and_admin() -> None:
    """Create DB tables and bootstrap admin if ADMIN_ID is set."""
    with app.app_context():
        db.create_all()
        admin_id = os.environ.get("ADMIN_ID")
        if admin_id:
            existing = User.query.filter_by(supervisor_code=admin_id).first()
            if not existing:
                admin = User(supervisor_code=admin_id, name="Admin", role="admin", is_active=True)
                db.session.add(admin)
                db.session.commit()
                
# parse_date و validate_week_sun_to_thu: النسخ الصحيحة موجودة أسفل في قسم Auth & Helpers
def _safe_date(s):
    """Parse date or return None if invalid/empty."""
    try:
        return parse_date(s) if s else None
    except Exception:
        return None

def cur_user():
    uid = session.get("user_id")
    return db.session.get(User, uid) if uid else None

def login_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        if not cur_user():
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return inner

def admin_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        u = cur_user()
        if not u or u.role != "admin":
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return inner

def hr_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        u = cur_user()
        if not u or u.role != "hr":
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return inner

def super_admin_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        u = cur_user()
        if not u or u.role != "super_admin":
            abort(403)
        return f(*args, **kwargs)
    return inner

def cid():
    """Returns current company_id. None for super_admin (sees all)."""
    u = cur_user()
    if u and u.role == "super_admin":
        return None
    return session.get("company_id")

def apply_company_filter(query, model):
    """Adds company_id filter. Super_admin sees all; everyone else is scoped."""
    u = cur_user()
    if u and u.role == "super_admin":
        return query  # super_admin sees everything
    company_id = session.get("company_id")
    # Always filter — even if company_id is None (returns only unscoped rows,
    # which is safe; after migration all rows have company_id set)
    query = query.filter(model.company_id == company_id)
    return query

# يحدد مشرفًا من نص البحث (ID مطابق أو اسم مطابق بالكامل)
# يحدد مشرفًا من نص البحث (ID مطابق أو اسم مطابق بالكامل)
def find_supervisor_from_query(q: str):
    if not q:
        return None
    q = q.strip()
    # جرّب Supervisor ID (supervisor_code) أولًا
    sup = User.query.filter(
        User.role == "supervisor",
        User.supervisor_code == q
    ).first()
    if sup:
        return sup
    # جرّب اسم مطابق بالكامل (case-insensitive)
    return (User.query
            .filter(User.role == "supervisor",
                    func.lower(User.name) == q.lower())
            .first())


# ===================== Auth & Helpers =====================
# parse_date و validate_week_sun_to_thu: النسخ الصحيحة موجودة أسفل
def compute_scores(
    ev,
    target_percent_attrs=("t1_percent","t2_percent","t3_percent","t4_percent"),
    rating_attrs=("p_punctuality","p_quality","p_productivity",
                  "p_communication","p_problemsolving","p_compliance"),
):
    def clamp(x, lo, hi):
        try:
            return max(lo, min(hi, x))
        except Exception:
            return lo

    # ===== Targets (40) =====
    target_values = []
    for attr in target_percent_attrs:
        v = getattr(ev, attr, None)
        try:
            v = float(v) if v not in (None, "") else 0.0
        except Exception:
            v = 0.0
        target_values.append(clamp(v, 0.0, 100.0))

    n_targets = len(target_values) or 1
    per_target_points = 40.0 / n_targets
    ev.targets_score = round(sum((v / 100.0) * per_target_points for v in target_values), 2)

    # ===== Performance (60) — الآن الافتراضي 0 =====
    rating_values = []
    for attr in rating_attrs:
        raw = getattr(ev, attr, None)

        # لو فاضية أو None → 0
        if raw in (None, "", 0):
            r = 0
        else:
            try:
                r = int(raw)
            except Exception:
                r = 0

        # خليه يقبل 0 إلى 5
        rating_values.append(clamp(r, 0, 5))

    n_ratings = len(rating_values) or 1
    per_item_points = 60.0 / n_ratings
    ev.perf_score = round(sum((r / 5.0) * per_item_points for r in rating_values), 2)

    # ===== Total & Band =====
    ev.total_score = round(ev.targets_score + ev.perf_score, 2)
    ev.overall_band = ("Excellent" if ev.total_score >= 90 else
                       "Good" if ev.total_score >= 80 else
                       "Satisfactory" if ev.total_score >= 70 else
                       "Needs Improvement")

# (اختياري) لو عندك استدعاءات قديمة:
# compute_weekly_scores = compute_scores
                       
# ==== Auth & helpers glue (ضعها بعد الـ Helpers) ====


# 5) parse_date بسيطة متسامحة مع YYYY-MM-DD
def parse_date(s: str) -> date:
    s = (s or "").strip()
    try:
        return date.fromisoformat(s)          # 2025-10-16
    except Exception:
        # fallback: dd/mm/yyyy أو dd-mm-yyyy
        for sep in ("/", "-"):
            parts = s.split(sep)
            if len(parts) == 3 and len(parts[0]) <= 2:
                d, m, y = map(int, parts)
                return date(y, m, d)
        raise

# 6) تحقق أن الأسبوع أحد→خميس
def validate_week_sun_to_thu(ws: date, we: date):
    if not ws or not we:
        return False, "Week start/end are required."
    if we != ws + timedelta(days=4):
        return False, "Week must be 5 days (Sun→Thu)."
    if ws.weekday() != WEEK_START_WEEKDAY or we.weekday() != WEEK_END_WEEKDAY:
        return False, "Start must be Sunday and end must be Thursday."
    return True, ""
# --- Safe scoring function for daily evaluation ---

# --- Safe scoring function for daily evaluation ---
def compute_daily_scores(de):
    """
    حساب الديلي بناءً على البنود المدخلة فقط.
    """
    # ===== Daily Targets (40) =====
    filled_percents = []
    for i in range(1, 5):
        txt = getattr(de, f"t{i}_text", "") or ""
        pct = getattr(de, f"t{i}_percent", 0) or 0
        try:
            pct = float(pct)
        except Exception:
            pct = 0.0

        # احسبه لو فيه نص أو نسبة
        if txt.strip() or pct > 0:
            pct = max(0.0, min(100.0, pct))
            filled_percents.append(pct)

    if filled_percents:
        n_targets = len(filled_percents)
        per_target_points = 40.0 / n_targets
        de.targets_score = round(
            sum((p / 100.0) * per_target_points for p in filled_percents),
            2
        )
    else:
        de.targets_score = 0.0

    # ===== Performance (60) =====
    rating_fields = [
        "p_punctuality",
        "p_quality",
        "p_productivity",
        "p_communication",
        "p_problemsolving",
        "p_compliance",
    ]
    rating_values = []
    for f in rating_fields:
        r = getattr(de, f, 0) or 0
        try:
            r = int(r)
        except Exception:
            r = 0
        r = max(0, min(5, r))
        rating_values.append(r)

    if rating_values:
        n_ratings = len(rating_values)
        per_item_points = 60.0 / n_ratings
        de.perf_score = round(
            sum((r / 5.0) * per_item_points for r in rating_values),
            2
        )
    else:
        de.perf_score = 0.0

    # ===== Total & Band =====
    de.total_score = round(de.targets_score + de.perf_score, 2)
    de.overall_band = (
        "Excellent" if de.total_score >= 90 else
        "Good" if de.total_score >= 80 else
        "Satisfactory" if de.total_score >= 70 else
        "Needs Improvement"
    )


@app.get("/logo")
def logo():
    return app.send_static_file("img/logo.png")
    

@app.route("/daily2/evaluate/<int:emp_id>/new", methods=["GET", "POST"])
@login_required
def daily2_evaluate_new(emp_id):
    u = cur_user()
    emp = (Employee.query.get_or_404(emp_id) if u.role=="admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404())

    if request.method == "POST":
        # تاريخ اليوم أو اللي اختاره
        d_raw = request.form.get("eval_date") or date.today().isoformat()
        y, m, d = map(int, d_raw.split("-"))
        eval_date = date(y, m, d)

        # منع تكرار نفس اليوم
        exists = DailyEvaluation.query.filter_by(employee_id=emp.id, eval_date=eval_date).first()
        if exists:
            flash("Daily evaluation for this date already exists.", "warning")
            return redirect(url_for("daily_report_employee", emp_id=emp.id, d=eval_date.isoformat()))

        de = DailyEvaluation(
            employee_id=emp.id,
            evaluator_id=u.id,
            eval_date=eval_date,

            # Daily targets
            t1_text=request.form.get("t1_text",""),
            t1_percent=request.form.get("t1_percent") or 0,
            t1_remarks=request.form.get("t1_remarks",""),
            t2_text=request.form.get("t2_text",""),
            t2_percent=request.form.get("t2_percent") or 0,
            t2_remarks=request.form.get("t2_remarks",""),
            t3_text=request.form.get("t3_text",""),
            t3_percent=request.form.get("t3_percent") or 0,
            t3_remarks=request.form.get("t3_remarks",""),
            t4_text=request.form.get("t4_text",""),
            t4_percent=request.form.get("t4_percent") or 0,
            t4_remarks=request.form.get("t4_remarks",""),

            # Performance
            p_punctuality=request.form.get("p_punctuality", type=int),
            c_punctuality=request.form.get("c_punctuality",""),
            p_quality=request.form.get("p_quality", type=int),
            c_quality=request.form.get("c_quality",""),
            p_productivity=request.form.get("p_productivity", type=int),
            c_productivity=request.form.get("c_productivity",""),
            p_communication=request.form.get("p_communication", type=int),
            c_communication=request.form.get("c_communication",""),
            p_problemsolving=request.form.get("p_problemsolving", type=int),
            c_problemsolving=request.form.get("c_problemsolving",""),
            p_compliance=request.form.get("p_compliance", type=int),
            c_compliance=request.form.get("c_compliance",""),

            strengths=request.form.get("strengths",""),
            improvements=request.form.get("improvements",""),
            training_needed=request.form.get("training_needed",""),
            company_id=cid(),
        )

        compute_daily_scores(de)
        db.session.add(de)
        db.session.commit()

        flash("Daily evaluation saved.", "success")
        return redirect(url_for("daily_report_employee", emp_id=emp.id, d=eval_date.isoformat()))

    # GET
    today = date.today()
    return render_template("daily_eval_form_v2.html", employee=emp, d=today, weights=WEIGHTS)
    
@app.route("/daily2/reports/employee/<int:emp_id>")
@login_required
def daily2_report_employee(emp_id):
    u = cur_user()
    emp = (Employee.query.get_or_404(emp_id) if u.role == "admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404())

    d_str = request.args.get("d")
    if not d_str:
        flash("Missing date.", "warning")
        return redirect(url_for("employees"))

    d_val = date.fromisoformat(d_str)
    de = DailyEvaluation.query.filter_by(employee_id=emp.id, eval_date=d_val).first_or_404()

    evaluator = db.session.get(User, de.evaluator_id)
    return render_template(
        "daily_report_employee_v2.html",
        employee=emp,
        de=de,
        evaluator_code=evaluator.supervisor_code if evaluator else "",
        evaluator_name=(evaluator.name if (evaluator and evaluator.name) else "")
    )


# ===================== Routes =====================
@app.route("/requests/new", methods=["GET", "POST"])
@login_required
def request_new():
    u = cur_user()
    if not (u and getattr(u, "role", None) in ("supervisor", "site_supervisor", "admin")):
        abort(403)

    # فلترة موظفي المشرف فقط (حسب ما طبّقناه سابقًا بالـ join على User أو user_id=u.id)
    employees = (Employee.query
                 .filter_by(user_id=u.id, is_active=True)
                 .order_by(Employee.name.asc())
                 .all())

    def _parse_date(s):
        s = (s or "").strip()
        if not s: return None
        from datetime import datetime
        try: return datetime.strptime(s, "%Y-%m-%d").date()
        except: return None

    if request.method == "POST":
        # التقاط
        emp_id     = request.form.get("employee_id")
        rtype      = (request.form.get("type") or "").strip().lower()
        reason     = (request.form.get("reason") or "").strip()
        start_date = _parse_date(request.form.get("start_date"))
        end_date   = _parse_date(request.form.get("end_date"))
        from_time  = (request.form.get("from_time") or "").strip()
        to_time    = (request.form.get("to_time") or "").strip()
        priority   = (request.form.get("priority") or "").strip().lower()

        # تحقق أساسي
        try: emp_id = int(emp_id or 0)
        except: emp_id = 0
        emp = db.session.get(Employee, emp_id) if emp_id else None
        if not emp or emp.user_id != u.id:
            flash("Please choose a valid employee.", "danger")
            return render_template("requests_new.html", employees=employees,
                                   pre_emp_id=emp_id, pre_type=rtype,
                                   pre_start=request.form.get("start_date"),
                                   pre_end=request.form.get("end_date"),
                                   pre_from_time=from_time, pre_to_time=to_time,
                                   pre_priority=priority, pre_reason=reason)

        # قواعد خفيفة لكل نوع (بدون ترحيل DB)
        if rtype in {"leave","shift","overtime","absence","late"}:
            if not start_date:
                flash("Start date is required for the selected type.", "danger")
                return render_template("requests_new.html", employees=employees,
                                       pre_emp_id=emp_id, pre_type=rtype,
                                       pre_start=request.form.get("start_date"),
                                       pre_end=request.form.get("end_date"),
                                       pre_priority=priority, pre_reason=reason)
        if rtype == "permission":
            # ننسّق الوقت داخل الـ reason
            tag = f"[Permission {from_time or '?'}→{to_time or '?'}]"
            reason = f"{tag} {reason}".strip()
            # تواريخ ليست مطلوبة هنا
            start_date = start_date or None
            end_date   = end_date or None
        if rtype == "warning":
            # نضيف أولوية لو موجودة
            if priority in {"low","normal","high"}:
                reason = f"[Warning {priority}] {reason}".strip()

        # الإنشاء
        r = Request(
            supervisor_id=u.id,
            employee_id=emp.id,
            type=rtype if rtype in {"leave","permission","warning","late","other"} else "other",
            start_date=start_date,
            end_date=end_date,
            reason=reason,
            status="pending",
            company_id=cid(),
        )
        db.session.add(r); db.session.commit()
        flash("Request submitted.", "success")
        return redirect(url_for("requests_mine"))

    # GET
    return render_template("requests_new.html", employees=employees,
                           pre_emp_id=None, pre_type="leave")



@app.route("/requests/mine")
@login_required
def requests_mine():
    u = cur_user()
    reqs = Request.query.filter(Request.supervisor_id == u.id) \
                        .order_by(Request.created_at.desc()).all()
    # نجلب HRTask لكل طلب لمعرفة official_reason وحالة التطبيق
    req_ids = [r.id for r in reqs]
    tasks_map = {}
    if req_ids:
        for t in HRTask.query.filter(HRTask.request_id.in_(req_ids)).all():
            tasks_map[t.request_id] = t
    return render_template("requests_mine.html", rows=reqs, tasks_map=tasks_map)


@app.post("/daily/aggregate/week")
@login_required
def daily_aggregate_week():
    # احصل على أسبوع مستهدف
    ws = parse_date(request.form.get("week_start"))
    we = parse_date(request.form.get("week_end"))
    ok, msg = validate_week_sun_to_thu(ws, we)
    if not ok:
        flash(msg, "danger")
        return redirect(request.referrer or url_for("admin_report_picker"))

    u = cur_user()
    # من يشمل؟
    if u.role == "admin":
        emp_ids = [e.id for e in Employee.query.filter_by(is_active=True).all()]
    else:
        emp_ids = [e.id for e in Employee.query.filter_by(user_id=u.id, is_active=True).all()]

    done = 0
    for emp_id in emp_ids:
        if aggregate_week_from_dailies(emp_id, ws, we):
            done += 1

    flash(f"Aggregated weekly evaluations for {done} employees.", "success")
    # رجّع للأدمن إلى /admin/reports/all أو للمشرف لصفحته
    if u.role == "admin":
        return redirect(url_for("admin_reports_all",
                                week_start=ws.isoformat(), week_end=we.isoformat()))
    return redirect(url_for("report_supervisor",
                            week_start=ws.isoformat(), week_end=we.isoformat()))



@app.route("/admin/employee/<int:emp_id>/summary", methods=["GET"])
@login_required
@admin_required
def employee_summary(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    summary = get_employee_summary(emp_id)
    return render_template("employee_summary.html", emp=emp, summary=summary)


@app.route("/daily/evaluate/<int:emp_id>/new", methods=["GET", "POST"])
@login_required
def daily_evaluate_new(emp_id):
    u = cur_user()
    # المشرف يشوف موظفينه فقط، الأدمن يشوف الكل
    emp = (Employee.query.get_or_404(emp_id) if u.role=="admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404())

    if request.method == "POST":
        try:
            # 1) التاريخ
            eval_date_raw = request.form.get("eval_date")
            if not eval_date_raw:
                flash("Date is required.", "danger")
                return redirect(url_for("daily_evaluate_new", emp_id=emp.id))
            eval_date = parse_date(eval_date_raw)  # تأكد إنها موجودة: strptime("%Y-%m-%d")

            # 2) منع التكرار
            exists = DailyEvaluation.query.filter_by(
                employee_id=emp.id, eval_date=eval_date
            ).first()
            if exists:
                flash("A daily evaluation for this date already exists.", "warning")
                return redirect(url_for("daily_report_employee", emp_id=emp.id, d=eval_date.isoformat()))

            # 3) قراءة آمنة للقيم + قصّ المدى
            gi = lambda n, lo=0, hi=5: max(lo, min(hi, int(request.form.get(n) or 0)))
            gf = lambda n, lo=0, hi=100: max(lo, min(hi, float(request.form.get(n) or 0)))

            de = DailyEvaluation(
                employee_id=emp.id,
                evaluator_id=u.id,
                eval_date=eval_date,

                t1_text=request.form.get("t1_text",""),
                t1_percent=gf("t1_percent"), t1_remarks=request.form.get("t1_remarks",""),
                t2_text=request.form.get("t2_text",""),
                t2_percent=gf("t2_percent"), t2_remarks=request.form.get("t2_remarks",""),
                t3_text=request.form.get("t3_text",""),
                t3_percent=gf("t3_percent"), t3_remarks=request.form.get("t3_remarks",""),
                t4_text=request.form.get("t4_text",""),
                t4_percent=gf("t4_percent"), t4_remarks=request.form.get("t4_remarks",""),

                p_punctuality=gi("p_punctuality"), c_punctuality=request.form.get("c_punctuality",""),
                p_quality=gi("p_quality"),         c_quality=request.form.get("c_quality",""),
                p_productivity=gi("p_productivity"), c_productivity=request.form.get("c_productivity",""),
                p_communication=gi("p_communication"), c_communication=request.form.get("c_communication",""),
                p_problemsolving=gi("p_problemsolving"), c_problemsolving=request.form.get("c_problemsolving",""),
                p_compliance=gi("p_compliance"),     c_compliance=request.form.get("c_compliance",""),

                strengths=request.form.get("strengths",""),
                improvements=request.form.get("improvements",""),
                training_needed=request.form.get("training_needed",""),
                company_id=cid(),
            )

            compute_daily_scores(de)
            db.session.add(de)
            db.session.commit()

            flash("Daily evaluation saved.", "success")
            return redirect(url_for("daily_report_employee", emp_id=emp.id, d=eval_date.isoformat()))

        except Exception as e:
            # مهم: رجوع المعاملة + تسجيل الخطأ
            db.session.rollback()
            import traceback
            app.logger.error("Error saving daily evaluation:\n%s", traceback.format_exc())
            flash("Error while saving the daily evaluation.", "danger")
            return redirect(url_for("daily_evaluate_new", emp_id=emp.id))

    # GET — نفس تمبليتك الحالية
    d = date.today()
    return render_template("daily_eval_form.html", employee=emp, d=d, weights=WEIGHTS)


@app.route("/daily/reports/employee/<int:emp_id>")
@login_required
def daily_report_employee(emp_id):
    u = cur_user()

    # 1) get employee
    emp = Employee.query.get_or_404(emp_id)

    # 2) permission check
    allowed = False

    # admin can see all
    if u.role == "admin":
        allowed = True

    # supervisor of this employee
    elif emp.user_id == u.id:
        allowed = True

    # site supervisor: check mapping
    elif u.role == "site_supervisor":
        link = (db.session.query(SiteSupervisorMap)
                .filter_by(site_sup_id=u.id, supervisor_id=emp.user_id)
                .first())
        if link:
            allowed = True

    if not allowed:
        abort(403)

    # 3) date
    d_str = request.args.get("d")
    if not d_str:
        flash("Missing date.", "warning")
        return redirect(url_for("employees"))
    d_val = parse_date(d_str)

    # 4) daily evaluation
    de = (DailyEvaluation.query
          .filter_by(employee_id=emp.id, eval_date=d_val)
          .first_or_404())

    # important: recompute
    compute_daily_scores(de)

    evaluator = db.session.get(User, de.evaluator_id)
    return render_template(
        "daily_report_employee_v2.html",
        employee=emp,
        de=de,
        evaluator_code=(evaluator.supervisor_code if evaluator else ""),
        evaluator_name=(evaluator.name if (evaluator and evaluator.name) else "")
    )





@app.route("/daily/reports")
@login_required
def daily_reports():
    u = cur_user()

    d_str = request.args.get("d")
    if d_str:
        d_val = parse_date(d_str)
    else:
        # لو ما وصل تاريخ، استخدم تاريخ اليوم
        d_val = date.today()

    q = (db.session.query(DailyEvaluation, Employee)
         .join(Employee, DailyEvaluation.employee_id == Employee.id))

    _cid = cid()
    if _cid is not None:
        q = q.filter(DailyEvaluation.company_id == _cid)

    if u.role != "admin":
        q = q.filter(Employee.user_id == u.id)

    rows = (q.filter(DailyEvaluation.eval_date == d_val)
              .order_by(Employee.name.asc()).all())

    return render_template("daily_reports.html", rows=rows, d=d_val)


from datetime import date  # تأكد أنه موجود فوق

@app.route("/admin/reports/daily/all")
@admin_required
def admin_reports_daily_all():
    # 1) التاريخ
    d_s = request.args.get("eval_date")
    if d_s:
        d = parse_date(d_s)
    else:
        d = date.today()

    # 2) الفلتر اللي يكتبه الأدمن
    q = (request.args.get("q") or "").strip()
    focus_sup = find_supervisor_from_query(q)

    # 3) الجدول الرئيسي (نفس اللي كان عندك)
    query = (db.session.query(DailyEvaluation, Employee, User)
             .join(Employee, DailyEvaluation.employee_id == Employee.id)
             .join(User, Employee.user_id == User.id)
             .filter(DailyEvaluation.eval_date == d))
    if cid():
        query = query.filter(Employee.company_id == cid())

    if focus_sup:
        query = query.filter(Employee.user_id == focus_sup.id)

    if q and not focus_sup:
        like = f"%{q}%"
        query = query.filter(or_(
            Employee.name.ilike(like),
            Employee.emp_number.ilike(like),
            Employee.department.ilike(like),
            Employee.site.ilike(like),
            User.supervisor_code.ilike(like),
            DailyEvaluation.overall_band.ilike(like),
        ))

    rows = query.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    # 4) كل المشرفين (علشان نطلع تحت "من اللي ما قيّم")
    supervisors = (
        apply_company_filter(User.query.filter_by(role="supervisor", is_active=True), User)
        .order_by(User.name.asc(), User.supervisor_code.asc())
        .all()
    )

    all_sup_stats = []  # هنا بنحط كل مشرف مع تغطيته
    for s in supervisors:
        # موظفين هذا المشرف
        emps = (Employee.query
                .filter_by(user_id=s.id, is_active=True)
                .order_by(Employee.name.asc())
                .all())
        total_emp = len(emps)
        emp_ids = [e.id for e in emps]

        if emp_ids:
            done_rows = (db.session.query(DailyEvaluation.employee_id)
                         .filter(DailyEvaluation.eval_date == d,
                                 DailyEvaluation.employee_id.in_(emp_ids))
                         .all())
            done_ids = {r[0] for r in done_rows}
        else:
            done_ids = set()

        done_emp_count = len(done_ids)
        coverage = (done_emp_count / total_emp * 100.0) if total_emp else 0.0
        missing_emps = [e for e in emps if e.id not in done_ids]

        all_sup_stats.append({
            "sup": s,
            "total_emp": total_emp,
            "done_emp_count": done_emp_count,
            "coverage": coverage,
            "missing_emps": missing_emps,
        })

    # 5) اقتراحات البحث اللي كانت عندك
    sup_suggestions, seen = [], set()
    for s in supervisors:
        val = (s.supervisor_code or "").strip()
        if val and val.lower() not in seen:
            sup_suggestions.append((val, f"{s.name or '—'} — ID: {s.supervisor_code}"))
            seen.add(val.lower())
    for s in supervisors:
        nm = (s.name or "").strip()
        if nm and nm.lower() not in seen:
            sup_suggestions.append((nm, f"{nm} — ID: {s.supervisor_code}"))
            seen.add(nm.lower())

    return render_template(
        "report_admin_daily_all.html",
        eval_date=d,
        rows=rows,
        q=q,
        focus_sup=focus_sup,
        supervisors=supervisors,
        sup_suggestions=sup_suggestions,
        all_sup_stats=all_sup_stats,   # <= الجديد
    )



# ---------- Admin: Requests ----------


@app.route("/attendance", methods=["GET", "POST"])
@admin_required
def attendance_admin():
    """
    صفحة الأدمن الرئيسية للحضور:
      - mode=day  : تاريخ واحد (d)
      - mode=week : أسبوع (ws..we) أحد->خميس
    """
    mode = request.args.get("mode") or "day"

    if request.method == "POST":
        mode = request.form.get("mode") or "day"
        if mode == "day":
            d = request.form.get("day_date") or ""
            return redirect(url_for("attendance_admin", mode="day", d=d))
        else:
            ws = request.form.get("week_start") or ""
            we = request.form.get("week_end") or ""
            # لو وضع أسبوع فقط تاريخ بداية، احسب النهاية = +4 أيام
            if ws and not we:
                try:
                    ws_dt = parse_date(ws)
                    we = (ws_dt + timedelta(days=4)).isoformat()
                except:
                    pass
            return redirect(url_for("attendance_admin", mode="week", ws=ws, we=we))

    # قراءة المعطيات من الquerystring
    if mode == "week":
        ws = _safe_date(request.args.get("ws")) or default_week_today()[0]
        we = _safe_date(request.args.get("we")) or (ws + timedelta(days=4))
        # تأكد أحد→خميس
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            ws, we = default_week_today()
    else:
        d = _safe_date(request.args.get("d")) or date.today()

    # الموظفون
    _emp_q = (db.session.query(Employee, User)
              .join(User, Employee.user_id == User.id)
              .filter(Employee.is_active == True))
    _cid = cid()
    if _cid is not None:
        _emp_q = _emp_q.filter(Employee.company_id == _cid)
    employees = _emp_q.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    rows = []
    kpi = {"total_days": 0, "total_present": 0}

    if mode == "day":
        emp_ids = [emp.id for emp, _ in employees]
        atts = []
        if emp_ids:
            atts = (Attendance.query
                    .filter(Attendance.employee_id.in_(emp_ids),
                            Attendance.date == d)
                    .all())
        att_by_emp = {a.employee_id: a for a in atts}
        for emp, sup in employees:
            a = att_by_emp.get(emp.id)
            rows.append({
                "emp": emp,
                "sup": sup,
                "status": (a.status if a else None),
                "remarks": (a.remarks if a else ""),
                "marked_by": (db.session.get(User, a.supervisor_id) if a else None)
            })
        # KPI يومي بسيط
        kpi["total_days"] = len([r for r in rows if r["status"]])
        kpi["total_present"] = len([r for r in rows if r["status"] == "present"])

        return render_template("attendance_admin.html",
                               mode="day", day_date=d, rows=rows, kpi=kpi)

    else:  # mode == "week"
        # اجلب كل الحضور ضمن المدى
        emp_ids = [emp.id for emp, _ in employees]
        atts = []
        if emp_ids:
            atts = (Attendance.query
                    .filter(Attendance.employee_id.in_(emp_ids),
                            Attendance.date >= ws,
                            Attendance.date <= we)
                    .all())
        # احصائيات لكل موظف
        stats = defaultdict(lambda: {"present": 0, "absent": 0, "leave": 0, "total": 0})
        for a in atts:
            stats[a.employee_id]["total"] += 1
            if a.status in ("present", "absent", "leave"):
                stats[a.employee_id][a.status] += 1
        for emp, sup in employees:
            s = stats[emp.id]
            pct = (s["present"] / s["total"] * 100.0) if s["total"] else 0.0
            kpi["total_days"] += s["total"]
            kpi["total_present"] += s["present"]
            rows.append({
                "emp": emp,
                "sup": sup,
                "present": s["present"],
                "absent": s["absent"],
                "leave": s["leave"],
                "total": s["total"],
                "percent": round(pct, 1)
            })

        return render_template("attendance_admin.html",
                               mode="week", ws=ws, we=we, rows=rows, kpi=kpi)


from sqlalchemy.exc import IntegrityError
from sqlalchemy import text
@app.route("/employees/add", methods=["POST"], endpoint="employees_add", strict_slashes=False)
@login_required
def employees_add():
    u = cur_user()
    
    # فقط السوبرفايزر
    if not u or getattr(u, "role", "") != "supervisor":
        abort(403)

    emp_number = (request.form.get("emp_number") or "").strip()
    name       = (request.form.get("name") or "").strip()
    department = (request.form.get("department") or "").strip()
    site       = (request.form.get("site") or "").strip()

    if not emp_number or not name:
        flash("Employee number and name are required.", "danger")
        return redirect(url_for("employees"))

    # رقم الموظف أرقام فقط
    if not emp_number.isdigit():
        flash("Employee number must contain digits only (0-9).", "danger")
        return redirect(url_for("employees"))

    # بحث عن الموظف (نشط أو غير نشط)
    emp = Employee.query.filter_by(emp_number=emp_number).first()

    if emp:
        # إعادة تفعيل إذا كان ملغي
        if not emp.is_active:
            emp.is_active = True

        # نقل الملكية إلى السوبرفايزر الحالي
        if emp.user_id != u.id:
            emp.user_id = u.id
            db.session.commit()
            flash(f"Employee {emp.name} ({emp.emp_number}) reactivated and assigned to you.", "success")
        else:
            flash(f"Employee {emp.name} ({emp.emp_number}) is already under your account.", "info")

        return redirect(url_for("employees"))

    # إنشاء موظف جديد
    try:
        emp = Employee(
            emp_number=emp_number,
            name=name,
            department=department,
            site=site,
            user_id=u.id,
            is_active=True,
            company_id=cid(),
        )
        db.session.add(emp)
        db.session.commit()
        flash(f"Employee {emp.name} ({emp.emp_number}) added successfully.", "success")

    except IntegrityError:
        db.session.rollback()
        flash("Duplicate employee number. Try again.", "danger")

    return redirect(url_for("employees"))

@app.route("/employees/<int:emp_id>/deactivate", methods=["POST"], endpoint="employees_deactivate", strict_slashes=False)
@login_required
def employees_deactivate(emp_id):
    u = cur_user()

    # فقط السوبرفايزر
    if not u or getattr(u, "role", "") != "supervisor":
        abort(403)

    emp = Employee.query.get_or_404(emp_id)

    # تأكد أنه تحت حساب السوبرفايزر نفسه
    if emp.user_id != u.id:
        abort(403)

    emp.is_active = False

    try:
        db.session.commit()
        flash(f"Employee {emp.name} ({emp.emp_number}) deactivated.", "success")
    except:
        db.session.rollback()
        flash("Error occurred while deactivating employee.", "danger")

    return redirect(url_for("employees"))



@app.route("/admin/requests", methods=["GET", "POST"])
@login_required
@admin_required
def requests_inbox():
    status = request.args.get("status", "pending")
    qterm  = (request.args.get("q") or "").strip()

    q = apply_company_filter(Request.query, Request)

    # فلتر الحالة (لو مو "all")
    if status and status != "all":
        q = q.filter(Request.status == status)

    # فلتر البحث بالاسم / رقم الموظف
    if qterm:
        q = (q.join(Employee, Request.employee_id == Employee.id)
               .filter(or_(
                   Employee.emp_number.like(f"%{qterm}%"),
                   Employee.name.ilike(f"%{qterm}%"),
               )))

    rows = q.order_by(Request.created_at.desc()).all()

    # 1) رابط "New" لو الراوت موجود
    request_new_url = None
    if "request_new" in current_app.view_functions:
        request_new_url = url_for("request_new")

    # 2) خرائط روابط ملخص الموظف حسب الراوتات المتاحة في مشروعك
    # جرّب بالترتيب: employee_summary → admin_employee_summary → admin_employee_history
    employee_summary_url_map = {}
    endpoint_candidates = ["employee_summary", "admin_employee_summary", "admin_employee_history"]
    target_endpoint = next((ep for ep in endpoint_candidates if ep in current_app.view_functions), None)

    if target_endpoint:
        for r in rows:
            if r.employee_id:
                try:
                    employee_summary_url_map[r.employee_id] = url_for(target_endpoint, emp_id=r.employee_id)
                except Exception:
                    # لو الراوت يطلب باراميتر باسم مختلف، نتركه بدون رابط
                    pass

    return render_template(
        "requests_inbox.html",
        rows=rows,
        status=status,
        q=qterm,
        request_new_url=request_new_url,
        employee_summary_url_map=employee_summary_url_map,
    )

@app.route("/admin/requests/<int:req_id>/decide", methods=["POST"])
@login_required
@admin_required
def request_decide(req_id):
    req = Request.query.get_or_404(req_id)
    decision = request.form.get("decision")  # "approve" أو "reject"
    comment  = (request.form.get("admin_comment") or "").strip()

    if decision not in ("approve", "reject"):
        flash("Invalid decision.", "danger")
        return redirect(url_for("requests_inbox"))

    # حدّث بيانات الطلب
    req.status = "approved" if decision == "approve" else "rejected"
    req.decided_by = cur_user().id
    req.decided_at = datetime.now(timezone.utc)
    req.admin_comment = comment

    if decision == "approve":
        # إجازة → تسجيل حضور + توليد PDF
        if req.type == "leave":
            _apply_leave_attendance_for_request(req)
            existing_form = LeaveForm.query.filter_by(request_id=req.id).first()
            if not existing_form:
                try:
                    pdf_name = generate_leave_pdf(req)
                    if pdf_name:
                        db.session.add(LeaveForm(request_id=req.id, filename=pdf_name))
                except Exception as e:
                    app.logger.error("leave PDF error: %s", e)

        # سكليف → تسجيل غياب تلقائي
        if req.type == "sick":
            _apply_sick_attendance_for_request(req)

        # إنشاء مهمة HR لجميع الأنواع
        existing = HRTask.query.filter_by(request_id=req.id).first()
        if not existing:
            try:
                db.session.add(HRTask(
                    request_id=req.id,
                    employee_id=req.employee_id,
                    type=req.type or "leave",
                    status="pending",
                    company_id=req.company_id,
                ))
            except IntegrityError:
                db.session.rollback()

    db.session.commit()
    flash("Request updated.", "success")
    return redirect(url_for("requests_inbox"))


@app.route("/requests/<int:req_id>/leave-sign", methods=["GET", "POST"])
@login_required
def web_leave_sign(req_id):
    req = Request.query.get_or_404(req_id)
    lf  = LeaveForm.query.filter_by(request_id=req_id).first()
    sig = LeaveSignature.query.filter_by(request_id=req_id).first()

    if request.method == "POST":
        signature  = (request.form.get("signature") or "").strip()
        emp_name   = (request.form.get("employee_name") or (req.employee.name if req.employee else "")).strip()
        if not signature:
            flash("التوقيع مطلوب", "danger")
            return redirect(url_for("web_leave_sign", req_id=req_id))

        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("Asia/Riyadh"))
        if sig:
            sig.signature = signature; sig.employee_name = emp_name; sig.signed_at = now
        else:
            sig = LeaveSignature(request_id=req_id, signature=signature,
                                 employee_name=emp_name, signed_at=now)
            db.session.add(sig)

        if lf:
            lf.signed = True; lf.signed_at = now
            try:
                pdf_name = generate_leave_pdf(req, sig)
                if pdf_name:
                    lf.filename = pdf_name
            except Exception as e:
                app.logger.error("leave sign PDF: %s", e)
        db.session.commit()
        flash("تم التوقيع بنجاح ✓", "success")
        return redirect(url_for("web_leave_sign", req_id=req_id))

    return render_template("leave_sign.html", req=req, lf=lf, sig=sig)


@app.route("/requests/<int:req_id>/permission-sign", methods=["GET", "POST"])
@login_required
def web_permission_sign(req_id):
    req  = Request.query.get_or_404(req_id)
    sig  = PermissionSignature.query.filter_by(request_id=req_id).first()
    task = HRTask.query.filter_by(request_id=req_id).first()

    if request.method == "POST":
        signature = (request.form.get("signature") or "").strip()
        emp_name  = (request.form.get("employee_name") or (req.employee.name if req.employee else "")).strip()
        if not signature:
            flash("التوقيع مطلوب", "danger")
            return redirect(url_for("web_permission_sign", req_id=req_id))

        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("Asia/Riyadh"))
        if sig:
            sig.signature = signature; sig.employee_name = emp_name; sig.signed_at = now
        else:
            sig = PermissionSignature(request_id=req_id, signature=signature,
                                      employee_name=emp_name, signed_at=now)
            db.session.add(sig)
        db.session.flush()

        try:
            pdf_name = generate_permission_pdf(req, sig)
            if pdf_name and task:
                task.warning_pdf = pdf_name
        except Exception as e:
            app.logger.error("permission sign PDF: %s", e)
        db.session.commit()
        flash("تم التوقيع بنجاح ✓", "success")
        return redirect(url_for("web_permission_sign", req_id=req_id))

    return render_template("permission_sign.html", req=req, sig=sig, task=task)


@app.get("/requests/<int:req_id>/leave-pdf")
@login_required
def web_leave_pdf(req_id):
    lf = LeaveForm.query.filter_by(request_id=req_id).first_or_404()
    return send_file(os.path.join(UPLOAD_FOLDER, lf.filename),
                     mimetype="application/pdf", as_attachment=False,
                     download_name=f"leave_form_{req_id}.pdf")


@app.route("/requests/<int:req_id>/warning-sign", methods=["GET", "POST"])
@login_required
def web_warning_sign(req_id):
    req  = Request.query.get_or_404(req_id)
    task = HRTask.query.filter_by(request_id=req_id).first()
    sig  = WarningSignature.query.filter_by(request_id=req_id).first()

    if request.method == "POST":
        signature = (request.form.get("signature") or "").strip()
        emp_name  = (request.form.get("employee_name") or (req.employee.name if req.employee else "")).strip()
        if not signature:
            flash("التوقيع مطلوب", "danger")
            return redirect(url_for("web_warning_sign", req_id=req_id))
        from zoneinfo import ZoneInfo
        now = datetime.now(ZoneInfo("Asia/Riyadh"))
        if sig:
            sig.signature = signature; sig.employee_name = emp_name; sig.signed_at = now
        else:
            sig = WarningSignature(request_id=req_id, signature=signature,
                                   employee_name=emp_name, signed_at=now)
            db.session.add(sig)
        # توليد PDF الإنذار
        try:
            pdf_name = generate_warning_pdf(req, sig)
            if pdf_name and task:
                task.warning_pdf = pdf_name
        except Exception as e:
            app.logger.error("warning sign PDF: %s", e)
        db.session.commit()
        flash("تم التوقيع بنجاح ✓", "success")
        return redirect(url_for("web_warning_sign", req_id=req_id))

    return render_template("warning_sign.html", req=req, task=task, sig=sig)


@app.get("/requests/<int:req_id>/warning-pdf")
@login_required
def web_warning_pdf(req_id):
    task = HRTask.query.filter_by(request_id=req_id).first_or_404()
    if not task.warning_pdf:
        abort(404)
    return send_file(os.path.join(UPLOAD_FOLDER, task.warning_pdf),
                     mimetype="application/pdf", as_attachment=False,
                     download_name=f"warning_{req_id}.pdf")


@app.get("/requests/attachment/<int:att_id>")
@login_required
def request_attachment_download(att_id):
    att = RequestAttachment.query.get_or_404(att_id)
    ext = att.filename.rsplit(".", 1)[-1].lower() if "." in att.filename else "pdf"
    mime = "application/pdf" if ext == "pdf" else f"image/{ext}"
    return send_file(os.path.join(UPLOAD_FOLDER, att.filename),
                     mimetype=mime, as_attachment=False,
                     download_name=att.original_name or att.filename)


@app.route("/attendance/report", methods=["GET"])
@admin_required
def attendance_report():
    # نطاق التاريخ (اختياري عبر ?from=YYYY-MM-DD&to=YYYY-MM-DD)
    frm = request.args.get("from"); to = request.args.get("to")
    try:
        d_from = parse_date(frm) if frm else date.today().replace(day=1)
    except: d_from = date.today().replace(day=1)
    try:
        d_to = parse_date(to) if to else date.today()
    except: d_to = date.today()

    employees = (db.session.query(Employee, User)
                 .join(User, Employee.user_id == User.id)
                 .filter(Employee.is_active == True)
                 .order_by(User.supervisor_code.asc(), Employee.name.asc())
                 .all())

    rows = []
    for emp, sup in employees:
        q = Attendance.query.filter(Attendance.employee_id==emp.id,
                                    Attendance.date>=d_from,
                                    Attendance.date<=d_to)
        total = q.count()
        present = q.filter_by(status="present").count()
        leave   = q.filter_by(status="leave").count()
        absent  = q.filter_by(status="absent").count()
        pct = (present / total * 100.0) if total else 0.0
        rows.append({
            "emp": emp, "sup": sup,
            "total": total, "present": present,
            "leave": leave, "absent": absent,
            "percent": round(pct,1)
        })

    # KPI إجمالي
    total_days = sum(r["total"] for r in rows)
    total_present = sum(r["present"] for r in rows)
    kpi_attendance = round((total_present/total_days*100.0),1) if total_days else 0.0

    return render_template("attendance_report.html",
                           rows=rows, d_from=d_from, d_to=d_to,
                           kpi_attendance=kpi_attendance)

@app.route("/attendance/print", methods=["GET"])
@admin_required
def attendance_print():
    mode = request.args.get("mode") or "day"

    if mode == "week":
        ws = _safe_date(request.args.get("ws")) or default_week_today()[0]
        we = _safe_date(request.args.get("we")) or (ws + timedelta(days=4))
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            ws, we = default_week_today()
    else:
        d = _safe_date(request.args.get("d")) or date.today()

    employees = (db.session.query(Employee, User)
                 .join(User, Employee.user_id == User.id)
                 .filter(Employee.is_active == True)
                 .order_by(User.supervisor_code.asc(), Employee.name.asc())
                 .all())

    if mode == "day":
        atts = Attendance.query.filter_by(date=d).all()
        att_by_emp = {a.employee_id: a for a in atts}
        rows = []
        for emp, sup in employees:
            a = att_by_emp.get(emp.id)
            rows.append({
                "emp": emp, "sup": sup,
                "status": (a.status if a else None),
                "remarks": (a.remarks if a else ""),
                "marked_by": (db.session.get(User, a.supervisor_id) if a else None)
            })
        return render_template("attendance_print.html", mode="day", day_date=d, rows=rows)

    else:
        emp_ids = [emp.id for emp, _ in employees]
        atts = []
        if emp_ids:
            atts = (Attendance.query
                    .filter(Attendance.employee_id.in_(emp_ids),
                            Attendance.date >= ws,
                            Attendance.date <= we).all())
        from collections import defaultdict
        stats = defaultdict(lambda: {"present": 0, "absent": 0, "leave": 0, "total": 0})
        for a in atts:
            stats[a.employee_id]["total"] += 1
            if a.status in ("present", "absent", "leave"):
                stats[a.employee_id][a.status] += 1
        rows = []
        for emp, sup in employees:
            s = stats[emp.id]
            pct = (s["present"]/s["total"]*100.0) if s["total"] else 0.0
            rows.append({
                "emp": emp, "sup": sup,
                "present": s["present"], "absent": s["absent"], "leave": s["leave"],
                "total": s["total"], "percent": round(pct,1)
            })
        return render_template("attendance_print.html", mode="week", ws=ws, we=we, rows=rows)

@app.route("/attendance/pdf", methods=["GET"])
@admin_required
def attendance_pdf():
    frm = request.args.get("from"); to = request.args.get("to")
    try:
        d_from = parse_date(frm) if frm else date.today().replace(day=1)
    except: d_from = date.today().replace(day=1)
    try:
        d_to = parse_date(to) if to else date.today()
    except: d_to = date.today()

    employees = (db.session.query(Employee, User)
                 .join(User, Employee.user_id == User.id)
                 .filter(Employee.is_active == True)
                 .order_by(User.supervisor_code.asc(), Employee.name.asc())
                 .all())

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, f"Attendance Report {d_from} to {d_to}", ln=True, align="C")

    pdf.set_font("Arial", "B", 10)
    pdf.cell(60, 8, "Employee", 1)
    pdf.cell(35, 8, "Supervisor ID", 1)
    pdf.cell(20, 8, "Total", 1)
    pdf.cell(25, 8, "Present", 1)
    pdf.cell(25, 8, "Leave", 1)
    pdf.cell(25, 8, "Absent", 1)
    pdf.cell(25, 8, "Attendance %", 1)
    pdf.ln()

    pdf.set_font("Arial", "", 10)
    for emp, sup in employees:
        q = Attendance.query.filter(Attendance.employee_id==emp.id,
                                    Attendance.date>=d_from,
                                    Attendance.date<=d_to)
        total = q.count()
        present = q.filter_by(status="present").count()
        leave   = q.filter_by(status="leave").count()
        absent  = q.filter_by(status="absent").count()
        pct = (present/total*100.0) if total else 0.0

        pdf.cell(60, 8, emp.name[:28], 1)
        pdf.cell(35, 8, (sup.supervisor_code or "")[:12], 1)
        pdf.cell(20, 8, str(total), 1, align="C")
        pdf.cell(25, 8, str(present), 1, align="C")
        pdf.cell(25, 8, str(leave), 1, align="C")
        pdf.cell(25, 8, str(absent), 1, align="C")
        pdf.cell(25, 8, f"{pct:.1f}%", 1, align="C")
        pdf.ln()

    return Response(pdf.output(dest="S").encode("latin1"),
                    mimetype="application/pdf",
                    headers={"Content-Disposition":"inline; filename=attendance.pdf"})

@app.route("/attendance/mark", methods=["GET", "POST"])
@login_required
def attendance_mark():
    u = cur_user()
    if not u or u.role not in ("supervisor", "admin"):
        flash("Unauthorized", "danger")
        return redirect(url_for("index"))

    # الموظفين للمشرف الحالي
    employees = (Employee.query
                 .filter_by(user_id=u.id, is_active=True)
                 .order_by(Employee.name.asc())
                 .all())

    # هذا اللي نعرضه في الفورم
    today = date.today()

    if request.method == "POST":
        # اقرأ التاريخ اللي جاي من الفورم (name="date" في التمبليت)
        d_raw = request.form.get("date")
        if d_raw:
            # نحلّلها يدوي عشان ما يصير خطأ 500 لو السيرفر قديم
            try:
                y, m, d = map(int, d_raw.split("-"))
                mark_date = date(y, m, d)
            except ValueError:
                mark_date = today
        else:
            mark_date = today

        changed = 0
        for emp in employees:
            status = (request.form.get(f"emp_{emp.id}_status") or "").strip()
            remarks = (request.form.get(f"emp_{emp.id}_remarks") or "").strip()
            if not status:
                continue

            # استخدم التاريخ اللي اختاره المستخدم
            rec = Attendance.query.filter_by(employee_id=emp.id, date=mark_date).first()
            if not rec:
                rec = Attendance(
                    employee_id=emp.id,
                    supervisor_id=u.id,
                    date=mark_date,
                    status=status,
                    remarks=remarks,
                    company_id=cid(),
                )
                db.session.add(rec)
                changed += 1
            else:
                if rec.status != status or rec.remarks != remarks:
                    rec.status = status
                    rec.remarks = remarks
                    changed += 1

        if changed:
            db.session.commit()
            flash("Attendance saved.", "success")
        else:
            flash("No changes.", "info")

        # نرجع لنفس الصفحة زي أول
        return redirect(url_for("attendance_mark"))

    # GET: اعرض حضور اليوم
    saved = {a.employee_id: a for a in Attendance.query.filter_by(date=today).all()}
    return render_template("attendance_mark.html",
                           employees=employees,
                           today=today,
                           saved=saved)


# ─────────────────────────────────────────────
#  صفحة الدعم — Apple App Store requirement
#  https://nsh.z-app.biz/support
# ─────────────────────────────────────────────
@app.route("/support")
def support_page():
    path = os.path.join(BASE_DIR, "static", "support.html")
    with open(path, "r", encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "text/html; charset=utf-8"}


@app.route("/")
def index():
    u = cur_user()
    if not u:
        return render_template("landing.html")
    role = u.role
    if role == "super_admin":
        return redirect(url_for("superadmin_dashboard"))
    if role == "admin":
        return redirect(url_for("admin_kpi"))
    if role == "site_supervisor":
        return redirect(url_for("site_supervisors"))
    if role == "hr":
        return redirect(url_for("hr_inbox"))
    if role == "safety_officer":
        return redirect(url_for("hse_checkin"))
    return redirect(url_for("employees"))

# اختيار المشرف والأسبوع لطباعة الحزمة
@app.route("/site/print", methods=["GET", "POST"])
@login_required
def site_print_select():
    u = cur_user()
    if not u or u.role != "site_supervisor":
        abort(403)

    # المشرفين المرتبطين بهذا الـ Site Supervisor
    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id)
             .order_by(User.supervisor_code.asc())
             .all())
    supervisors = [sup for _, sup in links]
    ws, we = default_week_today()

    if request.method == "POST":
        sup_user_id = int(request.form["sup_user_id"])
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("site_print_select"))
        # نستخدم GET حتى تكون قابلة لإعادة الفتح
        return redirect(url_for("site_print_bundle",
                                sup_user_id=sup_user_id,
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    return render_template("site_print_select.html", supervisors=supervisors, ws=ws, we=we)

@app.route("/admin/kpi", methods=["GET", "POST"])
@admin_required
def admin_kpi():
    if request.method == "POST":
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("admin_kpi"))
        return redirect(url_for("admin_kpi", week_start=ws.isoformat(), week_end=we.isoformat()))

    # أسبوع افتراضي (أو من الاستعلام)
    ws, we = default_week_today()
    if request.args.get("week_start") and request.args.get("week_end"):
        ws = parse_date(request.args["week_start"])
        we = parse_date(request.args["week_end"])
    pws, pwe = previous_week_range(ws)

    # ---------- Supervisor → Employees ----------
    supervisors = (
        apply_company_filter(User.query.filter_by(role="supervisor", is_active=True), User)
        .order_by(User.supervisor_code.asc()).all()
    )

    emp_totals_map = dict(
        apply_company_filter(
            db.session.query(Employee.user_id, func.count(Employee.id))
            .filter(Employee.is_active == True), Employee)
        .group_by(Employee.user_id).all()
    )

    emp_cnt_this = dict(
        db.session.query(Employee.user_id, func.count(Evaluation.id))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
        .group_by(Employee.user_id).all()
    )

    emp_cnt_prev = dict(
        db.session.query(Employee.user_id, func.count(Evaluation.id))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == pws, Evaluation.week_end == pwe)
        .group_by(Employee.user_id).all()
    )

    # avg score per supervisor this week
    avg_score_map = dict(
        db.session.query(Employee.user_id, func.avg(Evaluation.total_score))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
        .group_by(Employee.user_id).all()
    )

    sup_rows, sum_target_emp, sum_done_emp = [], 0, 0
    growth_pool_prev, growth_pool_curr = 0, 0

    for sup in supervisors:
        target = emp_totals_map.get(sup.id, 0)
        done_c = emp_cnt_this.get(sup.id, 0)
        done_p = emp_cnt_prev.get(sup.id, 0)
        coverage = (done_c / target * 100.0) if target else None
        prev_coverage = (done_p / target * 100.0) if target else None

        sum_target_emp += target
        sum_done_emp += done_c

        if done_c > 0 and done_p > 0:
            growth_pool_prev += done_p
            growth_pool_curr += done_c

        avg_sc = avg_score_map.get(sup.id)
        sup_rows.append({
            "id": sup.id,
            "code": sup.supervisor_code,
            "name": sup.name,
            "target": target,
            "done_c": done_c,
            "done_p": done_p,
            "coverage": coverage,
            "prev_coverage": prev_coverage,
            "avg_score": round(float(avg_sc), 1) if avg_sc else None,
            "delta": (done_c - done_p),
            "trend": ("up" if done_c > done_p else "down" if done_c < done_p else "flat"),
        })

    # ---------- Site Supervisor → Supervisors ----------
    site_sups = (
        apply_company_filter(User.query.filter_by(role="site_supervisor", is_active=True), User)
        .order_by(User.supervisor_code.asc()).all()
    )

    assigned_counts = dict(
        db.session.query(
            SiteSupervisorMap.site_sup_id,
            func.count(func.distinct(SiteSupervisorMap.supervisor_id)),
        )
        .group_by(SiteSupervisorMap.site_sup_id).all()
    )

    se_cnt_this = dict(
        db.session.query(SupervisorEvaluation.evaluator_id, func.count(SupervisorEvaluation.id))
        .filter(SupervisorEvaluation.week_start == ws, SupervisorEvaluation.week_end == we)
        .group_by(SupervisorEvaluation.evaluator_id).all()
    )

    se_cnt_prev = dict(
        db.session.query(SupervisorEvaluation.evaluator_id, func.count(SupervisorEvaluation.id))
        .filter(SupervisorEvaluation.week_start == pws, SupervisorEvaluation.week_end == pwe)
        .group_by(SupervisorEvaluation.evaluator_id).all()
    )

    site_rows, sum_target_sup, sum_done_sup = [], 0, 0
    for s in site_sups:
        target = assigned_counts.get(s.id, 0)
        done_c = se_cnt_this.get(s.id, 0)
        done_p = se_cnt_prev.get(s.id, 0)
        coverage = (done_c / target * 100.0) if target else None
        prev_coverage = (done_p / target * 100.0) if target else None

        sum_target_sup += target
        sum_done_sup += done_c

        if done_c > 0 and done_p > 0:
            growth_pool_prev += done_p
            growth_pool_curr += done_c

        site_rows.append({
            "id": s.id,
            "code": s.supervisor_code,
            "name": s.name,
            "target": target,
            "done_c": done_c,
            "done_p": done_p,
            "coverage": coverage,
            "prev_coverage": prev_coverage,
            "delta": (done_c - done_p),
            "trend": ("up" if done_c > done_p else "down" if done_c < done_p else "flat"),
        })

    # ---------- إجماليات ----------
    _ev_this = apply_company_filter(
        db.session.query(func.count(Evaluation.id))
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we), Evaluation)
    _se_this = apply_company_filter(
        db.session.query(func.count(SupervisorEvaluation.id))
        .filter(SupervisorEvaluation.week_start == ws, SupervisorEvaluation.week_end == we), SupervisorEvaluation)
    total_this_week = (_ev_this.scalar() or 0) + (_se_this.scalar() or 0)

    _ev_prev = apply_company_filter(
        db.session.query(func.count(Evaluation.id))
        .filter(Evaluation.week_start == pws, Evaluation.week_end == pwe), Evaluation)
    _se_prev = apply_company_filter(
        db.session.query(func.count(SupervisorEvaluation.id))
        .filter(SupervisorEvaluation.week_start == pws, SupervisorEvaluation.week_end == pwe), SupervisorEvaluation)
    total_prev_week = (_ev_prev.scalar() or 0) + (_se_prev.scalar() or 0)

    growth_pct = 0.0
    if growth_pool_prev > 0:
        growth_pct = (growth_pool_curr - growth_pool_prev) / growth_pool_prev * 100.0

    overall_emp_coverage = (sum_done_emp / sum_target_emp * 100.0) if sum_target_emp else None
    overall_sup_coverage = (sum_done_sup / sum_target_sup * 100.0) if sum_target_sup else None

    # ---------- Attendance & Leave Rates (من جميع الفرص) ----------
    # الموظفون النشطون + عدد أيام الأسبوع (Sun→Thu بعد التحقق)
    active_emp_ids = [e.id for e in apply_company_filter(Employee.query.filter_by(is_active=True), Employee).all()]
    active_count = len(active_emp_ids)
    workdays = (we - ws).days + 1  # يفترض Sun..Thu بعد validate_week_sun_to_thu

    total_opportunities = active_count * workdays  # المقام

    # تعداد الحالات المسجلة
    present_count = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date >= ws,
        Attendance.date <= we,
        Attendance.employee_id.in_(active_emp_ids),
        Attendance.status == 'present'
    ).scalar() or 0

    leave_count = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date >= ws,
        Attendance.date <= we,
        Attendance.employee_id.in_(active_emp_ids),
        Attendance.status == 'leave'
    ).scalar() or 0

    # الغياب = الباقي (يشمل الأيام/الموظفين غير المسجلين)
    absent_count = max(total_opportunities - (present_count + leave_count), 0)

    att_rate   = round(present_count / total_opportunities * 100.0, 1) if total_opportunities else None
    leave_rate = round(leave_count   / total_opportunities * 100.0, 1) if total_opportunities else None
    # (اختياري) احسب غياب:
    # absent_rate = round(absent_count / total_opportunities * 100.0, 1) if total_opportunities else None

    # band distribution this week
    band_rows = (db.session.query(Evaluation.overall_band, func.count(Evaluation.id))
                 .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
                 .group_by(Evaluation.overall_band).all())
    band_counts = {b: c for b, c in band_rows}

    # overall avg score this week
    avg_score_val = (db.session.query(func.avg(Evaluation.total_score))
                     .filter(Evaluation.week_start == ws, Evaluation.week_end == we).scalar())
    avg_score = round(float(avg_score_val), 1) if avg_score_val else None

    return render_template(
        "admin_kpi.html",
        ws=ws, we=we, pws=pws, pwe=pwe,
        total_this_week=total_this_week, total_prev_week=total_prev_week,
        growth_pct=growth_pct,
        overall_emp_cov=overall_emp_coverage,
        overall_sup_cov=overall_sup_coverage,
        att_rate=att_rate,
        leave_rate=leave_rate,
        avg_score=avg_score,
        band_counts=band_counts,
        sup_rows=sup_rows, site_rows=site_rows
    )


@app.route("/admin/supervisor/<int:sup_id>/history")
@admin_required
def admin_supervisor_history(sup_id):
    sup = User.query.filter_by(id=sup_id, role="supervisor").first_or_404()
    evals = (SupervisorEvaluation.query
             .filter_by(supervisor_id=sup.id)
             .order_by(SupervisorEvaluation.week_start.desc())
             .all())
    return render_template("supervisor_history.html", sup=sup, evals=evals)
@app.get("/_health")
def _health():
    return "ok"

# عرض الحزمة الجاهزة للطباعة
@app.get("/site/print/bundle")
@login_required
def site_print_bundle():
    u = cur_user()
    if not u or u.role != "site_supervisor":
        abort(403)

    sup_user_id = int(request.args.get("sup_user_id", "0"))
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")
    if not (sup_user_id and week_start and week_end):
        return redirect(url_for("site_print_select"))

    ws = parse_date(week_start); we = parse_date(week_end)

    # تأكد أن هذا المشرف ضمن قوائم هذا الـ Site Supervisor
    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_user_id).first()
    if not link:
        abort(403)

    supervisor = db.session.get(User, sup_user_id) or abort(404)

    # جميع موظفي هذا المشرف
    employees = Employee.query.filter_by(user_id=sup_user_id, is_active=True).order_by(Employee.name.asc()).all()
    emp_ids = [e.id for e in employees]

    # كل التقييمات لهؤلاء الموظفين في الأسبوع المحدد
    eval_rows = (db.session.query(Evaluation, Employee)
                 .join(Employee, Evaluation.employee_id == Employee.id)
                 .filter(Evaluation.week_start == ws,
                         Evaluation.week_end == we,
                         Employee.id.in_(emp_ids))
                 .order_by(Employee.name.asc())
                 .all())

    # جهّز قائمة مطبوعات: (ev, emp, evaluator_user)
    evals = []
    seen_emp = set()
    for ev, emp in eval_rows:
        evaluator = db.session.get(User, ev.evaluator_id)
        evals.append((ev, emp, evaluator))
        seen_emp.add(emp.id)

    # من لم يتم تقييمهم
    not_evaluated = [emp for emp in employees if emp.id not in seen_emp]

    return render_template(
        "site_print_bundle.html",
        supervisor=supervisor, ws=ws, we=we,
        evals=evals, not_evaluated=not_evaluated
    )

@app.get("/hr/inbox")
@hr_required
def hr_inbox():
    rows = apply_company_filter(HRTask.query, HRTask).order_by(HRTask.created_at.desc()).all()
    return render_template("hr_inbox.html", rows=rows)

@app.post("/hr/task/<int:task_id>/apply")
@hr_required
def hr_task_apply(task_id):
    t = db.session.get(HRTask, task_id)
    if not t:
        abort(404)
    if t.status == "pending":
        t.status = "applied"
        t.applied_at = datetime.now(timezone.utc)
        t.applied_by = cur_user().id
        db.session.commit()
    return redirect(url_for("hr_inbox"))


@app.post("/hr/task/<int:task_id>/set-reason")
@hr_required
def hr_task_set_reason(task_id):
    """HR يحدد السبب الرسمي للإنذار من الويب"""
    t = db.session.get(HRTask, task_id)
    if not t or t.type != "warning":
        abort(404)
    reason = (request.form.get("official_reason") or "").strip()
    if not reason:
        flash("السبب الرسمي مطلوب", "danger")
        return redirect(url_for("hr_inbox"))
    try:
        t.official_reason = reason
    except Exception:
        from sqlalchemy import text as _text
        with db.engine.connect() as _conn:
            _conn.execute(_text("ALTER TABLE hr_task ADD COLUMN IF NOT EXISTS official_reason TEXT NULL"))
            _conn.commit()
        t.official_reason = reason
    db.session.commit()
    # إشعار للمشرف
    if t.request:
        threading.Thread(
            target=_send_push_to_user,
            args=(t.request.supervisor_id, "إنذار — جاهز للتوقيع",
                  "تم تحديد السبب الرسمي، يمكنك الآن فتح نموذج الإنذار"),
            daemon=True,
        ).start()
    flash("تم تحديد السبب الرسمي ✓", "success")
    return redirect(url_for("hr_inbox"))

@app.route("/admin")
@admin_required
def admin_home():
    return redirect(url_for("admin_kpi"))

@app.route("/login", methods=["GET", "POST"])
@_login_limit
def login():
    if request.method == "POST":
        mode     = request.form.get("mode", "code")
        password = (request.form.get("password") or "").strip()

        user = None
        if mode == "email":
            email = (request.form.get("email") or "").strip().lower()
            if not email:
                flash("Please enter your email.", "danger")
                return render_template("login.html", login_mode="email")
            user = User.query.filter_by(email=email, is_active=True).first()
            if not user or not user.check_password(password):
                flash("Invalid email or password.", "danger")
                return render_template("login.html", login_mode="email")
        else:
            code = (request.form.get("code") or request.form.get("sup_code") or "").strip()
            if not code:
                flash("Please enter your Supervisor ID.", "danger")
                return redirect(url_for("login"))
            user = User.query.filter_by(supervisor_code=code, is_active=True).first()
            if not user:
                flash("ID not found or inactive.", "danger")
                return redirect(url_for("login"))

        # Check company is active (skip for super_admin)
        if user.role != "super_admin" and user.company_id:
            co = db.session.get(Company, user.company_id)
            if co and not co.is_active:
                flash("Your company account is pending approval.", "warning")
                return redirect(url_for("login"))

        session.clear()
        session["user_id"] = user.id
        # Always store company_id (None is fine for super_admin)
        session["company_id"] = user.company_id

        if user.role == "super_admin":
            return redirect(url_for("superadmin_dashboard"))
        elif user.role == "admin":
            return redirect(url_for("admin_kpi"))
        elif user.role == "site_supervisor":
            return redirect(url_for("site_supervisors"))
        elif user.role == "hr":
            return redirect(url_for("hr_inbox"))
        elif user.role == "safety_officer":
            return redirect(url_for("hse_checkin"))
        else:
            return redirect(url_for("employees"))

    return render_template("login.html")


@app.route("/register/choose")
def register_choose():
    if cur_user():
        return redirect(url_for("index"))
    return render_template("register_choose.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if cur_user():
        return redirect(url_for("index"))

    if request.method == "POST":
        company_name = (request.form.get("company_name") or "").strip()
        your_name    = (request.form.get("your_name") or "").strip()
        email        = (request.form.get("email") or "").strip().lower()
        password     = (request.form.get("password") or "").strip()
        password2    = (request.form.get("password2") or "").strip()

        errors = []
        if not company_name:
            errors.append("Company name is required.")
        if not your_name:
            errors.append("Your name is required.")
        if not email or "@" not in email:
            errors.append("Valid email is required.")
        if len(password) < 6:
            errors.append("Password must be at least 6 characters.")
        if password != password2:
            errors.append("Passwords do not match.")
        if User.query.filter_by(email=email).first():
            errors.append("This email is already registered.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("register.html",
                                   company_name=company_name, your_name=your_name, email=email)

        # Build slug from company name
        import re
        slug = re.sub(r"[^a-z0-9]+", "-", company_name.lower()).strip("-")
        base_slug = slug
        counter = 1
        while Company.query.filter_by(slug=slug).first():
            slug = f"{base_slug}-{counter}"
            counter += 1

        company = Company(name=company_name, slug=slug, owner_email=email, is_active=False)
        db.session.add(company)
        db.session.flush()  # get company.id

        # Generate a unique supervisor_code for the new admin
        import secrets as _sec
        sup_code = f"adm-{_sec.token_hex(4)}"
        while User.query.filter_by(supervisor_code=sup_code).first():
            sup_code = f"adm-{_sec.token_hex(4)}"

        admin_user = User(
            supervisor_code=sup_code,
            name=your_name,
            email=email,
            role="admin",
            is_active=True,
            company_id=company.id,
        )
        admin_user.set_password(password)
        db.session.add(admin_user)
        db.session.commit()

        flash("Registration submitted! Your account is pending approval. You will be notified once approved.", "success")
        return redirect(url_for("login"))

    return render_template("register.html", company_name="", your_name="", email="")


@app.route("/logout")
def logout():
    session.clear()
    flash("Signed out.", "info")
    return redirect(url_for("login"))

# ----- Admin: Users -----
@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    if request.method == "POST":
        code = (request.form.get("code") or "").strip()
        name = (request.form.get("name") or "").strip()
        role = request.form.get("role") or "supervisor"  # can be site_supervisor
        if not code:
            flash("Supervisor ID is required.", "danger")
        else:
            existing = User.query.filter_by(supervisor_code=code).first()
            if existing:
                flash("This ID already exists.", "warning")
            else:
                u = User(supervisor_code=code, name=name, role=role,
                         is_active=True, company_id=cid())
                db.session.add(u)
                db.session.commit()
                flash("User added.", "success")
        return redirect(url_for("admin_users"))
    show_hidden = request.args.get("show_hidden") == "1"
    q = apply_company_filter(User.query, User)
    if not show_hidden:
        q = q.filter(db.or_(User.is_hidden == False, User.is_hidden == None))
    users = q.order_by(User.role.desc(), User.supervisor_code.asc()).all()
    return render_template("admin_users.html", users=users, show_hidden=show_hidden)
# --- Admin: Print hub (select week) ---
@app.route("/admin/print", methods=["GET", "POST"])
@admin_required
def admin_print_select():
    ws, we = default_week_today()
    if request.method == "POST":
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("admin_print_select"))
        return redirect(url_for("admin_print_bundle",
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    return render_template("admin_print_select.html", ws=ws, we=we)


# --- Admin: Print bundle (all reports + blanks for missing) ---
@app.route("/admin/print/bundle")
@admin_required
def admin_print_bundle():
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")
    if not (week_start and week_end):
        return redirect(url_for("admin_print_select"))

    ws = parse_date(week_start); we = parse_date(week_end)

    # كل المشرفين الفعّالين
    supervisors = (
        apply_company_filter(User.query.filter_by(role="supervisor", is_active=True), User)
        .order_by(User.supervisor_code.asc()).all()
    )
    sup_by_id = {s.id: s for s in supervisors}

    # خريطة (المشرف ← أحد مشرفي السايت المربوطين به إن وجد)
    ssm_rows = SiteSupervisorMap.query.all()
    site_by_sup: dict[int, User | None] = {}
    for row in ssm_rows:
        # اختر أول مشرف سايت فقط للعرض
        if row.supervisor_id not in site_by_sup:
            site_by_sup[row.supervisor_id] = db.session.get(User, row.site_sup_id)

    # تقييمات السايت لهذا الأسبوع
    se_list = SupervisorEvaluation.query.filter_by(week_start=ws, week_end=we).all()
    se_by_sup: dict[int, SupervisorEvaluation] = {se.supervisor_id: se for se in se_list}

    # صفحات تقييم السايت الموجودة
    site_eval_pages = []
    for se in se_list:
        sup = sup_by_id.get(se.supervisor_id)
        if not sup:
            continue
        evaluator = db.session.get(User, se.evaluator_id)
        site_eval_pages.append((se, sup, evaluator))

    # صفحات السايت المفقودة (فارغة)
    site_missing_pages = []
    for sup in supervisors:
        if sup.id not in se_by_sup:
            site_eval_pages_sup = site_by_sup.get(sup.id)  # قد يكون None
            site_missing_pages.append((sup, site_eval_pages_sup))

    # الموظفون لكل مشرف
    employees = Employee.query.filter(Employee.user_id.in_([s.id for s in supervisors]),
                                      Employee.is_active == True)\
                              .order_by(Employee.name.asc()).all()
    emp_ids = [e.id for e in employees]
    emp_by_id = {e.id: e for e in employees}
    sup_id_by_emp_id = {e.id: e.user_id for e in employees}

    # كل تقييمات الموظفين لهذا الأسبوع
    ev_list = Evaluation.query.filter(Evaluation.week_start == ws,
                                      Evaluation.week_end == we,
                                      Evaluation.employee_id.in_(emp_ids)).all()
    ev_by_emp: dict[int, Evaluation] = {ev.employee_id: ev for ev in ev_list}

    # صفحات تقييم الموظفين الموجودة
    emp_eval_pages = []
    for ev in ev_list:
        emp = emp_by_id.get(ev.employee_id)
        if not emp:
            continue
        sup = sup_by_id.get(emp.user_id)
        evaluator = db.session.get(User, ev.evaluator_id)
        emp_eval_pages.append((ev, emp, sup, evaluator))

    # الموظفون غير المُقيّمين
    emp_missing_pages = []
    for emp in employees:
        if emp.id not in ev_by_emp:
            sup = sup_by_id.get(emp.user_id)
            emp_missing_pages.append((emp, sup))

    return render_template(
        "admin_print_bundle.html",
        ws=ws, we=we,
        site_eval_pages=site_eval_pages,
        site_missing_pages=site_missing_pages,
        emp_eval_pages=emp_eval_pages,
        emp_missing_pages=emp_missing_pages
    )

@app.post("/admin/users/<int:user_id>/toggle")
@admin_required
def admin_users_toggle(user_id):
    u = db.session.get(User, user_id) or abort(404)
    if u.role == "admin":
        flash("Cannot deactivate admin.", "warning")
    else:
        u.is_active = not u.is_active
        db.session.commit()
        flash("Status updated.", "success")
    return redirect(url_for("admin_users"))


@app.post("/admin/users/<int:user_id>/hide")
@admin_required
def admin_users_hide(user_id):
    u = db.session.get(User, user_id) or abort(404)
    if u.role == "admin":
        flash("Cannot hide admin.", "warning")
    else:
        current = getattr(u, "is_hidden", False) or False
        u.is_hidden = not current
        db.session.commit()
        flash("تم الإخفاء." if u.is_hidden else "تم الإظهار.", "success")
    return redirect(url_for("admin_users"))

@app.post("/admin/users/<int:user_id>/role")
@admin_required
def admin_users_set_role(user_id):
    u = db.session.get(User, user_id) or abort(404)
    new_role = (request.form.get("role") or "").strip()

    # لا نسمح بتعديل دور الأدمن من هنا
    if u.role == "admin":
        flash("Cannot change role of admin here.", "warning")
        return redirect(url_for("admin_users"))

    if new_role not in ["supervisor", "site_supervisor"]:
        flash("Invalid role.", "danger")
    else:
        u.role = new_role
        db.session.commit()
        flash("Role updated.", "success")

    return redirect(url_for("admin_users"))



# ----- Supervisor: Employees -----
@app.route("/employees", methods=["GET"])
@login_required
def employees():
    u = cur_user()
    if u.role not in ("supervisor", "admin"):
        abort(403)

    emps = (Employee.query
            .filter_by(user_id=u.id, is_active=True)
            .order_by(Employee.name)
            .all())
    return render_template("employees.html", user=u, employees=emps)



# ----- New Evaluation (employee) -----
@app.route("/evaluate/<int:emp_id>/new", methods=["GET", "POST"])
@login_required
def evaluate_new(emp_id):
    u = cur_user()

    if not u:
        abort(403)

    # admin و site_supervisor يشوفون أي موظف، supervisor العادي موظفيه فقط
    if u.role in ("admin", "site_supervisor"):
        emp = Employee.query.get_or_404(emp_id)
    else:
        emp = Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404()

    # POST = حفظ (إنشاء أو تعديل)
    if request.method == "POST":
        week_start_raw = request.form.get("week_start")
        week_end_raw   = request.form.get("week_end")

        if not week_start_raw or not week_end_raw:
            flash("Week dates are required.", "danger")
            return redirect(url_for("evaluate_new", emp_id=emp.id))

        ws = parse_date(week_start_raw)
        we = parse_date(week_end_raw)

        # تحقّق أن الأسبوع من الأحد إلى الخميس (نفس ما كنت تستخدم)
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("evaluate_new", emp_id=emp.id))

        # 🔴 هنا الفكرة المهمّة:
        # إذا في تقييم لنفس الموظف ونفس الأسبوع → استخدمه وحدثه
        # إذا مافي → أنشئ واحد جديد
        ev = (Evaluation.query
              .filter_by(employee_id=emp.id, week_start=ws, week_end=we)
              .first())

        if ev is None:
            ev = Evaluation(
                employee_id=emp.id,
                evaluator_id=u.id,
                week_start=ws,
                week_end=we,
                company_id=cid(),
            )
            db.session.add(ev)
        else:
            # لو حاب تعتبر أن آخر من عدّل هو المقيم الحالي
            ev.evaluator_id = u.id

        # 1) Weekly Targets
        ev.t1_text     = request.form.get("t1_text") or ""
        ev.t2_text     = request.form.get("t2_text") or ""
        ev.t3_text     = request.form.get("t3_text") or ""
        ev.t4_text     = request.form.get("t4_text") or ""

        ev.t1_percent  = request.form.get("t1_percent", type=float)
        ev.t2_percent  = request.form.get("t2_percent", type=float)
        ev.t3_percent  = request.form.get("t3_percent", type=float)
        ev.t4_percent  = request.form.get("t4_percent", type=float)

        ev.t1_remarks  = request.form.get("t1_remarks") or ""
        ev.t2_remarks  = request.form.get("t2_remarks") or ""
        ev.t3_remarks  = request.form.get("t3_remarks") or ""
        ev.t4_remarks  = request.form.get("t4_remarks") or ""

        # 2) Performance ratings
        ev.p_punctuality    = request.form.get("p_punctuality", type=int)
        ev.p_quality        = request.form.get("p_quality", type=int)
        ev.p_productivity   = request.form.get("p_productivity", type=int)
        ev.p_communication  = request.form.get("p_communication", type=int)
        ev.p_problemsolving = request.form.get("p_problemsolving", type=int)
        ev.p_compliance     = request.form.get("p_compliance", type=int)

        # 3) Performance comments
        ev.c_punctuality    = request.form.get("c_punctuality") or ""
        ev.c_quality        = request.form.get("c_quality") or ""
        ev.c_productivity   = request.form.get("c_productivity") or ""
        ev.c_communication  = request.form.get("c_communication") or ""
        ev.c_problemsolving = request.form.get("c_problemsolving") or ""
        ev.c_compliance     = request.form.get("c_compliance") or ""

        # 4) Summary
        ev.strengths        = request.form.get("strengths") or ""
        ev.improvements     = request.form.get("improvements") or ""
        ev.training_needed  = request.form.get("training_needed") or ""

        # إعادة حساب الدرجات (نفس الفنكشن اللي تستخدمه أصلًا)
        compute_scores(ev)

        db.session.commit()
        flash("Weekly evaluation saved.", "success")

        # بعد الحفظ، افتح تقرير الموظف لهذا الأسبوع
        return redirect(url_for(
            "report_employee",
            emp_id=emp.id,
            week_start=ws.isoformat(),
            week_end=we.isoformat()
        ))

    # GET = فتح النموذج لأول مرة
    # نفس منطقك القديم: ws / we يتم حسابها أو أخذها من الكويري
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")

    if week_start and week_end:
        ws = parse_date(week_start)
        we = parse_date(week_end)
    else:
        # لو ما جا شيء، استخدم الأسبوع الحالي (عدّلها لو عندك دالة خاصة)
        today = date.today()
        # مثال بسيط: نخلي ws = الأحد، we = الخميس لنفس الأسبوع
        # weekday(): الاثنين=0 ... الأحد=6
        offset_to_sun = (today.weekday() + 1) % 7  # يخلي الأحد = 0
        ws = today - timedelta(days=offset_to_sun)
        we = ws + timedelta(days=4)

    return render_template(
        "eval_form.html",
        employee=emp,
        ws=ws,
        we=we,
        weights=WEIGHTS,
        ev=None  # مهم: وضع إنشاء جديد
    )


@app.route("/evaluate/<int:ev_id>/edit", methods=["GET", "POST"])
@login_required
def evaluate_edit(ev_id):
    u = cur_user()
    ev = Evaluation.query.get_or_404(ev_id)
    emp = db.session.get(Employee, ev.employee_id)

    if not emp:
        abort(404)

    # نفس منطق الصلاحيات في report_employee
    if u.role != "admin" and emp.user_id != u.id:
        abort(403)

    if request.method == "POST":
        try:
            week_start = parse_date(request.form.get("week_start"))
            week_end   = parse_date(request.form.get("week_end"))

            ok, msg = validate_week_sun_to_thu(week_start, week_end)
            if not ok:
                flash(msg, "danger")
                return redirect(url_for("evaluate_edit", ev_id=ev.id))

            # منع تكرار نفس الأسبوع لنفس الموظف (ما عدا هذا السجل)
            dup = (Evaluation.query
                   .filter(
                       Evaluation.employee_id == emp.id,
                       Evaluation.week_start == week_start,
                       Evaluation.week_end == week_end,
                       Evaluation.id != ev.id
                   )
                   .first())
            if dup:
                flash("An evaluation for this week already exists.", "warning")
                return redirect(url_for(
                    "report_employee",
                    emp_id=emp.id,
                    week_start=week_start.isoformat(),
                    week_end=week_end.isoformat()
                ))

            # تحديث الحقول
            ev.week_start = week_start
            ev.week_end   = week_end

            # 1) Targets
            ev.t1_text     = request.form.get("t1_text") or ""
            ev.t2_text     = request.form.get("t2_text") or ""
            ev.t3_text     = request.form.get("t3_text") or ""
            ev.t4_text     = request.form.get("t4_text") or ""
            ev.t1_percent  = request.form.get("t1_percent", type=float)
            ev.t2_percent  = request.form.get("t2_percent", type=float)
            ev.t3_percent  = request.form.get("t3_percent", type=float)
            ev.t4_percent  = request.form.get("t4_percent", type=float)
            ev.t1_remarks  = request.form.get("t1_remarks") or ""
            ev.t2_remarks  = request.form.get("t2_remarks") or ""
            ev.t3_remarks  = request.form.get("t3_remarks") or ""
            ev.t4_remarks  = request.form.get("t4_remarks") or ""

            # 2) Performance ratings
            ev.p_punctuality   = request.form.get("p_punctuality", type=int)
            ev.p_quality       = request.form.get("p_quality", type=int)
            ev.p_productivity  = request.form.get("p_productivity", type=int)
            ev.p_communication = request.form.get("p_communication", type=int)
            ev.p_problemsolving = request.form.get("p_problemsolving", type=int)
            ev.p_compliance    = request.form.get("p_compliance", type=int)

            # 2) Performance comments
            ev.c_punctuality   = request.form.get("c_punctuality") or ""
            ev.c_quality       = request.form.get("c_quality") or ""
            ev.c_productivity  = request.form.get("c_productivity") or ""
            ev.c_communication = request.form.get("c_communication") or ""
            ev.c_problemsolving = request.form.get("c_problemsolving") or ""
            ev.c_compliance    = request.form.get("c_compliance") or ""

            # 3) Summary
            ev.strengths       = request.form.get("strengths") or ""
            ev.improvements    = request.form.get("improvements") or ""
            ev.training_needed = request.form.get("training_needed") or ""

            # إعادة حساب الدرجات
            compute_scores(ev)
            db.session.commit()

            flash("Evaluation updated.", "success")
            return redirect(url_for(
                "report_employee",
                emp_id=emp.id,
                week_start=week_start.isoformat(),
                week_end=week_end.isoformat()
            ))
        except Exception:
            db.session.rollback()
            flash("Error while updating evaluation.", "danger")
            return redirect(url_for("evaluate_edit", ev_id=ev.id))

    # GET → افتح نفس نموذج التقييم لكن مع تعبئة البيانات
    ws = ev.week_start
    we = ev.week_end
    return render_template("eval_form.html", employee=emp, ws=ws, we=we, weights=WEIGHTS, ev=ev)


# ----- Reports picker (generic) -----
@app.route("/reports")
@login_required
def reports_picker():
    ws, we = default_week_today()
    return render_template("report_picker.html", ws=ws, we=we)

# ----- Employee report (detailed) -----
@app.route("/reports/employee/<int:emp_id>")
@login_required
def report_employee(emp_id):
    u = cur_user()
    # admin can view any employee; supervisor only his own
    if u.role == "admin":
        emp = Employee.query.get_or_404(emp_id)
    else:
        emp = Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404()

    week_start = request.args.get("week_start")
    week_end = request.args.get("week_end")
    if not week_start or not week_end:
        flash("Missing week dates.", "warning")
        return redirect(url_for("employees"))

    ws = parse_date(week_start); we = parse_date(week_end)
    ev = Evaluation.query.filter_by(employee_id=emp.id, week_start=ws, week_end=we).first_or_404()

    evaluator = db.session.get(User, ev.evaluator_id)
    evaluator_code = evaluator.supervisor_code if evaluator else ""
    evaluator_name = evaluator.name if (evaluator and evaluator.name) else ""

    return render_template(
        "report_employee.html",
        employee=emp,
        ev=ev,
        evaluator_code=evaluator_code,
        evaluator_name=evaluator_name,
    )

# ----- Supervisor reports picker (for supervisors) -----
@app.route("/reports/supervisor/select", methods=["GET", "POST"])
@login_required
def supervisor_report_select():
    u = cur_user()
    employees = Employee.query.filter_by(user_id=u.id, is_active=True).order_by(Employee.name.asc()).all()

    if request.method == "POST":
        emp_id = int(request.form["emp_id"])
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("supervisor_report_select"))
        return redirect(url_for("report_employee",
                                emp_id=emp_id,
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    ws, we = default_week_today()
    return render_template("supervisor_report_picker.html", employees=employees, ws=ws, we=we)

# ----- Admin: Reports picker -----
@app.route("/admin/reports", methods=["GET", "POST"])
@admin_required
def admin_report_picker():
    if request.method == "POST":
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("admin_report_picker"))
        return redirect(url_for("admin_reports_all",
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    ws, we = default_week_today()
    return redirect(url_for("admin_reports_weekly",
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))

# ----- Admin: consolidated employee reports -----
@app.route("/admin/reports/all")
@admin_required
def admin_reports_all():
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")
    q = (request.args.get("q") or "").strip()

    if not (week_start and week_end):
        return redirect(url_for("admin_report_picker"))

    ws = parse_date(week_start)
    we = parse_date(week_end)

    # حدّد سوبرفايزر من نص البحث (ID أو اسم مطابق تمامًا)
    focus_sup = find_supervisor_from_query(q)

    # الاستعلام الأساسي
    query = (db.session.query(Evaluation, Employee, User)
             .join(Employee, Evaluation.employee_id == Employee.id)
             .join(User, Employee.user_id == User.id)
             .filter(Evaluation.week_start == ws, Evaluation.week_end == we))

    if cid():
        query = query.filter(Employee.company_id == cid())

    if focus_sup:
        query = query.filter(Employee.user_id == focus_sup.id)

    if q and not focus_sup:
        like = f"%{q}%"
        query = query.filter(or_(
            Employee.name.ilike(like),
            Employee.emp_number.ilike(like),
            Employee.department.ilike(like),
            Employee.site.ilike(like),
            User.supervisor_code.ilike(like),
            Evaluation.overall_band.ilike(like),
        ))

    rows = query.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    # ----- KPI + غير المُقيّمين عند تحديد سوبرفايزر -----
    sup_cov_pct = None
    missing_emps = []
    total_emp = 0
    done_emp_count = 0

    if focus_sup:
        emps = (Employee.query
                .filter_by(user_id=focus_sup.id, is_active=True)
                .order_by(Employee.name.asc())
                .all())
        total_emp = len(emps)
        if total_emp > 0:
            emp_ids = [e.id for e in emps]
            done_ids = set(
                r[0] for r in db.session.query(Evaluation.employee_id)
                .filter(Evaluation.week_start == ws,
                        Evaluation.week_end == we,
                        Evaluation.employee_id.in_(emp_ids))
                .all()
            )
            done_emp_count = len(done_ids)
            sup_cov_pct = (done_emp_count / total_emp) * 100.0
            missing_emps = [e for e in emps if e.id not in done_ids]

    # ----- قائمة السوبرفايزر + تفريد المقترحات للـ datalist -----
    supervisors = (
        apply_company_filter(User.query.filter_by(role="supervisor", is_active=True), User)
        .order_by(User.name.asc(), User.supervisor_code.asc())
        .all()
    )

    # نحضّر [(value, label)] بدون تكرار (case-insensitive)
    sup_suggestions = []
    seen = set()

    # أولاً: قيم الـID (هي فريدة غالبًا)
    for s in supervisors:
        val = (s.supervisor_code or "").strip()
        if not val:
            continue
        key = val.lower()
        if key in seen:
            continue
        label = f"{s.name or '—'} — ID: {s.supervisor_code}"
        sup_suggestions.append((val, label))
        seen.add(key)

    # ثانيًا: الأسماء (قد تتكرر، لذلك نفردها)
    for s in supervisors:
        nm = (s.name or "").strip()
        if not nm:
            continue
        key = nm.lower()
        if key in seen:
            continue
        label = f"{nm} — ID: {s.supervisor_code}"
        sup_suggestions.append((nm, label))
        seen.add(key)

    return render_template(
        "report_admin_all.html",
        ws=ws, we=we, rows=rows, q=q,
        focus_sup=focus_sup,
        sup_cov_pct=sup_cov_pct,
        total_emp=total_emp,
        done_emp_count=done_emp_count,
        missing_emps=missing_emps,
        supervisors=supervisors,
        sup_suggestions=sup_suggestions,  # ← استخدم هذه في القالب
    )

@app.route("/admin/reports/weekly")
@admin_required
def admin_reports_weekly():
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")
    q = (request.args.get("q") or "").strip()

    # لو ما فيه تاريخ نرجع لصفحة الاختيار
    if not (week_start and week_end):
        return redirect(url_for("admin_report_picker"))

    ws = parse_date(week_start)
    we = parse_date(week_end)

    # نحاول نحدد سوبرفايزر من البحث (ID أو اسم)
    focus_sup = find_supervisor_from_query(q)

    # الاستعلام الأساسي: التقييمات الأسبوعية
    query = (db.session.query(Evaluation, Employee, User)
             .join(Employee, Evaluation.employee_id == Employee.id)
             .join(User, Employee.user_id == User.id)
             .filter(Evaluation.week_start == ws,
                     Evaluation.week_end == we))

    # لو حددنا سوبرفايزر نفلتر عليه
    if focus_sup:
        query = query.filter(Employee.user_id == focus_sup.id)

    # لو فيه بحث عام وما لقينا سوبرفايزر
    if q and not focus_sup:
        like = f"%{q}%"
        query = query.filter(or_(
            Employee.name.ilike(like),
            Employee.emp_number.ilike(like),
            Employee.department.ilike(like),
            Employee.site.ilike(like),
            User.supervisor_code.ilike(like),
            Evaluation.overall_band.ilike(like),
        ))

    rows = query.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    # KPI لو كان فيه سوبرفايزر محدد
    sup_cov_pct = None
    missing_emps = []
    total_emp = 0
    done_emp_count = 0

    if focus_sup:
        emps = (Employee.query
                .filter_by(user_id=focus_sup.id, is_active=True)
                .order_by(Employee.name.asc())
                .all())
        total_emp = len(emps)
        if total_emp > 0:
            emp_ids = [e.id for e in emps]
            done_ids = set(
                r[0] for r in db.session.query(Evaluation.employee_id)
                .filter(Evaluation.week_start == ws,
                        Evaluation.week_end == we,
                        Evaluation.employee_id.in_(emp_ids))
                .all()
            )
            done_emp_count = len(done_ids)
            sup_cov_pct = (done_emp_count / total_emp) * 100.0
            missing_emps = [e for e in emps if e.id not in done_ids]

    # نجهز قائمة السوبرفايزر عشان الـ datalist
    supervisors = (
        apply_company_filter(User.query.filter_by(role="supervisor", is_active=True), User)
        .order_by(User.name.asc(), User.supervisor_code.asc())
        .all()
    )
    sup_suggestions = []
    seen = set()
    for s in supervisors:
        key = f"ID:{s.supervisor_code}"
        sup_suggestions.append((s.supervisor_code, f"{s.name or '—'} — ID: {s.supervisor_code}"))
        seen.add(key)
    for s in supervisors:
        nm = (s.name or "").strip()
        if not nm:
            continue
        key = nm.lower()
        if key in seen:
            continue
        sup_suggestions.append((nm, f"{nm} — ID: {s.supervisor_code}"))
        seen.add(key)

    return render_template(
        "report_admin_weekly.html",
        ws=ws, we=we,
        rows=rows,
        q=q,
        focus_sup=focus_sup,
        sup_cov_pct=sup_cov_pct,
        total_emp=total_emp,
        done_emp_count=done_emp_count,
        missing_emps=missing_emps,
        sup_suggestions=sup_suggestions,
    )


# ----- Supervisor weekly list (own employees) -----
@app.route("/reports/supervisor")
@login_required
def report_supervisor():
    u = cur_user()
    week_start = request.args.get("week_start"); week_end = request.args.get("week_end")
    if not week_start or not week_end:
        flash("Choose a week (start & end).", "warning")
        return redirect(url_for("employees"))
    ws = parse_date(week_start); we = parse_date(week_end)
    evals = (db.session.query(Evaluation, Employee)
             .join(Employee, Evaluation.employee_id == Employee.id)
             .filter(Employee.user_id == u.id,
                     Evaluation.week_start == ws, Evaluation.week_end == we)
             .order_by(Employee.name.asc()).all())
    return render_template("report_supervisor.html", user=u, ws=ws, we=we, evals=evals)

# ----- Admin: All Employees + search + history -----
@app.route("/admin/employees")
@admin_required
def admin_employees():
    q      = (request.args.get("q") or "").strip()
    flt    = (request.args.get("filter") or "").strip()   # unassigned | deactivated

    if flt == "unassigned":
        # موظفون بدون مشرف
        emp_q = apply_company_filter(
            Employee.query.filter(Employee.user_id.is_(None)), Employee)
        if q:
            like = f"%{q}%"
            emp_q = emp_q.filter(or_(
                Employee.name.ilike(like),
                Employee.emp_number.ilike(like),
                Employee.department.ilike(like),
                Employee.site.ilike(like),
            ))
        emp_rows = [(emp, None) for emp in emp_q.order_by(Employee.name.asc()).all()]

    elif flt == "deactivated":
        # موظفون حذفهم مشرفهم (is_active=False, لديهم مشرف)
        emp_q = (db.session.query(Employee, User)
                 .join(User, Employee.user_id == User.id)
                 .filter(Employee.is_active == False))
        if cid():
            emp_q = emp_q.filter(Employee.company_id == cid())
        if q:
            like = f"%{q}%"
            emp_q = emp_q.filter(or_(
                Employee.name.ilike(like),
                Employee.emp_number.ilike(like),
                Employee.department.ilike(like),
                Employee.site.ilike(like),
                User.supervisor_code.ilike(like),
                User.name.ilike(like),
            ))
        emp_rows = emp_q.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    else:
        # الكل (نشطون فقط مع مشرف)
        emp_q = (db.session.query(Employee, User)
                 .join(User, Employee.user_id == User.id)
                 .filter(Employee.is_active == True))
        if cid():
            emp_q = emp_q.filter(Employee.company_id == cid())
        if q:
            like = f"%{q}%"
            emp_q = emp_q.filter(or_(
                Employee.name.ilike(like),
                Employee.emp_number.ilike(like),
                Employee.department.ilike(like),
                Employee.site.ilike(like),
                User.supervisor_code.ilike(like),
                User.name.ilike(like),
            ))
        emp_rows = emp_q.order_by(User.supervisor_code.asc(), Employee.name.asc()).all()

    # عدد كل فئة
    _base = apply_company_filter(Employee.query, Employee)
    count_unassigned  = _base.filter(Employee.user_id.is_(None)).count()
    count_deactivated = _base.filter(Employee.is_active == False, Employee.user_id.isnot(None)).count()

    users_q = apply_company_filter(
        User.query.filter(User.role.in_(["supervisor", "site_supervisor"])), User)
    if q:
        like = f"%{q}%"
        users_q = users_q.filter(or_(
            User.supervisor_code.ilike(like),
            User.name.ilike(like),
            User.role.ilike(like),
        ))
    users = users_q.order_by(User.supervisor_code.asc()).all()

    return render_template("admin_employees.html", rows=emp_rows, users=users,
                           q=q, flt=flt,
                           count_unassigned=count_unassigned,
                           count_deactivated=count_deactivated)

@app.route("/admin/employee/<int:emp_id>/history")
@admin_required
def admin_employee_history(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    evals = (Evaluation.query.filter_by(employee_id=emp.id)
             .order_by(Evaluation.week_start.desc()).all())
    return render_template("employee_history.html", emp=emp, evals=evals)

# ===================== Site Supervisor Features =====================
# Manage assigned supervisors
@app.route("/site/supervisors", methods=["GET", "POST"])
@login_required
def site_supervisors():
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    if request.method == "POST":
        sup_code = (request.form.get("supervisor_code") or "").strip()
        sup_name = (request.form.get("supervisor_name") or "").strip()
        if not sup_code:
            flash("Please enter a Supervisor ID.", "danger")
            return redirect(url_for("site_supervisors"))

        # جرّب نلقى مستخدم بهذا الـID
        sup_user = User.query.filter_by(supervisor_code=sup_code).first()

        # لو ما وُجد: أنشئه كمشرف (Supervisor) مفعَّل
        if not sup_user:
            sup_user = User(
                supervisor_code=sup_code,
                name=sup_name,
                role="supervisor",
                is_active=True,
                company_id=cid(),
            )
            db.session.add(sup_user)
            db.session.commit()
            flash("Supervisor user created and assigned.", "success")
        else:
            # لو موجود لكنه مو مشرف، ما نسمح بربطه
            if sup_user.role != "supervisor":
                flash("This ID exists but is not a Supervisor role.", "danger")
                return redirect(url_for("site_supervisors"))
            if not sup_user.is_active:
                flash("This Supervisor is inactive. Ask Admin to activate.", "warning")
                return redirect(url_for("site_supervisors"))

        # اربطه إن ما كان مرتبط مسبقًا
        exists = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_user.id).first()
        if exists:
            flash("This supervisor is already assigned.", "warning")
        else:
            link = SiteSupervisorMap(site_sup_id=u.id, supervisor_id=sup_user.id, company_id=cid())
            db.session.add(link)
            db.session.commit()
            flash("Supervisor assigned.", "success")

        return redirect(url_for("site_supervisors"))

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    return render_template("site_supervisors.html", links=links)

@app.post("/site/supervisors/<int:link_id>/remove")
@login_required
def site_supervisors_remove(link_id):
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)
    link = SiteSupervisorMap.query.get_or_404(link_id)
    if link.site_sup_id != u.id:
        abort(403)
    db.session.delete(link)
    db.session.commit()
    flash("Removed.", "success")
    return redirect(url_for("site_supervisors"))

# New evaluation for a supervisor
@app.route("/site/evaluate/<int:sup_user_id>/new", methods=["GET", "POST"])
@login_required
def site_evaluate_new(sup_user_id):
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    # تأكد أنه ضمن قائمته
    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_user_id).first()
    if not link:
        abort(403)

    supervisor = User.query.get_or_404(sup_user_id)
    if request.method == "POST":
        ws = parse_date(request.form.get("week_start"))
        we = parse_date(request.form.get("week_end"))
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("site_evaluate_new", sup_user_id=sup_user_id))

        if SupervisorEvaluation.query.filter_by(supervisor_id=sup_user_id, week_start=ws, week_end=we).first():
            flash("An evaluation for this supervisor already exists this week.", "warning")
            return redirect(url_for("site_report_supervisor", sup_user_id=sup_user_id,
                                    week_start=ws.isoformat(), week_end=we.isoformat()))

        se = SupervisorEvaluation(
            supervisor_id=sup_user_id, evaluator_id=u.id,
            week_start=ws, week_end=we,

            t1_text=request.form.get("t1_text",""), t1_percent=float(request.form.get("t1_percent") or 0), t1_remarks=request.form.get("t1_remarks",""),
            t2_text=request.form.get("t2_text",""), t2_percent=float(request.form.get("t2_percent") or 0), t2_remarks=request.form.get("t2_remarks",""),
            t3_text=request.form.get("t3_text",""), t3_percent=float(request.form.get("t3_percent") or 0), t3_remarks=request.form.get("t3_remarks",""),
            t4_text=request.form.get("t4_text",""), t4_percent=float(request.form.get("t4_percent") or 0), t4_remarks=request.form.get("t4_remarks",""),

            p_punctuality=int(request.form.get("p_punctuality") or 0), c_punctuality=request.form.get("c_punctuality",""),
            p_quality=int(request.form.get("p_quality") or 0), c_quality=request.form.get("c_quality",""),
            p_productivity=int(request.form.get("p_productivity") or 0), c_productivity=request.form.get("c_productivity",""),
            p_communication=int(request.form.get("p_communication") or 0), c_communication=request.form.get("c_communication",""),
            p_problemsolving=int(request.form.get("p_problemsolving") or 0), c_problemsolving=request.form.get("c_problemsolving",""),
            p_compliance=int(request.form.get("p_compliance") or 0), c_compliance=request.form.get("c_compliance",""),

            strengths=request.form.get("strengths",""),
            improvements=request.form.get("improvements",""),
            training_needed=request.form.get("training_needed",""),
            company_id=cid(),
        )

        # نفس المعادلة بالضبط
        compute_scores(se)
        db.session.add(se)
        db.session.commit()
        flash("Supervisor evaluation saved.", "success")
        return redirect(url_for("site_report_supervisor", sup_user_id=sup_user_id,
                                week_start=ws.isoformat(), week_end=we.isoformat()))

    ws, we = default_week_today()
    return render_template("site_eval_form.html", supervisor=supervisor, ws=ws, we=we, weights=WEIGHTS)

# Site supervisor report picker
@app.route("/site/reports/select", methods=["GET", "POST"])
@login_required
def site_report_select():
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)
    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    supervisors = [su for _, su in links]

    if request.method == "POST":
        sup_user_id = int(request.form["sup_user_id"])
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("site_report_select"))
        return redirect(url_for("site_report_supervisor",
                                sup_user_id=sup_user_id,
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    ws, we = default_week_today()
    return render_template("site_report_picker.html", supervisors=supervisors, ws=ws, we=we)

# Detailed supervisor report
@app.route("/site/reports/supervisor/<int:sup_user_id>")
@login_required
def site_report_supervisor(sup_user_id):
    u = cur_user()
    if u.role not in ["site_supervisor", "admin"]:
        abort(403)
    if u.role == "site_supervisor":
        link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_user_id).first()
        if not link:
            abort(403)

    week_start = request.args.get("week_start")
    week_end = request.args.get("week_end")
    if not (week_start and week_end):
        flash("Missing week dates.", "warning")
        return redirect(url_for("site_report_select") if u.role == "site_supervisor" else url_for("admin_site_report_picker"))

    ws = parse_date(week_start); we = parse_date(week_end)
    se = SupervisorEvaluation.query.filter_by(supervisor_id=sup_user_id, week_start=ws, week_end=we).first_or_404()
    supervisor = User.query.get_or_404(sup_user_id)
    evaluator = db.session.get(User, se.evaluator_id)
    return render_template("site_report_supervisor.html", se=se, supervisor=supervisor, evaluator=evaluator)

@app.route("/site/daily/reports")
@login_required
def site_daily_reports_site():
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    # تاريخ اليوم أو اللي اختاره
    d_s = request.args.get("d")
    if d_s:
        d_val = parse_date(d_s)
    else:
        d_val = date.today()

    # المشرفين اللي تابعين لهذا الـ site supervisor
    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id)
             .all())
    sup_ids = [su.id for _, su in links]

    # نفس استعلام صفحة /daily/reports لكن مفلتر على اللي فوق
    q = (db.session.query(DailyEvaluation, Employee)
         .join(Employee, DailyEvaluation.employee_id == Employee.id)
         .filter(DailyEvaluation.eval_date == d_val))

    if sup_ids:
        q = q.filter(Employee.user_id.in_(sup_ids))

    rows = (q.order_by(Employee.name.asc()).all())

    # نستخدم نفس التمبليت الجاهز
    return render_template("daily_reports.html", rows=rows, d=d_val)
    
@app.route("/site/attendance")
@login_required
def site_attendance_site():
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    d_s = request.args.get("d")
    if d_s:
        d_val = parse_date(d_s)
    else:
        d_val = date.today()

    # اشوف المشرفين اللي تحت هذا الـ site sup
    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id)
             .all())
    sup_ids = [su.id for _, su in links]

    # هذا يفترض إن عندك موديل Attendance بنفس شكل admin
    q = (db.session.query(Attendance, Employee, User)
         .join(Employee, Attendance.employee_id == Employee.id)
         .join(User, Employee.user_id == User.id)
         .filter(Attendance.date == d_val))

    if sup_ids:
        q = q.filter(Employee.user_id.in_(sup_ids))

    rows = q.order_by(User.name.asc(), Employee.name.asc()).all()

    # نرجّع نفس قالب الحضور اللي عندك
    return render_template("attendance_admin.html", rows=rows, d=d_val)

@app.route("/site/requests")
@login_required
def site_requests_site():
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    # المشرفين اللي تحته
    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id)
             .all())
    sup_ids = [su.id for _, su in links]

    # نجيب الطلبات اللي أنشأها هؤلاء المشرفون
    req_q = (Request.query
             .filter(Request.supervisor_id.in_(sup_ids))
             .order_by(Request.created_at.desc()))
    items = req_q.all()

    return render_template("requests_inbox.html", items=items)

@app.route("/site/daily/report/<int:emp_id>")
@login_required
def site_daily_report_employee(emp_id):
    u = cur_user()
    if u.role != "site_supervisor":
        abort(403)

    d_str = request.args.get("d")
    if not d_str:
        flash("Missing date.", "warning")
        return redirect(url_for("site_daily_reports_site"))
    d_val = parse_date(d_str)

    # الموظف
    emp = Employee.query.get_or_404(emp_id)

    # تأكد أن الموظف يتبع مشرف من مشرفين هذا الـ site supervisor
    link = (db.session.query(SiteSupervisorMap)
            .filter_by(site_sup_id=u.id, supervisor_id=emp.user_id)
            .first())
    if not link:
        abort(403)

    # التقييم اليومي نفسه
    de = (DailyEvaluation.query
          .filter_by(employee_id=emp_id, eval_date=d_val)
          .first_or_404())

    # نستعمل نفس القالب اللي تستعمله للمشرف العادي
    return render_template("daily_report_employee.html",
                           employee=emp,
                           de=de,
                           d=d_val)



# ----- Admin: Site reviews picker & consolidated -----
@app.route("/admin/site/reports", methods=["GET", "POST"])
@admin_required
def admin_site_report_picker():
    if request.method == "POST":
        ws = parse_date(request.form["week_start"])
        we = parse_date(request.form["week_end"])
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            flash(msg, "danger")
            return redirect(url_for("admin_site_report_picker"))
        return redirect(url_for("admin_site_reports_all",
                                week_start=ws.isoformat(),
                                week_end=we.isoformat()))
    ws, we = default_week_today()
    return render_template("site_admin_picker.html", ws=ws, we=we)

@app.route("/admin/site/reports/all")
@admin_required
def admin_site_reports_all():
    week_start = request.args.get("week_start")
    week_end   = request.args.get("week_end")
    q_raw = request.args.get("q") or ""
    q = q_raw.strip().lower()

    if not (week_start and week_end):
        return redirect(url_for("admin_site_report_picker"))

    ws = parse_date(week_start); we = parse_date(week_end)

    # اجلب جميع تقييمات الأسبوع واربط الأسماء يدويًا (أضمن من joins متعددة على User)
    rows = []
    se_list = SupervisorEvaluation.query.filter_by(week_start=ws, week_end=we).all()
    for se in se_list:
        sup_user  = db.session.get(User, se.supervisor_id)   # المشرف المُقيَّم
        site_user = db.session.get(User, se.evaluator_id)    # مشرف السايت المُقيِّم

        if q:
            hay = " ".join([
                sup_user.name or "", sup_user.supervisor_code or "",
                site_user.name or "", site_user.supervisor_code or "",
                se.overall_band or ""
            ]).lower()
            if q not in hay:
                continue
        rows.append((se, sup_user, site_user))

    return render_template("site_admin_all.html", ws=ws, we=we, rows=rows, q=q_raw)

import traceback

@app.errorhandler(404)
def _not_found(e):
    app.logger.error("404 path=%s method=%s", request.path, request.method)
    return "Not Found", 404

# ============================================================
#  ملف التعديلات للتطبيق — أضف هذا الكود في آخر ملف main.py
#  قبل سطر: ensure_db_and_admin()
# ============================================================

# المكتبات المُستوردة أعلى الملف — uuid, json, secrets, freq مستوردة هنا للـ API section
import uuid
import json
import secrets
from flask import jsonify, request as freq

from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

# ─────────────────────────────────────────────
#  إعداد رفع الملفات
# ─────────────────────────────────────────────
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 ميجا كحد أقصى

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# ─────────────────────────────────────────────
#  موديل التوكن للتطبيق (جدول منفصل)
# ─────────────────────────────────────────────
class MobileToken(db.Model):
    __tablename__ = "mobile_token"
    id           = db.Column(db.Integer, primary_key=True)
    user_id      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    token        = db.Column(db.String(64), unique=True, nullable=False)
    device_token = db.Column(db.String(255), nullable=True)
    company_id   = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    user         = db.relationship("User", backref="mobile_tokens")

# ─────────────────────────────────────────────
#  موديل مرفق الطلب (PDF)
# ─────────────────────────────────────────────
class RequestAttachment(db.Model):
    __tablename__ = "request_attachment"
    id          = db.Column(db.Integer, primary_key=True)
    request_id  = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False)
    filename    = db.Column(db.String(255), nullable=False)   # اسم الملف المحفوظ
    original_name = db.Column(db.String(255), nullable=True)  # الاسم الأصلي
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    request     = db.relationship("Request", backref="attachments")

# ─────────────────────────────────────────────
#  موديل نموذج الإجازة الثابت
# ─────────────────────────────────────────────
class LeaveForm(db.Model):
    __tablename__ = "leave_form"
    id          = db.Column(db.Integer, primary_key=True)
    request_id  = db.Column(db.Integer, db.ForeignKey("requests.id"), nullable=False, unique=True)
    filename    = db.Column(db.String(255), nullable=False)   # اسم ملف PDF المُولّد
    generated_at = db.Column(db.DateTime, default=datetime.utcnow)
    signed      = db.Column(db.Boolean, default=False)
    signed_at   = db.Column(db.DateTime, nullable=True)
    request     = db.relationship("Request", backref="leave_form", uselist=False)

# db.create_all() تُستدعى مرة واحدة في نهاية الملف

# ─────────────────────────────────────────────
#  مساعد: التحقق من توكن التطبيق
# ─────────────────────────────────────────────
def get_api_user():
    auth = freq.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return None
    token_str = auth[7:]
    tok = MobileToken.query.filter_by(token=token_str).first()
    return tok.user if tok else None

def get_api_token():
    auth = freq.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return None
    return MobileToken.query.filter_by(token=auth[7:]).first()

def api_cid():
    """Returns company_id from the API token (for data isolation in API routes)."""
    tok = get_api_token()
    if not tok:
        return None
    u = tok.user
    if u and u.role == "super_admin":
        return None
    return tok.company_id or (u.company_id if u else None)

def api_login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = get_api_user()
        if not u or not u.is_active:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return wrapper

def api_admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = get_api_user()
        if not u or u.role != "admin":
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def api_hr_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = get_api_user()
        if not u or u.role != "hr":
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

# ─────────────────────────────────────────────
#  مساعد توليد نموذج الإجازة PDF
# ─────────────────────────────────────────────

def generate_warning_pdf(req: "Request", sig: "WarningSignature") -> str:
    """يولّد نموذج إنذار PDF مع توقيع الموظف"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import base64, io
    except ImportError:
        return ""

    # ── تسجيل الخط العربي ──
    ARABIC_FONT = "Arabic"
    ARABIC_TTF  = "/usr/share/fonts/google-droid/DroidSansArabic.ttf"
    LATIN_FONT  = "Helvetica"
    try:
        pdfmetrics.registerFont(TTFont(ARABIC_FONT, ARABIC_TTF))
    except Exception:
        ARABIC_FONT = "Helvetica"  # fallback

    def ar(text):
        """عكس النص العربي فقط لـ ReportLab (RTL)"""
        t = str(text)
        if not any("؀" <= c <= "ۿ" for c in t):
            return t
        try:
            from arabic_reshaper import reshape
            from bidi.algorithm import get_display
            return get_display(reshape(t))
        except Exception:
            try:
                from bidi.algorithm import get_display
                return get_display(t)
            except Exception:
                return t

    def font_for(text):
        """اختر الخط المناسب حسب المحتوى"""
        if any("؀" <= c <= "ۿ" for c in str(text)):
            return ARABIC_FONT
        return LATIN_FONT

    emp  = req.employee
    sup  = req.supervisor
    task_item = HRTask.query.filter_by(request_id=req.id).first()
    try:
        reason_text = (task_item.official_reason if task_item and task_item.official_reason else req.reason) or "—"
    except Exception:
        reason_text = req.reason or "—"
    fname = f"warning_{req.id}_{uuid.uuid4().hex[:8]}.pdf"
    fpath = os.path.join(UPLOAD_FOLDER, fname)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("wtitle", parent=styles["Title"],
                                 fontName=ARABIC_FONT,
                                 fontSize=18, spaceAfter=6,
                                 alignment=1,
                                 textColor=colors.HexColor("#c0392b"))
    sub_style   = ParagraphStyle("wsub", parent=styles["Normal"],
                                 fontName=ARABIC_FONT,
                                 fontSize=11, spaceAfter=10, alignment=1)
    body_style  = ParagraphStyle("wbody", parent=styles["Normal"],
                                 fontName=ARABIC_FONT,
                                 fontSize=11, spaceAfter=6, alignment=2)  # right align
    elements = []

    # ── شعار الشركة ──
    logo_path = os.path.join(BASE_DIR, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        elements.append(RLImage(logo_path, width=4*cm, height=2*cm))
        elements.append(Spacer(1, 0.3*cm))

    # ── عنوان النموذج ──
    elements.append(Paragraph(ar("نموذج إنذار رسمي"), title_style))
    elements.append(Paragraph("Official Warning Notice", ParagraphStyle("wsub_en", parent=styles["Normal"], fontName=LATIN_FONT, fontSize=11, spaceAfter=10, alignment=1)))
    elements.append(Spacer(1, 0.5*cm))

    # ── بيانات الموظف ──
    from zoneinfo import ZoneInfo
    now_str = datetime.now(ZoneInfo("Asia/Riyadh")).strftime("%Y/%m/%d")
    def cell(text):
        t = ar(text)
        return Paragraph(t, ParagraphStyle("c", fontName=font_for(text), fontSize=10, alignment=2 if any("؀"<=c<="ۿ" for c in str(text)) else 0))

    rows = [
        [cell("البند"), cell("البيانات")],
        [cell("اسم الموظف"),   cell(emp.name if emp else "—")],
        [cell("رقم الموظف"),   cell(emp.emp_number if emp else "—")],
        [cell("القسم"),        cell(emp.department or "—")],
        [cell("الموقع"),       cell(emp.site or "—")],
        [cell("المشرف"),       cell(sup.name if sup else "—")],
        [cell("سبب الإنذار"),  cell(reason_text)],
        [cell("تاريخ الإنذار"), cell(now_str)],
    ]

    tbl = Table(rows, colWidths=[5*cm, 11*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#c0392b")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, -1), ARABIC_FONT),
        ("FONTSIZE",      (0, 0), (-1, 0), 12),
        ("ALIGN",         (0, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE",      (0, 1), (-1, -1), 10),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.HexColor("#fdf2f2"), colors.white]),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 1*cm))

    # ── نص الإنذار ──
    elements.append(Paragraph(
        ar("بناءً على المخالفة المذكورة أعلاه، يُنذر الموظف رسمياً ويُطلب منه الالتزام بالأنظمة والتعليمات."),
        ParagraphStyle("warn_body", parent=styles["Normal"], fontName=ARABIC_FONT, fontSize=10, leading=16, alignment=2)
    ))
    elements.append(Spacer(1, 1*cm))

    # ── توقيع الموظف ──
    sign_elements = []
    try:
        img_data = base64.b64decode(sig.signature)
        img_buf  = io.BytesIO(img_data)
        sign_img = RLImage(img_buf, width=5*cm, height=2*cm)
        sign_elements.append(sign_img)
    except Exception:
        sign_elements.append(Paragraph("_" * 30, styles["Normal"]))

    sig_name = sig.employee_name or (emp.name if emp else "")
    signed_date = sig.signed_at.strftime("%Y/%m/%d") if sig.signed_at else now_str

    sign_rows = [
        [cell("توقيع الموظف"), cell("اسم الموظف"), cell("التاريخ")],
        [sign_elements, cell(sig_name), cell(signed_date)],
    ]
    sign_tbl = Table(sign_rows, colWidths=[6*cm, 5*cm, 5*cm])
    sign_tbl.setStyle(TableStyle([
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",      (0, 0), (-1, -1), LATIN_FONT),
        ("FONTSIZE",      (0, 0), (-1, -1), 10),
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("BOX",           (0, 0), (-1, -1), 0.5, colors.grey),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(sign_tbl)

    doc.build(elements)
    return fname

def generate_leave_pdf(req: "Request", sig: "LeaveSignature" = None) -> str:
    """نموذج إجازة رسمي عربي/إنجليزي صفحة واحدة مع إمكانية التوقيع."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, Image as RLImage,
                                        HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import base64, io
    except ImportError:
        return ""

    # ── خط عربي ──
    ARABIC_FONT = "Arabic"
    ARABIC_TTF  = "/usr/share/fonts/google-droid/DroidSansArabic.ttf"
    LATIN_FONT  = "Helvetica"
    try:
        pdfmetrics.registerFont(TTFont(ARABIC_FONT, ARABIC_TTF))
        use_arabic = True
    except Exception:
        ARABIC_FONT = "Helvetica"
        use_arabic  = False

    def ar(txt: str) -> str:
        if not use_arabic:
            return txt
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(txt))
        except Exception:
            return txt

    emp  = req.employee
    sup  = req.supervisor
    company_name = ""
    try:
        from sqlalchemy.orm import relationship
        if emp and emp.company_id:
            from main import Company
            c = Company.query.get(emp.company_id)
            if c:
                company_name = c.name or ""
    except Exception:
        pass

    days_count = "—"
    if req.start_date and req.end_date:
        days_count = str((req.end_date - req.start_date).days + 1)

    fname = f"leave_{req.id}_{uuid.uuid4().hex[:8]}.pdf"
    fpath = os.path.join(UPLOAD_FOLDER, fname)

    W, H = A4
    doc  = SimpleDocTemplate(fpath, pagesize=A4,
                              rightMargin=1.8*cm, leftMargin=1.8*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles   = getSampleStyleSheet()
    ar_style = ParagraphStyle("ar", fontName=ARABIC_FONT, fontSize=10,
                               alignment=2, leading=16)
    en_style = ParagraphStyle("en", fontName=LATIN_FONT,  fontSize=9,
                               alignment=0, textColor=colors.HexColor("#555555"), leading=14)
    hdr_style= ParagraphStyle("hdr",fontName=ARABIC_FONT, fontSize=14,
                               alignment=1, spaceAfter=2, textColor=colors.HexColor("#1a1f3a"))

    DARK  = colors.HexColor("#1a1f3a")
    BLUE  = colors.HexColor("#2563eb")
    LIGHT = colors.HexColor("#f0f4ff")
    GREY  = colors.HexColor("#f8f8f8")

    def cell(ar_text, en_text="", bold=False, bg=None, center=False):
        align = 1 if center else 2
        f     = ARABIC_FONT
        p_ar  = Paragraph(ar(ar_text), ParagraphStyle("c", fontName=f,
                          fontSize=10, alignment=align, leading=14,
                          textColor=colors.black))
        p_en  = Paragraph(en_text, ParagraphStyle("ce", fontName=LATIN_FONT,
                          fontSize=8,  alignment=1 if center else 0,
                          textColor=colors.HexColor("#777777"), leading=12))
        return [p_ar, p_en] if en_text else p_ar

    elements = []

    # ══ رأس الصفحة: شعار + اسم الشركة + عنوان ══
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    header_cols = []
    if os.path.exists(logo_path):
        try:
            header_cols.append(RLImage(logo_path, width=2.5*cm, height=2.5*cm))
        except Exception:
            header_cols.append(Paragraph("", styles["Normal"]))
    else:
        header_cols.append(Paragraph("", styles["Normal"]))

    header_cols.append(
        Table([[Paragraph(ar("نموذج طلب إجازة"), ParagraphStyle("t1",
                          fontName=ARABIC_FONT, fontSize=16, alignment=1,
                          textColor=DARK, spaceAfter=4))],
               [Paragraph("LEAVE REQUEST FORM", ParagraphStyle("t2",
                          fontName=LATIN_FONT,  fontSize=10, alignment=1,
                          textColor=BLUE))],
               [Paragraph(ar(company_name) if company_name else "",
                          ParagraphStyle("t3", fontName=ARABIC_FONT, fontSize=9,
                          alignment=1, textColor=colors.HexColor("#888888")))]],
              colWidths=[W - 3.6*cm - 2.5*cm])
    )
    header_cols.append(Paragraph("", styles["Normal"]))

    hdr_tbl = Table([header_cols], colWidths=[2.5*cm, W - 3.6*cm - 2.5*cm, None])
    hdr_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    elements.append(hdr_tbl)
    elements.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=10))

    # ══ بيانات الموظف ══
    def info_row(label_ar, label_en, val):
        val_str = val or "—"
        return [
            Paragraph(ar(label_ar), ParagraphStyle("lbl", fontName=ARABIC_FONT,
                      fontSize=9, alignment=2, textColor=colors.HexColor("#555"))),
            Paragraph(label_en, ParagraphStyle("lble", fontName=LATIN_FONT,
                      fontSize=7.5, alignment=0, textColor=colors.HexColor("#999"))),
            Paragraph(ar(str(val_str)), ParagraphStyle("val", fontName=ARABIC_FONT,
                      fontSize=10, alignment=2, textColor=DARK)),
        ]

    info_data = [
        info_row("اسم الموظف",    "Employee Name",   emp.name if emp else ""),
        info_row("رقم الموظف",    "Employee No.",    emp.emp_number if emp else ""),
        info_row("القسم",         "Department",      emp.department if emp else ""),
        info_row("الموقع",        "Site",            emp.site if emp else ""),
        info_row("المشرف المباشر","Direct Supervisor",sup.name if sup else ""),
    ]
    info_tbl = Table(info_data, colWidths=[3.5*cm, 3*cm, 9*cm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), GREY),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [colors.white, GREY]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#dddddd")),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 0.4*cm))

    # ══ بيانات الإجازة ══
    elements.append(
        Paragraph(ar("تفاصيل الإجازة  |  Leave Details"),
                  ParagraphStyle("sec", fontName=ARABIC_FONT, fontSize=11,
                  alignment=1, textColor=BLUE, spaceAfter=4))
    )

    leave_data = [
        info_row("نوع الإجازة",    "Leave Type",     "إجازة اعتيادية" if req.type == "leave" else req.type),
        info_row("تاريخ البداية",  "Start Date",     str(req.start_date) if req.start_date else ""),
        info_row("تاريخ الانتهاء", "End Date",       str(req.end_date)   if req.end_date   else ""),
        info_row("عدد الأيام",     "No. of Days",    days_count),
        info_row("السبب",          "Reason",         req.reason or ""),
    ]
    leave_tbl = Table(leave_data, colWidths=[3.5*cm, 3*cm, 9*cm])
    leave_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), LIGHT),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [LIGHT, colors.white]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#c0d0ff")),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    elements.append(leave_tbl)
    elements.append(Spacer(1, 0.5*cm))

    # ══ التوقيعات ══
    elements.append(
        Paragraph(ar("التوقيعات  |  Signatures"),
                  ParagraphStyle("sec2", fontName=ARABIC_FONT, fontSize=11,
                  alignment=1, textColor=BLUE, spaceAfter=4))
    )

    # بناء خلية توقيع الموظف
    emp_sig_cell = []
    if sig and sig.signature:
        try:
            img_data = base64.b64decode(sig.signature)
            img_buf  = io.BytesIO(img_data)
            emp_sig_cell.append(RLImage(img_buf, width=4.5*cm, height=1.8*cm))
        except Exception:
            emp_sig_cell.append(Paragraph("_" * 20, styles["Normal"]))
        emp_sig_cell.append(Paragraph(
            ar(sig.employee_name or (emp.name if emp else "")),
            ParagraphStyle("sn", fontName=ARABIC_FONT, fontSize=8.5,
                           alignment=1, textColor=DARK)))
        signed_date = sig.signed_at.strftime("%Y/%m/%d") if sig.signed_at else ""
        emp_sig_cell.append(Paragraph(signed_date,
            ParagraphStyle("sd", fontName=LATIN_FONT, fontSize=8,
                           alignment=1, textColor=colors.HexColor("#777"))))
    else:
        emp_sig_cell = [Paragraph("\n\n\n", styles["Normal"]),
                        Paragraph(ar("توقيع الموظف"),
                            ParagraphStyle("sl", fontName=ARABIC_FONT, fontSize=8.5,
                                           alignment=1, textColor=colors.HexColor("#aaa")))]

    sup_sig_cell  = [Paragraph("\n\n\n", styles["Normal"]),
                     Paragraph(ar("توقيع المشرف"),
                         ParagraphStyle("s2", fontName=ARABIC_FONT, fontSize=8.5,
                                        alignment=1, textColor=colors.HexColor("#aaa")))]
    hr_sig_cell   = [Paragraph("\n\n\n", styles["Normal"]),
                     Paragraph(ar("توقيع الموارد البشرية"),
                         ParagraphStyle("s3", fontName=ARABIC_FONT, fontSize=8.5,
                                        alignment=1, textColor=colors.HexColor("#aaa")))]

    sig_tbl = Table([[emp_sig_cell, sup_sig_cell, hr_sig_cell]],
                    colWidths=[5.3*cm, 5.3*cm, 5.3*cm])
    sig_tbl.setStyle(TableStyle([
        ("BOX",           (0,0), (0,0), 1, colors.HexColor("#2563eb")),
        ("BOX",           (1,0), (1,0), 1, colors.HexColor("#cccccc")),
        ("BOX",           (2,0), (2,0), 1, colors.HexColor("#cccccc")),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "BOTTOM"),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("BACKGROUND",    (0,0), (0,0), colors.HexColor("#f0f4ff")),
    ]))
    elements.append(sig_tbl)

    # ══ تذييل ══
    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    from datetime import datetime as dt_cls
    elements.append(Paragraph(
        f"Generated: {dt_cls.utcnow().strftime('%Y-%m-%d')}  |  NSH SafeTrack",
        ParagraphStyle("ft", fontName=LATIN_FONT, fontSize=7.5,
                       alignment=1, textColor=colors.HexColor("#aaaaaa"))))

    doc.build(elements)
    return fname


def generate_permission_pdf(req: "Request", sig: "PermissionSignature" = None) -> str:
    """نموذج استئذان رسمي عربي/إنجليزي مع التوقيع."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, Image as RLImage,
                                        HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import base64, io
    except ImportError:
        return ""

    ARABIC_FONT = "Arabic"
    ARABIC_TTF  = "/usr/share/fonts/google-droid/DroidSansArabic.ttf"
    LATIN_FONT  = "Helvetica"
    try:
        pdfmetrics.registerFont(TTFont(ARABIC_FONT, ARABIC_TTF))
        use_arabic = True
    except Exception:
        ARABIC_FONT = "Helvetica"
        use_arabic  = False

    def ar(txt):
        if not use_arabic:
            return txt
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(txt))
        except Exception:
            return txt

    emp = req.employee
    sup = req.supervisor
    fname = f"permission_{req.id}_{uuid.uuid4().hex[:8]}.pdf"
    fpath = os.path.join(UPLOAD_FOLDER, fname)
    W, H  = A4

    styles = getSampleStyleSheet()
    DARK   = colors.HexColor("#1a1f3a")
    BLUE   = colors.HexColor("#2563eb")
    LIGHT  = colors.HexColor("#f0f4ff")
    GREY   = colors.HexColor("#f8f8f8")

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    elements = []

    def info_row(label_ar, label_en, val):
        return [
            Paragraph(ar(label_ar), ParagraphStyle("lbl", fontName=ARABIC_FONT, fontSize=9,
                      alignment=2, textColor=colors.HexColor("#555"))),
            Paragraph(label_en, ParagraphStyle("lble", fontName=LATIN_FONT, fontSize=7.5,
                      alignment=0, textColor=colors.HexColor("#999"))),
            Paragraph(ar(str(val or "—")), ParagraphStyle("val", fontName=ARABIC_FONT, fontSize=10,
                      alignment=2, textColor=DARK)),
        ]

    # رأس
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    logo_cell = RLImage(logo_path, width=2.5*cm, height=2.5*cm) if os.path.exists(logo_path) \
                else Paragraph("", styles["Normal"])
    title_tbl = Table([[logo_cell,
        Table([[Paragraph(ar("نموذج استئذان"), ParagraphStyle("t1", fontName=ARABIC_FONT,
                              fontSize=16, alignment=1, textColor=DARK))],
               [Paragraph("PERMISSION FORM", ParagraphStyle("t2", fontName=LATIN_FONT,
                              fontSize=10, alignment=1, textColor=BLUE))]],
              colWidths=[W - 3.6*cm - 2.5*cm])]], colWidths=[2.5*cm, W - 3.6*cm - 2.5*cm, None])
    title_tbl.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                                   ("TOPPADDING",(0,0),(-1,-1),6),
                                   ("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    elements.append(title_tbl)
    elements.append(HRFlowable(width="100%", thickness=2, color=BLUE, spaceAfter=10))

    rows = [
        info_row("اسم الموظف",   "Employee Name",  emp.name if emp else ""),
        info_row("رقم الموظف",   "Employee No.",   emp.emp_number if emp else ""),
        info_row("القسم",        "Department",     emp.department if emp else ""),
        info_row("المشرف",       "Supervisor",     sup.name if sup else ""),
        info_row("التاريخ",      "Date",           str(req.start_date or req.created_at.date())),
        info_row("السبب",        "Reason",         req.reason or ""),
    ]
    tbl = Table(rows, colWidths=[3.5*cm, 3*cm, 9*cm])
    tbl.setStyle(TableStyle([
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[LIGHT, colors.white]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#c0d0ff")),
        ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 0.6*cm))

    # توقيعات
    elements.append(Paragraph(ar("التوقيعات  |  Signatures"),
        ParagraphStyle("sec", fontName=ARABIC_FONT, fontSize=11, alignment=1,
                       textColor=BLUE, spaceAfter=4)))

    def sig_cell(sig_obj, label_ar, label_en, name_str=""):
        if sig_obj and sig_obj.signature:
            try:
                img_buf = io.BytesIO(base64.b64decode(sig_obj.signature))
                img     = RLImage(img_buf, width=4.5*cm, height=1.8*cm)
                return [img,
                        Paragraph(ar(sig_obj.employee_name or name_str),
                            ParagraphStyle("sn", fontName=ARABIC_FONT, fontSize=8.5, alignment=1, textColor=DARK)),
                        Paragraph(sig_obj.signed_at.strftime("%Y/%m/%d") if sig_obj.signed_at else "",
                            ParagraphStyle("sd", fontName=LATIN_FONT, fontSize=8, alignment=1,
                                           textColor=colors.HexColor("#777")))]
            except Exception:
                pass
        return [Paragraph("\n\n\n", styles["Normal"]),
                Paragraph(ar(label_ar), ParagraphStyle("s_lbl", fontName=ARABIC_FONT, fontSize=8.5,
                             alignment=1, textColor=colors.HexColor("#aaa"))),
                Paragraph(label_en, ParagraphStyle("s_lble", fontName=LATIN_FONT, fontSize=7.5,
                             alignment=1, textColor=colors.HexColor("#ccc")))]

    emp_col = sig_cell(sig, "توقيع الموظف", "Employee Signature", emp.name if emp else "")
    sup_col = sig_cell(None, "توقيع المشرف", "Supervisor Signature")
    mgr_col = sig_cell(None, "توقيع الإدارة", "Management Signature")

    sig_tbl = Table([[emp_col, sup_col, mgr_col]], colWidths=[5.3*cm, 5.3*cm, 5.3*cm])
    sig_tbl.setStyle(TableStyle([
        ("BOX",(0,0),(0,0),1,BLUE),("BOX",(1,0),(1,0),1,colors.HexColor("#cccccc")),
        ("BOX",(2,0),(2,0),1,colors.HexColor("#cccccc")),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"BOTTOM"),
        ("TOPPADDING",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("BACKGROUND",(0,0),(0,0),LIGHT),
    ]))
    elements.append(sig_tbl)
    elements.append(Spacer(1, 0.4*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    from datetime import datetime as dt_cls
    elements.append(Paragraph(
        f"Generated: {dt_cls.utcnow().strftime('%Y-%m-%d')}  |  NSH SafeTrack",
        ParagraphStyle("ft", fontName=LATIN_FONT, fontSize=7.5, alignment=1,
                       textColor=colors.HexColor("#aaaaaa"))))
    doc.build(elements)
    return fname


# ═══════════════════════════════════════════════════════════════
#  API Routes — كلها تبدأ بـ /api/
# ═══════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────
#  1. تسجيل الدخول
# ─────────────────────────────────────────────
# Rate limiting بسيط للـ login — بدون مكتبات خارجية
_login_attempts: dict = {}  # {ip: [timestamps]}
_LOGIN_MAX    = 10   # محاولات
_LOGIN_WINDOW = 60   # ثانية
_LOGIN_CLEANUP_INTERVAL = 300  # تنظيف كل 5 دقائق
_last_cleanup = 0.0

def _is_rate_limited(ip: str) -> bool:
    import time
    global _last_cleanup
    now = time.time()

    # تنظيف دوري لمنع تسرب الذاكرة
    if now - _last_cleanup > _LOGIN_CLEANUP_INTERVAL:
        stale_ips = [k for k, v in _login_attempts.items()
                     if all(now - t >= _LOGIN_WINDOW for t in v)]
        for k in stale_ips:
            del _login_attempts[k]
        _last_cleanup = now

    attempts = _login_attempts.get(ip, [])
    attempts = [t for t in attempts if now - t < _LOGIN_WINDOW]
    if len(attempts) >= _LOGIN_MAX:
        _login_attempts[ip] = attempts
        return True
    attempts.append(now)
    _login_attempts[ip] = attempts
    return False

@app.post("/api/register")
def api_register():
    """Public company registration from mobile app — creates pending company + admin user."""
    data     = request.get_json(force=True) or {}
    name     = (data.get("name") or "").strip()
    email    = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()
    if not name or not email or not password:
        return jsonify({"error": "جميع الحقول مطلوبة"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "البريد الإلكتروني مسجّل مسبقاً"}), 400
    import re, unicodedata
    slug_base = re.sub(r"[^a-z0-9]+", "-",
        unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode().lower()
    ).strip("-") or "co"
    slug = slug_base
    i = 1
    while Company.query.filter_by(slug=slug).first():
        slug = f"{slug_base}-{i}"; i += 1
    co = Company(name=name, slug=slug, plan="free", is_active=False, owner_email=email)
    db.session.add(co)
    db.session.flush()
    admin_u = User(
        name=name + " Admin",
        supervisor_code=str(co.id) + "adm",
        role="admin", is_active=True,
        company_id=co.id, email=email
    )
    admin_u.set_password(password)
    db.session.add(admin_u)
    db.session.commit()
    return jsonify({"ok": True, "company_id": co.id}), 201


@app.post("/api/login")
@_login_limit
def api_login():
    # فحص rate limiting
    client_ip = freq.headers.get("X-Forwarded-For", freq.remote_addr or "unknown").split(",")[0].strip()
    if _is_rate_limited(client_ip):
        return jsonify({"error": "Too many attempts, please wait"}), 429

    data = freq.get_json(force=True) or {}
    code         = (data.get("code") or "").strip()
    email        = (data.get("email") or "").strip().lower()
    password     = (data.get("password") or "").strip()
    device_token = (data.get("device_token") or "").strip()

    user = None
    if email:
        user = User.query.filter_by(email=email, is_active=True).first()
        if user and not user.check_password(password):
            return jsonify({"error": "Invalid email or password"}), 401
        if not user:
            return jsonify({"error": "Account not found or inactive"}), 401
    elif code:
        user = User.query.filter_by(supervisor_code=code, is_active=True).first()
        if not user:
            return jsonify({"error": "ID not found or inactive"}), 401
    else:
        return jsonify({"error": "email or code is required"}), 400

    # Check company active (skip for super_admin)
    if user.role != "super_admin" and user.company_id:
        co = db.session.get(Company, user.company_id)
        if co and not co.is_active:
            return jsonify({"error": "Company account pending approval"}), 403

    # حذف tokens القديمة لهذا المستخدم (أقدم من 30 يوم) لمنع التراكم
    cutoff = datetime.now(timezone.utc) - timedelta(days=30)
    MobileToken.query.filter(
        MobileToken.user_id == user.id,
        MobileToken.created_at < cutoff
    ).delete(synchronize_session=False)

    # أنشئ توكن جلسة جديد
    token_str = secrets.token_hex(32)
    tok = MobileToken(user_id=user.id, token=token_str,
                      device_token=device_token or None,
                      company_id=user.company_id)
    db.session.add(tok)
    db.session.commit()

    co = db.session.get(Company, user.company_id) if user.company_id else None
    return jsonify({
        "token": token_str,
        "user": {
            "id":           user.id,
            "name":         user.name,
            "code":         user.supervisor_code,
            "role":         user.role,
            "hse_access":   is_hse_supervisor(user),
            "company_id":   user.company_id,
            "company_name": co.name if co else None,
        }
    })


# ─────────────────────────────────────────────
#  2. تسجيل الخروج
# ─────────────────────────────────────────────
@app.post("/api/logout")
@api_login_required
def api_logout():
    auth = freq.headers.get("Authorization", "")[7:]
    MobileToken.query.filter_by(token=auth).delete()
    db.session.commit()
    return jsonify({"message": "logged out"})


# ─────────────────────────────────────────────
#  3. بيانات المستخدم الحالي
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
#  badge count — لتحديث أيقونة التطبيق
# ─────────────────────────────────────────────
@app.get("/api/badge")
@api_login_required
def api_badge_count():
    u = get_api_user()
    cid = api_cid()
    if u.role == "admin":
        q = Request.query.filter_by(status="pending")
        if cid:
            q = q.filter_by(company_id=cid)
        count = q.count()
    elif u.role == "hr":
        q = HRTask.query.filter_by(status="pending")
        if cid:
            q = q.filter_by(company_id=cid)
        count = q.count()
    else:
        count = Request.query.filter_by(supervisor_id=u.id, status="pending").count()
    return jsonify({"badge": count})

@app.get("/api/me")
@api_login_required
def api_me():
    u = get_api_user()
    return jsonify({
        "id":   u.id,
        "name": u.name,
        "code": u.supervisor_code,
        "role": u.role,
    })


# ─────────────────────────────────────────────
#  4. قائمة الموظفين (للسوبرفايزر)
# ─────────────────────────────────────────────
@app.get("/api/employees")
@api_login_required
def api_employees():
    u = get_api_user()
    if u.role == "admin":
        emps = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    else:
        emps = Employee.query.filter_by(user_id=u.id, is_active=True).order_by(Employee.name).all()

    return jsonify([{
        "id":         e.id,
        "name":       e.name,
        "emp_number": e.emp_number,
        "department": e.department,
        "site":       e.site,
    } for e in emps])


# ─────────────────────────────────────────────
#  5. ملخص موظف (حضور + متوسط تقييم)
# ─────────────────────────────────────────────
@app.get("/api/employees/<int:emp_id>/summary")
@api_login_required
def api_employee_summary(emp_id):
    u = get_api_user()
    emp = (Employee.query.get_or_404(emp_id) if u.role == "admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first_or_404())
    summary = get_employee_summary(emp_id)
    return jsonify({
        "employee": {"id": emp.id, "name": emp.name, "emp_number": emp.emp_number},
        "summary":  summary,
    })


# ─────────────────────────────────────────────
#  6. رفع طلب جديد (إجازة أو سكليف) مع PDF اختياري
# ─────────────────────────────────────────────
@app.post("/api/requests/new")
@api_login_required
def api_request_new():
    u = get_api_user()

    # البيانات تأتي إما JSON أو Form (لو فيه ملف)
    if freq.content_type and "multipart" in freq.content_type:
        emp_id     = freq.form.get("employee_id")
        rtype      = (freq.form.get("type") or "leave").strip().lower()
        reason     = (freq.form.get("reason") or "").strip()
        start_date = freq.form.get("start_date")
        end_date   = freq.form.get("end_date")
    else:
        data       = freq.get_json(force=True) or {}
        emp_id     = data.get("employee_id")
        rtype      = (data.get("type") or "leave").strip().lower()
        reason     = (data.get("reason") or "").strip()
        start_date = data.get("start_date")
        end_date   = data.get("end_date")

    # تحقق من الموظف
    try:
        emp_id = int(emp_id or 0)
    except Exception:
        emp_id = 0

    emp = db.session.get(Employee, emp_id) if emp_id else None
    if not emp or (u.role != "admin" and emp.user_id != u.id):
        return jsonify({"error": "Invalid employee"}), 400

    # تحويل التواريخ
    def _pd(s):
        try:
            return parse_date(s) if s else None
        except Exception:
            return None

    r = Request(
        supervisor_id=u.id,
        employee_id=emp.id,
        type=rtype if rtype in {"leave", "sick", "permission", "warning", "late", "other", "secondment"} else "other",
        start_date=_pd(start_date),
        end_date=_pd(end_date),
        reason=reason,
        status="pending",
        company_id=api_cid(),
    )
    db.session.add(r)
    db.session.flush()   # نحتاج r.id قبل الحفظ النهائي

    # ─── رفع ملف PDF (إن وُجد) ───
    uploaded_file = freq.files.get("pdf_file")
    if uploaded_file and allowed_file(uploaded_file.filename):
        original_name = secure_filename(uploaded_file.filename)
        ext           = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else "pdf"
        saved_name    = f"req_{r.id}_{uuid.uuid4().hex[:8]}.{ext}"
        uploaded_file.save(os.path.join(UPLOAD_FOLDER, saved_name))

        att = RequestAttachment(
            request_id=r.id,
            filename=saved_name,
            original_name=original_name,
        )
        db.session.add(att)

    db.session.commit()

    # ── إشعار لجميع الأدمن ──
    threading.Thread(target=_notify_admins_new_request,
                     args=(r.id, emp.name, u.name), daemon=True).start()

    return jsonify({
        "message":    "Request submitted",
        "request_id": r.id,
        "status":     r.status,
    }), 201


# ─────────────────────────────────────────────
#  7. طلباتي (السوبرفايزر يشوف طلباته)
# ─────────────────────────────────────────────
@app.get("/api/requests/mine")
@api_login_required
def api_requests_mine():
    u = get_api_user()
    from sqlalchemy.orm import joinedload
    reqs = (Request.query
            .options(joinedload(Request.employee), joinedload(Request.attachments))
            .filter_by(supervisor_id=u.id)
            .order_by(Request.created_at.desc())
            .all())

    result = []
    for r in reqs:
        emp = r.employee
        try:
            _task     = HRTask.query.filter_by(request_id=r.id).first()
            _official = _task.official_reason if _task and r.type == "warning" else None
        except Exception:
            _official = None
        # نموذج الإجازة
        _leave_form = LeaveForm.query.filter_by(request_id=r.id).first() if r.type == "leave" else None
        _leave_sig  = LeaveSignature.query.filter_by(request_id=r.id).first() if r.type == "leave" else None
        _perm_sig   = PermissionSignature.query.filter_by(request_id=r.id).first() if r.type == "permission" else None
        _warn_sig   = WarningSignature.query.filter_by(request_id=r.id).first() if r.type == "warning" else None
        result.append({
            "id":            r.id,
            "type":          r.type,
            "status":        r.status,
            "start_date":    str(r.start_date) if r.start_date else None,
            "end_date":      str(r.end_date)   if r.end_date   else None,
            "reason":        r.reason,
            "created_at":    r.created_at.isoformat() if r.created_at else None,
            "admin_comment": r.admin_comment,
            "employee": {
                "id":   emp.id   if emp else None,
                "name": emp.name if emp else None,
            },
            "has_attachment":  len(r.attachments) > 0,
            "official_reason": _official,
            # إجازة
            "has_leave_form":  _leave_form is not None,
            "leave_signed":    _leave_form.signed if _leave_form else False,
            "leave_pdf_url":   f"/api/leave-form/{r.id}/download" if _leave_form else None,
            # استئذان
            "permission_signed": _perm_sig is not None,
            # إنذار
            "warning_signed":    _warn_sig is not None,
            "warning_pdf_url":   f"/api/warning/{r.id}/download" if (_task and _task.warning_pdf) else None,
        })

    return jsonify(result)


# ─────────────────────────────────────────────
#  8. صندوق طلبات الأدمن/المانجر
# ─────────────────────────────────────────────
@app.get("/api/admin/requests")
@api_admin_required
def api_admin_requests():
    status = freq.args.get("status", "pending")
    q_str  = (freq.args.get("q") or "").strip()

    from sqlalchemy.orm import joinedload
    q = Request.query.options(
        joinedload(Request.employee),
        joinedload(Request.supervisor),
        joinedload(Request.attachments),
    )
    if status and status != "all":
        q = q.filter(Request.status == status)
    if q_str:
        q = (q.join(Employee, Request.employee_id == Employee.id)
               .filter(Employee.name.ilike(f"%{q_str}%")))

    reqs = q.order_by(Request.created_at.desc()).all()

    result = []
    for r in reqs:
        emp = r.employee
        sup = r.supervisor
        result.append({
            "id":           r.id,
            "type":         r.type,
            "status":       r.status,
            "start_date":   str(r.start_date) if r.start_date else None,
            "end_date":     str(r.end_date)   if r.end_date   else None,
            "reason":       r.reason,
            "created_at":   r.created_at.isoformat() if r.created_at else None,
            "admin_comment": r.admin_comment,
            "employee": {
                "id":         emp.id         if emp else None,
                "name":       emp.name       if emp else None,
                "emp_number": emp.emp_number if emp else None,
                "department": emp.department if emp else None,
            },
            "supervisor": {
                "id":   sup.id   if sup else None,
                "name": sup.name if sup else None,
            },
            "has_attachment": len(r.attachments) > 0,
        })

    return jsonify(result)


# ─────────────────────────────────────────────
#  9. قرار الأدمن على الطلب (قبول / رفض)
# ─────────────────────────────────────────────
@app.post("/api/admin/requests/<int:req_id>/decide")
@api_admin_required
def api_request_decide(req_id):
    r = Request.query.get_or_404(req_id)
    data     = freq.get_json(force=True) or {}
    decision = data.get("decision")   # "approve" أو "reject"
    comment  = (data.get("comment") or "").strip()

    # نقبل approve/approved و reject/rejected
    if decision in ("approved",): decision = "approve"
    if decision in ("rejected",): decision = "reject"
    if decision not in ("approve", "reject"):
        return jsonify({"error": "decision must be approve or reject"}), 400

    u = get_api_user()
    r.status       = "approved" if decision == "approve" else "rejected"
    r.decided_by   = u.id
    r.decided_at   = datetime.now(timezone.utc)
    r.admin_comment = comment

    leave_pdf_url = None

    if decision == "approve":
        # إجازة → تسجيل حضور + نموذج PDF
        if r.type == "leave":
            _apply_leave_attendance_for_request(r)
            existing_form = LeaveForm.query.filter_by(request_id=r.id).first()
            if not existing_form:
                pdf_name = generate_leave_pdf(r)
                if pdf_name:
                    lf = LeaveForm(request_id=r.id, filename=pdf_name)
                    db.session.add(lf)
                    leave_pdf_url = f"/api/leave-form/{r.id}/download"

        # سكليف → تسجيل غياب تلقائي
        if r.type == "sick":
            _apply_sick_attendance_for_request(r)

        # إنشاء مهمة HR لجميع الأنواع
        existing = HRTask.query.filter_by(request_id=r.id).first()
        if not existing:
            db.session.add(HRTask(
                request_id=r.id,
                employee_id=r.employee_id,
                type=r.type or "leave",
                status="pending",
                company_id=r.company_id,
            ))

    db.session.commit()

    # ─── إشعار للسوبرفايزر ───
    type_labels = {"leave":"إجازة","sick":"سكليف","permission":"استئذان",
                   "warning":"إنذار","secondment":"إعارة","late":"تأخر"}
    type_ar = type_labels.get(r.type, r.type or "")
    _send_push_to_user(r.supervisor_id,
                       title="تحديث حالة الطلب",
                       body=f"طلب {type_ar} للموظف {r.employee.name if r.employee else ''} — {'مقبول' if r.status=='approved' else 'مرفوض'}")

    return jsonify({
        "message":       "Decision recorded",
        "status":        r.status,
        "leave_pdf_url": leave_pdf_url,
    })


# ─────────────────────────────────────────────
#  10. تحميل مرفق PDF الطلب
# ─────────────────────────────────────────────
@app.get("/api/requests/<int:req_id>/attachment")
@api_login_required
def api_request_attachment(req_id):
    from flask import send_from_directory
    r   = Request.query.get_or_404(req_id)
    att = RequestAttachment.query.filter_by(request_id=req_id).first()
    if not att:
        return jsonify({"error": "No attachment found"}), 404
    return send_from_directory(UPLOAD_FOLDER, att.filename,
                               as_attachment=True,
                               download_name=att.original_name or att.filename)


# ─────────────────────────────────────────────
#  11. تحميل نموذج الإجازة (للسوبرفايزر)
# ─────────────────────────────────────────────
@app.get("/api/leave-form/<int:req_id>/download")
@api_login_required
def api_leave_form_download(req_id):
    from flask import send_from_directory
    lf = LeaveForm.query.filter_by(request_id=req_id).first()
    if not lf:
        return jsonify({"error": "Leave form not generated yet"}), 404
    return send_from_directory(UPLOAD_FOLDER, lf.filename,
                               as_attachment=True,
                               download_name=f"leave_form_{req_id}.pdf")


# ─────────────────────────────────────────────
#  12. صندوق الموارد البشرية
# ─────────────────────────────────────────────
@app.get("/api/hr/inbox")
@api_hr_required
def api_hr_inbox():
    from sqlalchemy.orm import joinedload
    tasks = (HRTask.query
             .options(joinedload(HRTask.employee), joinedload(HRTask.request))
             .order_by(HRTask.created_at.desc()).all())
    result = []
    for t in tasks:
        r   = t.request
        emp = t.employee
        try:
            _warning_pdf     = t.warning_pdf
            _official_reason = t.official_reason
        except Exception:
            _warning_pdf     = None
            _official_reason = None
        result.append({
            "id":         t.id,
            "type":       t.type,
            "status":     t.status,
            "created_at": t.created_at.isoformat() if t.created_at else None,
            "applied_at": t.applied_at.isoformat() if t.applied_at else None,
            "employee": {
                "id":         emp.id         if emp else None,
                "name":       emp.name       if emp else None,
                "emp_number": emp.emp_number if emp else None,
            },
            "request": {
                "id":         r.id         if r else None,
                "start_date": str(r.start_date) if r and r.start_date else None,
                "end_date":   str(r.end_date)   if r and r.end_date   else None,
                "reason":     r.reason          if r else None,
            },
            "has_leave_form": (r and r.type == "leave" and
                               LeaveForm.query.filter_by(request_id=r.id).first() is not None)
                               if r else False,
            "warning_pdf_url":        f"/api/warning/{r.id}/download" if (r and _warning_pdf) else None,
            "official_reason":        _official_reason,
            "supervisor_description": r.reason if r and t.type == "warning" else None,
            "reason_set":             _official_reason is not None,
        })
    return jsonify(result)


# ─────────────────────────────────────────────
#  HR يختار السبب الرسمي للإنذار
# ─────────────────────────────────────────────
@app.post("/api/hr/tasks/<int:task_id>/set-reason")
@api_hr_required
def api_hr_set_warning_reason(task_id):
    t    = HRTask.query.get_or_404(task_id)
    data = freq.get_json(force=True) or {}
    reason = (data.get("official_reason") or "").strip()

    if not reason:
        return jsonify(error="السبب مطلوب"), 400
    if t.type != "warning":
        return jsonify(error="هذه المهمة ليست إنذاراً"), 400

    try:
        t.official_reason = reason
    except Exception:
        from sqlalchemy import text as _text
        with db.engine.connect() as _conn:
            _conn.execute(_text("ALTER TABLE hr_task ADD COLUMN IF NOT EXISTS official_reason TEXT NULL"))
            _conn.execute(_text("ALTER TABLE hr_task ADD COLUMN IF NOT EXISTS warning_pdf VARCHAR(255) NULL"))
            _conn.commit()
        t.official_reason = reason
    # لا نحدث request.reason — يبقى وصف المشرف كما هو
    db.session.commit()

    # إشعار للمشرف
    if t.request:
        threading.Thread(
            target=_send_push_to_user,
            args=(t.request.supervisor_id,
                  "إنذار جاهز للتوقيع",
                  f"نموذج إنذار الموظف {t.employee.name if t.employee else ''} جاهز"),
            daemon=True
        ).start()

    return jsonify(message="تم إرسال السبب الرسمي للمشرف", request_id=t.request_id), 200


# ─────────────────────────────────────────────
#  13. تنفيذ مهمة HR (تحويل إلى "تم التنفيذ")
# ─────────────────────────────────────────────
@app.post("/api/hr/tasks/<int:task_id>/apply")
@api_hr_required
def api_hr_task_apply(task_id):
    t = HRTask.query.get_or_404(task_id)
    if t.status == "pending":
        u = get_api_user()
        t.status     = "applied"
        t.applied_at = datetime.now(timezone.utc)
        t.applied_by = u.id
        db.session.commit()

        # إشعار للسوبرفايزر
        if t.request:
            _send_push_to_user(t.request.supervisor_id,
                               title="تم تنفيذ الطلب",
                               body=f"تم تنفيذ طلب {t.type} للموظف {t.employee.name if t.employee else ''}")

    return jsonify({"message": "Task applied", "status": t.status})


# ─────────────────────────────────────────────
#  14. لوحة KPI للأدمن (ملخص سريع)
# ─────────────────────────────────────────────
@app.get("/api/admin/kpi")
@api_admin_required
def api_admin_kpi():
    ws, we = default_week_today()
    ws_str = freq.args.get("week_start")
    we_str = freq.args.get("week_end")
    if ws_str and we_str:
        try:
            ws = parse_date(ws_str)
            we = parse_date(we_str)
        except Exception:
            pass

    pws, pwe = previous_week_range(ws)

    # ── إحصائيات أساسية ──
    total_employees  = Employee.query.filter_by(is_active=True).count()
    total_evals_week = Evaluation.query.filter_by(week_start=ws, week_end=we).count()
    total_evals_prev = Evaluation.query.filter_by(week_start=pws, week_end=pwe).count()
    pending_requests = Request.query.filter_by(status="pending").count()
    hr_pending       = HRTask.query.filter_by(status="pending").count()

    # ── تغطية الموظفين ──
    sum_target_emp = Employee.query.filter_by(is_active=True).count()
    sum_done_emp   = db.session.query(func.count(func.distinct(Evaluation.employee_id)))\
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we).scalar() or 0
    emp_coverage = round(sum_done_emp / sum_target_emp * 100.0, 1) if sum_target_emp else 0.0

    # ── معدل الحضور ──
    active_emp_ids = [e.id for e in Employee.query.filter_by(is_active=True).with_entities(Employee.id).all()]
    workdays = (we - ws).days + 1
    total_opp = len(active_emp_ids) * workdays

    present_count = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date >= ws, Attendance.date <= we,
        Attendance.employee_id.in_(active_emp_ids),
        Attendance.status == 'present'
    ).scalar() or 0

    leave_count = db.session.query(func.count(Attendance.id)).filter(
        Attendance.date >= ws, Attendance.date <= we,
        Attendance.employee_id.in_(active_emp_ids),
        Attendance.status == 'leave'
    ).scalar() or 0

    att_rate   = round(present_count / total_opp * 100.0, 1) if total_opp else 0.0
    leave_rate = round(leave_count   / total_opp * 100.0, 1) if total_opp else 0.0

    return jsonify({
        "week_start":       ws.isoformat(),
        "week_end":         we.isoformat(),
        "total_employees":  total_employees,
        "evals_this_week":  total_evals_week,
        "evals_last_week":  total_evals_prev,
        "pending_requests": pending_requests,
        "hr_pending_tasks": hr_pending,
        "emp_coverage":     emp_coverage,
        "att_rate":         att_rate,
        "leave_rate":       leave_rate,
        "present_count":    present_count,
        "leave_count":      leave_count,
    })



# ─────────────────────────────────────────────
#  مقارنة المشرفين
# ─────────────────────────────────────────────
@app.get("/api/admin/supervisors/ranking")
@api_admin_required
def api_supervisors_ranking():
    ws, we = default_week_today()
    ws_str = freq.args.get("week_start")
    we_str = freq.args.get("week_end")
    if ws_str and we_str:
        try: ws = parse_date(ws_str); we = parse_date(we_str)
        except: pass

    pws, pwe = previous_week_range(ws)

    supervisors = User.query.filter(
        User.role == "supervisor",
        User.is_active == True,
        db.or_(User.is_hidden == False, User.is_hidden == None)
    ).order_by(User.supervisor_code.asc()).all()

    emp_totals = dict(
        db.session.query(Employee.user_id, func.count(Employee.id))
        .filter(Employee.is_active == True)
        .group_by(Employee.user_id).all()
    )
    emp_done_this = dict(
        db.session.query(Employee.user_id, func.count(Evaluation.id))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
        .group_by(Employee.user_id).all()
    )
    emp_done_prev = dict(
        db.session.query(Employee.user_id, func.count(Evaluation.id))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == pws, Evaluation.week_end == pwe)
        .group_by(Employee.user_id).all()
    )
    avg_scores = dict(
        db.session.query(Employee.user_id, func.avg(Evaluation.total_score))
        .join(Evaluation, Evaluation.employee_id == Employee.id)
        .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
        .group_by(Employee.user_id).all()
    )

    rows = []
    for sup in supervisors:
        target   = emp_totals.get(sup.id, 0)
        done_c   = emp_done_this.get(sup.id, 0)
        done_p   = emp_done_prev.get(sup.id, 0)
        avg      = avg_scores.get(sup.id)
        coverage = round(done_c / target * 100.0, 1) if target else 0.0
        delta    = done_c - done_p
        trend    = "up" if delta > 0 else "down" if delta < 0 else "flat"
        rows.append({
            "id":         sup.id,
            "name":       sup.name,
            "code":       sup.supervisor_code,
            "target":     target,
            "done":       done_c,
            "done_prev":  done_p,
            "coverage":   coverage,
            "avg_score":  round(float(avg), 1) if avg else None,
            "delta":      delta,
            "trend":      trend,
        })

    rows.sort(key=lambda x: x["coverage"], reverse=True)
    return jsonify({"week_start": ws.isoformat(), "week_end": we.isoformat(), "rows": rows})


# ─────────────────────────────────────────────
#  تقرير مشرف محدد
# ─────────────────────────────────────────────
@app.get("/api/admin/supervisors/<int:sup_id>/report")
@api_admin_required
def api_supervisor_report(sup_id):
    sup = User.query.get_or_404(sup_id)
    ws, we = default_week_today()
    ws_str = freq.args.get("week_start")
    we_str = freq.args.get("week_end")
    if ws_str and we_str:
        try: ws = parse_date(ws_str); we = parse_date(we_str)
        except: pass

    pws, pwe = previous_week_range(ws)

    # موظفو المشرف
    employees = Employee.query.filter_by(user_id=sup_id, is_active=True).order_by(Employee.name).all()
    emp_ids   = [e.id for e in employees]

    # تقييمات الأسبوع
    evals = {ev.employee_id: ev for ev in
             Evaluation.query.filter(
                 Evaluation.employee_id.in_(emp_ids),
                 Evaluation.week_start == ws,
                 Evaluation.week_end   == we).all()}

    # حضور الأسبوع
    att_records = Attendance.query.filter(
        Attendance.employee_id.in_(emp_ids),
        Attendance.date >= ws,
        Attendance.date <= we).all()
    att_map = {}
    for a in att_records:
        att_map.setdefault(a.employee_id, []).append({
            "date": str(a.date), "status": a.status})

    # إحصائيات سريعة
    total      = len(employees)
    evaluated  = len(evals)
    avg_score  = round(sum(ev.total_score or 0 for ev in evals.values()) / evaluated, 1) if evaluated else None
    present_cnt = sum(1 for a in att_records if a.status == "present")

    emp_rows = []
    for e in employees:
        ev = evals.get(e.id)
        emp_rows.append({
            "emp_id":     e.id,
            "emp_number": e.emp_number,
            "name":       e.name,
            "department": e.department,
            "evaluated":  ev is not None,
            "total":      ev.total_score  if ev else None,
            "band":       ev.overall_band if ev else None,
            "att":        att_map.get(e.id, []),
        })

    return jsonify({
        "supervisor":  {"id": sup.id, "name": sup.name, "code": sup.supervisor_code},
        "week_start":  ws.isoformat(),
        "week_end":    we.isoformat(),
        "total_emp":   total,
        "evaluated":   evaluated,
        "avg_score":   avg_score,
        "present_cnt": present_cnt,
        "employees":   emp_rows,
    })



@app.get("/api/weekly-eval/<int:ev_id>")
@api_login_required
def api_weekly_eval_detail(ev_id):
    u  = get_api_user()
    ev = Evaluation.query.get_or_404(ev_id)
    emp = Employee.query.get_or_404(ev.employee_id)
    if u.role != "admin" and emp.user_id != u.id:
        return jsonify({"error": "Forbidden"}), 403
    return jsonify({
        "t1_text": ev.t1_text, "t1_percent": ev.t1_percent, "t1_remarks": ev.t1_remarks,
        "t2_text": ev.t2_text, "t2_percent": ev.t2_percent, "t2_remarks": ev.t2_remarks,
        "t3_text": ev.t3_text, "t3_percent": ev.t3_percent, "t3_remarks": ev.t3_remarks,
        "t4_text": ev.t4_text, "t4_percent": ev.t4_percent, "t4_remarks": ev.t4_remarks,
        "p_punctuality":   ev.p_punctuality,   "c_punctuality":   ev.c_punctuality,
        "p_quality":       ev.p_quality,       "c_quality":       ev.c_quality,
        "p_productivity":  ev.p_productivity,  "c_productivity":  ev.c_productivity,
        "p_communication": ev.p_communication, "c_communication": ev.c_communication,
        "p_problemsolving":ev.p_problemsolving,"c_problemsolving":ev.c_problemsolving,
        "p_compliance":    ev.p_compliance,    "c_compliance":    ev.c_compliance,
        "strengths":       ev.strengths,
        "improvements":    ev.improvements,
        "training_needed": ev.training_needed,
        "week_start":      ev.week_start.isoformat(),
        "week_end":        ev.week_end.isoformat(),
    })

# ─────────────────────────────────────────────
#  تعديل التقييم الأسبوعي
# ─────────────────────────────────────────────
@app.put("/api/weekly-eval/<int:ev_id>")
@api_login_required
def api_weekly_eval_edit(ev_id):
    u  = get_api_user()
    ev = Evaluation.query.get_or_404(ev_id)
    emp = Employee.query.get_or_404(ev.employee_id)

    if u.role != "admin" and emp.user_id != u.id:
        return jsonify({"error": "Forbidden"}), 403

    data = freq.get_json(force=True) or {}

    # تحقق من الأسبوع إذا تغيّر
    ws_str = data.get("week_start")
    we_str = data.get("week_end")
    if ws_str and we_str:
        try:
            ws = parse_date(ws_str)
            we = parse_date(we_str)
        except:
            return jsonify({"error": "Invalid dates"}), 400
        ok, msg = validate_week_sun_to_thu(ws, we)
        if not ok:
            return jsonify({"error": msg}), 400
        dup = Evaluation.query.filter(
            Evaluation.employee_id == emp.id,
            Evaluation.week_start  == ws,
            Evaluation.week_end    == we,
            Evaluation.id          != ev.id).first()
        if dup:
            return jsonify({"error": "تقييم لهذا الأسبوع موجود مسبقاً"}), 409
        ev.week_start = ws
        ev.week_end   = we

    ev.t1_text       = data.get("t1_text",       ev.t1_text)
    ev.t2_text       = data.get("t2_text",       ev.t2_text)
    ev.t3_text       = data.get("t3_text",       ev.t3_text)
    ev.t4_text       = data.get("t4_text",       ev.t4_text)
    ev.t1_percent    = data.get("t1_percent",    ev.t1_percent)
    ev.t2_percent    = data.get("t2_percent",    ev.t2_percent)
    ev.t3_percent    = data.get("t3_percent",    ev.t3_percent)
    ev.t4_percent    = data.get("t4_percent",    ev.t4_percent)
    ev.t1_remarks    = data.get("t1_remarks",    ev.t1_remarks)
    ev.t2_remarks    = data.get("t2_remarks",    ev.t2_remarks)
    ev.t3_remarks    = data.get("t3_remarks",    ev.t3_remarks)
    ev.t4_remarks    = data.get("t4_remarks",    ev.t4_remarks)
    ev.p_punctuality   = data.get("p_punctuality",   ev.p_punctuality)
    ev.p_quality       = data.get("p_quality",       ev.p_quality)
    ev.p_productivity  = data.get("p_productivity",  ev.p_productivity)
    ev.p_communication = data.get("p_communication", ev.p_communication)
    ev.p_problemsolving= data.get("p_problemsolving",ev.p_problemsolving)
    ev.p_compliance    = data.get("p_compliance",    ev.p_compliance)
    ev.c_punctuality   = data.get("c_punctuality",   ev.c_punctuality)
    ev.c_quality       = data.get("c_quality",       ev.c_quality)
    ev.c_productivity  = data.get("c_productivity",  ev.c_productivity)
    ev.c_communication = data.get("c_communication", ev.c_communication)
    ev.c_problemsolving= data.get("c_problemsolving",ev.c_problemsolving)
    ev.c_compliance    = data.get("c_compliance",    ev.c_compliance)
    ev.strengths       = data.get("strengths",       ev.strengths)
    ev.improvements    = data.get("improvements",    ev.improvements)
    ev.training_needed = data.get("training_needed", ev.training_needed)

    compute_scores(ev)
    db.session.commit()
    return jsonify({"message": "Updated", "total": ev.total_score, "band": ev.overall_band})

# ─────────────────────────────────────────────
#  15. تسجيل توكن الإشعارات (APNs)
# ─────────────────────────────────────────────
@app.post("/api/device-token")
@api_login_required
def api_register_device_token():
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    device_token = (data.get("device_token") or "").strip()
    if not device_token:
        return jsonify({"error": "device_token required"}), 400

    auth_token = freq.headers.get("Authorization", "")[7:]
    tok = MobileToken.query.filter_by(token=auth_token).first()
    if tok:
        tok.device_token = device_token
        db.session.commit()

    return jsonify({"message": "Device token registered"})


# ─────────────────────────────────────────────
#  مساعد إشعار الأدمن عند طلب جديد
# ─────────────────────────────────────────────
def _notify_admins_new_request(req_id: int, emp_name: str, sup_name: str):
    with app.app_context():
        admins = User.query.filter_by(role="admin", is_active=True).all()
        for admin in admins:
            _send_push_to_user(
                admin.id,
                title="طلب جديد 📋",
                body=f"{sup_name} رفع طلباً للموظف {emp_name}",
            )


# ─────────────────────────────────────────────
#  مساعد إرسال الإشعارات (APNs)
#  — يحتاج مكتبة: pip install httpx PyJWT
# ─────────────────────────────────────────────
def _send_push_to_user(user_id: int, title: str, body: str, badge: int = None):
    with app.app_context():
        try:
            import jwt as pyjwt
            import httpx
            KEY_ID    = os.getenv("APNS_KEY_ID")    or "69QBYQL8PG"
            TEAM_ID   = os.getenv("APNS_TEAM_ID")   or "76JS5SQN7L"
            BUNDLE_ID = os.getenv("APNS_BUNDLE_ID") or "com.z.nsh"
            KEY_PATH  = os.getenv("APNS_KEY_PATH")  or os.path.join(BASE_DIR, "keys", "AuthKey_69QBYQL8PG.p8")
            if not os.path.exists(KEY_PATH):
                app.logger.error("APNs: ملف .p8 غير موجود في %s", KEY_PATH)
                return
            with open(KEY_PATH, "r") as f:
                private_key = f.read()
            token = pyjwt.encode(
                {"iss": TEAM_ID, "iat": datetime.now(timezone.utc)},
                private_key,
                algorithm="ES256",
                headers={"kid": KEY_ID},
            )
            device_tokens = [
                t.device_token for t in
                MobileToken.query.filter_by(user_id=user_id).all()
                if t.device_token
            ]
            app.logger.info("APNs: إرسال لـ user_id=%s، عدد الأجهزة=%s", user_id, len(device_tokens))
            if not device_tokens:
                app.logger.warning("APNs: لا يوجد device_token مسجل للمستخدم %s", user_id)
                return
            # حساب الـ badge الفعلي حسب دور المستخدم
            if badge is None:
                try:
                    u = db.session.get(User, user_id)
                    if u and u.role == "admin":
                        badge = Request.query.filter_by(status="pending").count()
                    elif u and u.role == "hr":
                        badge = HRTask.query.filter_by(status="pending").count()
                    elif u:
                        badge = Request.query.filter_by(supervisor_id=user_id, status="pending").count()
                    else:
                        badge = 1
                except Exception:
                    badge = 1

            for dt in device_tokens:
                payload = json.dumps({
                    "aps": {
                        "alert": {"title": title, "body": body},
                        "sound": "default",
                        "badge": badge,
                    }
                })
                hdrs = {
                    "authorization": f"bearer {token}",
                    "apns-topic":    BUNDLE_ID,
                    "apns-push-type": "alert",
                    "content-type":  "application/json",
                }
                sent = False
                for host in ["api.push.apple.com", "api.sandbox.push.apple.com"]:
                    url = f"https://{host}/3/device/{dt}"
                    with httpx.Client(http2=True, timeout=5) as client:
                        resp = client.post(url, content=payload, headers=hdrs)
                    app.logger.info("APNs: host=%s status=%s device=%s…", host, resp.status_code, dt[:20])
                    if resp.status_code == 200:
                        sent = True
                        break
                if not sent:
                    app.logger.warning("APNs: فشل الإرسال لـ device=%s…", dt[:20])
        except Exception as e:
            app.logger.error("APNs error: %s", e, exc_info=True)


# ─────────────────────────────────────────────
#  16. قائمة طلبات السوبرفايزر كاملة (للأدمن يشوف)
# ─────────────────────────────────────────────
@app.get("/api/admin/requests/supervisor/<int:sup_id>")
@api_admin_required
def api_admin_supervisor_requests(sup_id):
    reqs = (Request.query
            .filter_by(supervisor_id=sup_id)
            .order_by(Request.created_at.desc())
            .all())
    result = []
    for r in reqs:
        emp = r.employee
        result.append({
            "id":         r.id,
            "type":       r.type,
            "status":     r.status,
            "start_date": str(r.start_date) if r.start_date else None,
            "end_date":   str(r.end_date)   if r.end_date   else None,
            "reason":     r.reason,
            "employee":   {"id": emp.id, "name": emp.name} if emp else None,
        })
    return jsonify(result)


# ═══════════════════════════════════════════════════════════════
#  نهاية ملف التعديلات
# ═══════════════════════════════════════════════════════════════
# ─────────────────────────────────────────────
#  17. Allowed Users — قائمة المستخدمين
# ─────────────────────────────────────────────
@app.get("/api/admin/users")
@api_admin_required
def api_admin_users():
    cid = api_cid()
    q = User.query
    if cid:
        q = q.filter_by(company_id=cid)
    users = q.order_by(User.id.asc()).all()
    return jsonify([{
        "id":        u.id,
        "name":      u.name,
        "code":      u.supervisor_code,
        "role":      u.role,
        "is_active": u.is_active,
        "is_hidden": bool(u.is_hidden) if u.is_hidden is not None else False,
    } for u in users])


# ─────────────────────────────────────────────
#  18. Allowed Users — إضافة مستخدم
# ─────────────────────────────────────────────
@app.post("/api/admin/users/add")
@api_admin_required
def api_admin_users_add():
    data = freq.get_json(force=True) or {}

    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    role = (data.get("role") or "supervisor").strip().lower()

    if not code:
        return jsonify({"error": "code is required"}), 400

    # التحقق إن الاسم إنجليزي فقط
    if name and not all(ord(c) < 128 for c in name):
        return jsonify({"error": "الاسم يجب أن يكون بالإنجليزية فقط"}), 400

    if role not in {"admin", "hr", "supervisor"}:
        return jsonify({"error": "invalid role"}), 400

    existing = User.query.filter_by(supervisor_code=code).first()
    if existing:
        return jsonify({"error": "user code already exists"}), 400

    u = User(
        supervisor_code=code,
        name=name or code,
        role=role,
        is_active=True,
    )
    db.session.add(u)
    db.session.commit()

    return jsonify({
        "message": "user added",
        "user": {
            "id": u.id,
            "name": u.name,
            "code": u.supervisor_code,
            "role": u.role,
            "is_active": u.is_active,
        }
    }), 201


# ─────────────────────────────────────────────
#  19. Allowed Users — تعطيل مستخدم
# ─────────────────────────────────────────────
@app.post("/api/admin/users/<int:user_id>/deactivate")
@api_admin_required
def api_admin_users_deactivate(user_id):
    u = User.query.get_or_404(user_id)

    if u.role == "admin" and User.query.filter_by(role="admin", is_active=True).count() <= 1:
        return jsonify({"error": "cannot deactivate the last active admin"}), 400

    u.is_active = False
    db.session.commit()

    return jsonify({"message": "user deactivated", "id": u.id, "is_active": u.is_active})


# ─────────────────────────────────────────────
#  إخفاء/إظهار مستخدم
# ─────────────────────────────────────────────
@app.post("/api/admin/users/<int:user_id>/hide")
@api_admin_required
def api_admin_users_hide(user_id):
    u = User.query.get_or_404(user_id)
    if u.role == "admin":
        return jsonify({"error": "Cannot hide admin"}), 400
    u.is_hidden = not (u.is_hidden or False)
    db.session.commit()
    return jsonify({"message": "updated", "is_hidden": u.is_hidden})



# ─────────────────────────────────────────────
#  استقالة مشرف — يُعطَّل ويتحول موظفوه لـ unassigned
#  POST /api/admin/users/:id/resign
# ─────────────────────────────────────────────
@app.post("/api/admin/users/<int:user_id>/resign")
@api_admin_required
def api_admin_users_resign(user_id):
    u = User.query.get_or_404(user_id)

    # لا يمكن تطبيق الاستقالة على admin
    if u.role == "admin":
        return jsonify({"error": "لا يمكن تطبيق الاستقالة على حساب أدمن"}), 400

    # تعطيل الحساب وإخفاؤه
    u.is_active = False
    u.is_hidden = True

    # تحويل موظفيه لـ unassigned
    employees = Employee.query.filter_by(user_id=user_id, is_active=True).all()
    count = len(employees)
    for emp in employees:
        emp.user_id = None
        emp.status  = "unassigned"

    db.session.commit()

    app.logger.info("استقالة مشرف: user_id=%s name=%s — %s موظف تحول لـ unassigned",
                    user_id, u.name, count)

    return jsonify({
        "message":            f"تم تسجيل استقالة {u.name}",
        "user_id":            u.id,
        "employees_unassigned": count,
    })

# ─────────────────────────────────────────────
#  20. Daily Reports — تقارير اليوم المحدد
# ─────────────────────────────────────────────
@app.get("/api/admin/reports/daily")
@api_admin_required
def api_admin_reports_daily():
    date_str = freq.args.get("date")
    if not date_str:
        return jsonify({"error": "date is required"}), 400

    try:
        d = parse_date(date_str)
    except Exception:
        return jsonify({"error": "invalid date"}), 400

    rows = (DailyEvaluation.query
            .join(Employee, DailyEvaluation.employee_id == Employee.id)
            .join(User, Employee.user_id == User.id, isouter=True)
            .filter(DailyEvaluation.eval_date == d)
            .order_by(User.name.asc(), Employee.name.asc())
            .all())

    result = []
    for r in rows:
        emp = r.employee
        sup = emp.user if emp else None

        result.append({
            "id": r.id,
            "eval_date": str(r.eval_date) if r.eval_date else None,
            "supervisor": {
                "id": sup.id if sup else None,
                "name": sup.name if sup else None,
                "code": sup.supervisor_code if sup else None,
            },
            "employee": {
                "id": emp.id if emp else None,
                "name": emp.name if emp else None,
                "emp_number": emp.emp_number if emp else None,
                "department": emp.department if emp else None,
                "site": emp.site if emp else None,
            },
            "targets": r.targets_score,
            "perf":    r.performance_score,
            "total":   r.total_score,
            "band":    r.overall_band,
        })

    return jsonify(result)


# ─────────────────────────────────────────────
#  21. Supervisor Reports — أسبوعي مع فلتر
# ─────────────────────────────────────────────
@app.get("/api/admin/reports/supervisors")
@api_admin_required
def api_admin_reports_supervisors():
    ws, we = default_week_today()

    ws_str = freq.args.get("week_start")
    we_str = freq.args.get("week_end")
    sup_code = (freq.args.get("supervisor_code") or "").strip()

    if ws_str and we_str:
        try:
            ws = parse_date(ws_str)
            we = parse_date(we_str)
        except Exception:
            return jsonify({"error": "invalid week_start/week_end"}), 400

    q = (Evaluation.query
         .join(Employee, Evaluation.employee_id == Employee.id)
         .join(User, Employee.user_id == User.id)
         .filter(Evaluation.week_start == ws, Evaluation.week_end == we))

    if sup_code:
        q = q.filter(User.supervisor_code == sup_code)

    rows = q.order_by(User.name.asc(), Employee.name.asc()).all()

    result = []
    for r in rows:
        emp = r.employee
        sup = emp.user if emp else None

        result.append({
            "id": getattr(r, "id", None),
            "week_start": str(r.week_start) if r.week_start else None,
            "week_end": str(r.week_end) if r.week_end else None,
            "supervisor": {
                "id": sup.id if sup else None,
                "name": sup.name if sup else None,
                "code": sup.supervisor_code if sup else None,
            },
            "employee": {
                "id": emp.id if emp else None,
                "name": emp.name if emp else None,
                "emp_number": emp.emp_number if emp else None,
                "department": emp.department if emp else None,
                "site": emp.site if emp else None,
            },
            "targets": r.targets_score,
            "perf":    r.perf_score,
            "total":   r.total_score,
            "band":    r.overall_band,
        })

    return jsonify(result)


# ─────────────────────────────────────────────
#  22. Attendance Admin — يومي / أسبوعي
# ─────────────────────────────────────────────
@app.get("/api/admin/attendance")
@api_admin_required
def api_admin_attendance():
    mode = (freq.args.get("mode") or "day").strip().lower()
    date_str = freq.args.get("date")
    week_start_str = freq.args.get("week_start")
    week_end_str = freq.args.get("week_end")

    result = {
        "mode": mode,
        "summary": {},
        "records": []
    }

    if mode == "week":
        ws, we = default_week_today()
        if week_start_str and week_end_str:
            try:
                ws = parse_date(week_start_str)
                we = parse_date(week_end_str)
            except Exception:
                return jsonify({"error": "invalid week range"}), 400

        rows = (Attendance.query
                .join(Employee, Attendance.employee_id == Employee.id)
                .join(User, Employee.user_id == User.id, isouter=True)
                .filter(Attendance.date >= ws, Attendance.date <= we)
                .order_by(Attendance.date.desc(), Employee.name.asc())
                .all())

        present = 0
        absent = 0
        leave = 0

        for a in rows:
            status = (a.status or "").lower()
            if status == "present":
                present += 1
            elif status == "absent":
                absent += 1
            elif status == "leave":
                leave += 1

            emp = a.employee
            sup = emp.user if emp else None

            result["records"].append({
                "id": a.id,
                "date": str(a.date) if a.date else None,
                "employee": {
                    "id": emp.id if emp else None,
                    "name": emp.name if emp else None,
                    "emp_number": emp.emp_number if emp else None,
                    "department": emp.department if emp else None,
                    "site": emp.site if emp else None,
                },
                "supervisor": {
                    "id": sup.id if sup else None,
                    "name": sup.name if sup else None,
                    "code": sup.supervisor_code if sup else None,
                },
                "status": a.status,
                "remarks": getattr(a, "remarks", None),
            })

        result["summary"] = {
            "week_start": str(ws),
            "week_end": str(we),
            "present": present,
            "absent": absent,
            "leave": leave,
            "count": len(rows),
        }
        return jsonify(result)

    # day mode
    if not date_str:
        date_str = date.today().isoformat()

    try:
        d = parse_date(date_str)
    except Exception:
        return jsonify({"error": "invalid date"}), 400

    rows = (Attendance.query
            .join(Employee, Attendance.employee_id == Employee.id)
            .join(User, Employee.user_id == User.id, isouter=True)
            .filter(Attendance.date == d)
            .order_by(Employee.name.asc())
            .all())

    present = 0
    absent = 0
    leave = 0

    for a in rows:
        status = (a.status or "").lower()
        if status == "present":
            present += 1
        elif status == "absent":
            absent += 1
        elif status == "leave":
            leave += 1

        emp = a.employee
        sup = emp.user if emp else None

        result["records"].append({
            "id": a.id,
            "date": str(a.date) if a.date else None,
            "employee": {
                "id": emp.id if emp else None,
                "name": emp.name if emp else None,
                "emp_number": emp.emp_number if emp else None,
                "department": emp.department if emp else None,
                "site": emp.site if emp else None,
            },
            "supervisor": {
                "id": sup.id if sup else None,
                "name": sup.name if sup else None,
                "code": sup.supervisor_code if sup else None,
            },
            "status": a.status,
            "remarks": getattr(a, "remarks", None),
        })

    result["summary"] = {
        "date": str(d),
        "present": present,
        "absent": absent,
        "leave": leave,
        "count": len(rows),
    }
    return jsonify(result)

# ─────────────────────────────────────────────
#  ملخص الحضور الشهري للأدمن
#  GET /api/admin/attendance/monthly?year=2025&month=3
# ─────────────────────────────────────────────
@app.get("/api/admin/attendance/monthly")
@api_admin_required
def api_admin_attendance_monthly():
    from calendar import monthrange
    year_str  = freq.args.get("year")
    month_str = freq.args.get("month")

    try:
        today = datetime.now(RIYADH_TZ).date()
        year  = int(year_str)  if year_str  else today.year
        month = int(month_str) if month_str else today.month
    except (ValueError, TypeError):
        return jsonify({"error": "invalid year or month"}), 400

    # أول وآخر يوم في الشهر
    _, days_in_month = monthrange(year, month)
    month_start = date(year, month, 1)
    month_end   = date(year, month, days_in_month)

    # كل الموظفين النشطين مع بيانات المشرف دفعة واحدة
    from sqlalchemy.orm import joinedload
    emps = (Employee.query
            .filter_by(is_active=True)
            .order_by(Employee.name)
            .all())

    emp_ids = [e.id for e in emps]

    # جلب بيانات المشرفين دفعة واحدة بدل N+1
    sup_map = {u.id: u.name for u in User.query.filter(
        User.id.in_([e.user_id for e in emps if e.user_id])
    ).all()}

    # كل سجلات الحضور في الشهر دفعة واحدة
    all_att = Attendance.query.filter(
        Attendance.employee_id.in_(emp_ids),
        Attendance.date >= month_start,
        Attendance.date <= month_end
    ).all()

    # نبني dict سريع: {emp_id: {date_str: status}}
    att_map = {}
    for a in all_att:
        if a.employee_id not in att_map:
            att_map[a.employee_id] = {}
        att_map[a.employee_id][str(a.date)] = a.status

    # نحسب ملخص كل موظف
    rows = []
    total_present = total_absent = total_leave = total_days = 0

    for emp in emps:
        emp_att = att_map.get(emp.id, {})
        present  = sum(1 for s in emp_att.values() if s == "present")
        absent   = sum(1 for s in emp_att.values() if s == "absent")
        leave    = sum(1 for s in emp_att.values() if s == "leave")
        recorded = present + absent + leave

        total_present += present
        total_absent  += absent
        total_leave   += leave
        total_days    += recorded

        sup_name = sup_map.get(emp.user_id, "—") if emp.user_id else "—"

        # النسبة = حاضر ÷ أيام الشهر — يكشف من ما سُجّل حضوره أصلاً
        rate = round(present / days_in_month * 100, 1) if present > 0 else 0.0

        rows.append({
            "emp_id":     emp.id,
            "emp_number": emp.emp_number,
            "name":       emp.name,
            "department": emp.department or "",
            "site":       emp.site or "",
            "supervisor": sup_name,
            "present":    present,
            "absent":     absent,
            "leave":      leave,
            "recorded":   recorded,
            "rate":       rate,
        })

    # متوسط الحضور = حاضر ÷ كل الموظفين النشطين (يكشف من ما سُجّل حضوره)
    total_active = len(emps)
    avg_rate = round(total_present / total_active * 100, 1) if total_active > 0 else 0.0

    return jsonify({
        "year":    year,
        "month":   month,
        "days":    days_in_month,
        "summary": {
            "total_employees": len(emps),
            "total_present":   total_present,
            "total_absent":    total_absent,
            "total_leave":     total_leave,
            "avg_rate":        avg_rate,
        },
        "employees": rows,
    })


# ─────────────────────────────────────────────
#  تقرير شهري شامل للأدمن — PDF
#  GET /api/admin/monthly-report?year=2026&month=3
# ─────────────────────────────────────────────
@app.get("/api/admin/monthly-report")
@api_admin_required
def api_admin_monthly_report():
    """يولّد تقرير PDF شهري شامل: ملخص + صفحة لكل مشرف"""
    from calendar import monthrange
    from datetime import date as dt_date
    from zoneinfo import ZoneInfo

    year_str  = freq.args.get("year")
    month_str = freq.args.get("month")
    try:
        today = datetime.now(ZoneInfo("Asia/Riyadh")).date()
        year  = int(year_str)  if year_str  else today.year
        month = int(month_str) if month_str else today.month
    except (ValueError, TypeError):
        return jsonify({"error": "invalid year or month"}), 400

    report_type = (freq.args.get("type") or "detailed").strip().lower()
    # type=summary  → تقرير إجمالي للشركة (صفحة واحدة)
    # type=detailed → تقرير مفصّل (صفحة لكل مشرف)

    _, days_in_month = monthrange(year, month)
    month_start = dt_date(year, month, 1)
    month_end   = dt_date(year, month, days_in_month)

    month_names_ar = ["","يناير","فبراير","مارس","أبريل","مايو","يونيو",
                      "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]
    month_name = month_names_ar[month]

    # ── جلب البيانات ──
    supervisors = User.query.filter(
        User.role == "supervisor",
        User.is_active == True,
        User.is_hidden == False
    ).order_by(User.name).all()

    sup_ids = [s.id for s in supervisors]

    # كل الموظفين النشطين
    all_emps = Employee.query.filter_by(is_active=True).all()
    emp_by_sup = {}
    for e in all_emps:
        if e.user_id:
            emp_by_sup.setdefault(e.user_id, []).append(e)

    # كل سجلات الحضور في الشهر
    emp_ids = [e.id for e in all_emps]
    sup_map = {u.id: u.name for u in supervisors}

    all_att = Attendance.query.filter(
        Attendance.employee_id.in_(emp_ids),
        Attendance.date >= month_start,
        Attendance.date <= month_end
    ).all()
    att_map = {}
    for a in all_att:
        att_map.setdefault(a.employee_id, {})[str(a.date)] = a.status

    # كل التقييمات الأسبوعية في الشهر
    all_evals = Evaluation.query.filter(
        Evaluation.employee_id.in_(emp_ids),
        Evaluation.week_start >= month_start,
        Evaluation.week_end   <= month_end
    ).all()
    eval_map = {}
    for ev in all_evals:
        eval_map.setdefault(ev.employee_id, []).append(ev.total_score or 0)

    # ── توليد PDF ──
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return jsonify({"error": "reportlab غير مثبت"}), 500

    # تسجيل الخطوط
    ARABIC_FONT = "Arabic"
    LATIN_FONT  = "Helvetica"
    try:
        pdfmetrics.registerFont(TTFont(ARABIC_FONT,
            "/usr/share/fonts/google-droid/DroidSansArabic.ttf"))
    except Exception:
        ARABIC_FONT = LATIN_FONT

    def ar(text):
        t = str(text)
        if not any("؀" <= c <= "ۿ" for c in t):
            return t
        try:
            from arabic_reshaper import reshape
            from bidi.algorithm import get_display
            return get_display(reshape(t))
        except Exception:
            return t

    def font_for(text):
        return ARABIC_FONT if any("؀" <= c <= "ۿ" for c in str(text)) else LATIN_FONT

    fname = f"monthly_{'summary' if report_type == 'summary' else 'detailed'}_{year}_{month:02d}_{uuid.uuid4().hex[:6]}.pdf"
    fpath = os.path.join(UPLOAD_FOLDER, fname)

    doc = SimpleDocTemplate(fpath, pagesize=A4,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", fontName=ARABIC_FONT, fontSize=16,
                                  alignment=1, spaceAfter=6,
                                  textColor=colors.HexColor("#1a1f3a"))
    sub_style   = ParagraphStyle("s", fontName=ARABIC_FONT, fontSize=11,
                                  alignment=1, spaceAfter=4,
                                  textColor=colors.HexColor("#7b5ea7"))
    normal_r    = ParagraphStyle("n", fontName=ARABIC_FONT, fontSize=9,
                                  alignment=2)
    small_r     = ParagraphStyle("sm", fontName=LATIN_FONT,  fontSize=8,
                                  alignment=1)

    HDR_COLOR  = colors.HexColor("#1a1f3a")
    SUB_COLOR  = colors.HexColor("#7b5ea7")
    ROW1_COLOR = colors.HexColor("#f0f4ff")
    GREEN      = colors.HexColor("#27ae60")
    RED        = colors.HexColor("#e74c3c")
    ORANGE     = colors.HexColor("#f39c12")

    elements = []

    # ══════════════════════════════════════════
    # صفحة 1: ملخص الشركة
    # ══════════════════════════════════════════
    logo_path = os.path.join(BASE_DIR, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        from reportlab.platypus import Image as RLImage
        elements.append(RLImage(logo_path, width=4*cm, height=2*cm))
        elements.append(Spacer(1, 0.3*cm))

    elements.append(Paragraph(ar(f"التقرير الشهري — {month_name} {year}"), title_style))
    elements.append(Paragraph(ar("ملخص الأداء والحضور"), sub_style))
    elements.append(Spacer(1, 0.5*cm))

    # إحصائيات عامة
    total_emps    = len(all_emps)
    total_present = sum(sum(1 for s in att_map.get(e.id, {}).values() if s=="present") for e in all_emps)
    total_evals   = len(all_evals)
    avg_score     = round(sum(ev.total_score or 0 for ev in all_evals) / total_evals, 1) if total_evals else 0
    att_rate      = round(total_present / total_emps * 100, 1) if total_emps else 0

    summary_data = [
        [Paragraph(ar("القيمة"), normal_r), Paragraph(ar("البند"), normal_r)],
        [Paragraph(str(total_emps),    small_r), Paragraph(ar("إجمالي الموظفين"),     normal_r)],
        [Paragraph(str(len(supervisors)), small_r), Paragraph(ar("عدد المشرفين"),      normal_r)],
        [Paragraph(f"{att_rate}%",     small_r), Paragraph(ar("نسبة الحضور الكلية"),   normal_r)],
        [Paragraph(str(total_present), small_r), Paragraph(ar("إجمالي أيام الحضور"), normal_r)],
        [Paragraph(str(total_evals),   small_r), Paragraph(ar("تقييمات أسبوعية"),    normal_r)],
        [Paragraph(str(avg_score),     small_r), Paragraph(ar("متوسط الدرجات"),      normal_r)],
    ]
    sum_tbl = Table(summary_data, colWidths=[4*cm, 10*cm])
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), HDR_COLOR),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ROW1_COLOR, colors.white]),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.grey),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
    ]))
    elements.append(sum_tbl)
    elements.append(Spacer(1, 0.8*cm))

    # ملخص المشرفين
    elements.append(Paragraph(ar("ملخص المشرفين"), sub_style))
    elements.append(Spacer(1, 0.3*cm))

    sup_summary_data = [[
        Paragraph(ar("النسبة"), normal_r),
        Paragraph(ar("متوسط الدرجة"), normal_r),
        Paragraph(ar("تقييمات"), normal_r),
        Paragraph(ar("حضور"), normal_r),
        Paragraph(ar("موظفون"), normal_r),
        Paragraph(ar("المشرف"), normal_r),
    ]]

    for sup in supervisors:
        s_emps = emp_by_sup.get(sup.id, [])
        s_present = sum(sum(1 for st in att_map.get(e.id,{}).values() if st=="present") for e in s_emps)
        s_evals   = [ev for e in s_emps for ev in eval_map.get(e.id, [])]
        s_avg     = round(sum(s_evals)/len(s_evals),1) if s_evals else 0
        s_rate    = round(s_present/len(s_emps)*100,1) if s_emps else 0

        latin_r = ParagraphStyle("lat", fontName=LATIN_FONT, fontSize=8, alignment=1)
        sup_summary_data.append([
            Paragraph(f"{s_rate}%", latin_r),
            Paragraph(str(s_avg),   latin_r),
            Paragraph(str(len(s_evals)), latin_r),
            Paragraph(str(s_present), latin_r),
            Paragraph(str(len(s_emps)), latin_r),
            Paragraph(sup.name, latin_r),
        ])

    sup_tbl = Table(sup_summary_data, colWidths=[2.5*cm,3*cm,2.5*cm,2.5*cm,2.5*cm,5.5*cm])
    sup_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), SUB_COLOR),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ROW1_COLOR, colors.white]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.grey),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ]))
    elements.append(sup_tbl)

    # ══════════════════════════════════════════
    # تقرير إجمالي فقط (بدون تفاصيل المشرفين)
    # ══════════════════════════════════════════
    if report_type == "summary":
        doc.build(elements)
        return send_file(fpath, mimetype="application/pdf",
                         as_attachment=False,
                         download_name=fname)

    # ══════════════════════════════════════════
    # صفحة لكل مشرف (التقرير المفصّل)
    # ══════════════════════════════════════════
    for sup in supervisors:
        s_emps = emp_by_sup.get(sup.id, [])
        if not s_emps:
            continue

        elements.append(PageBreak())
        elements.append(Paragraph(sup.name, title_style))
        elements.append(Paragraph(ar(f"كود: {sup.supervisor_code} | {month_name} {year}"), sub_style))
        elements.append(Spacer(1, 0.4*cm))

        # جدول موظفي المشرف
        emp_data = [[
            Paragraph(ar("النسبة"),     normal_r),
            Paragraph(ar("متوسط الدرجة"), normal_r),
            Paragraph(ar("إجازة"),      normal_r),
            Paragraph(ar("غياب"),       normal_r),
            Paragraph(ar("حضور"),       normal_r),
            Paragraph(ar("القسم"),      normal_r),
            Paragraph(ar("الموظف"),     normal_r),
        ]]

        for emp in sorted(s_emps, key=lambda e: e.name):
            emp_att  = att_map.get(emp.id, {})
            present  = sum(1 for s in emp_att.values() if s=="present")
            absent   = sum(1 for s in emp_att.values() if s=="absent")
            leave    = sum(1 for s in emp_att.values() if s=="leave")
            scores   = eval_map.get(emp.id, [])
            avg_s    = round(sum(scores)/len(scores),1) if scores else 0
            rate     = round(present/days_in_month*100,1) if present else 0

            # لون النسبة
            rate_color = GREEN if rate >= 80 else (ORANGE if rate >= 50 else RED)

            emp_data.append([
                Paragraph(f"{rate}%", ParagraphStyle("rc", fontName=LATIN_FONT,
                          fontSize=8, alignment=1, textColor=rate_color)),
                Paragraph(str(avg_s) if avg_s else "—", small_r),
                Paragraph(str(leave),   small_r),
                Paragraph(str(absent),  small_r),
                Paragraph(str(present), small_r),
                Paragraph(emp.department or "—", ParagraphStyle("dep",
                           fontName=LATIN_FONT, fontSize=8, alignment=1)),
                Paragraph(emp.name, ParagraphStyle("emp_n",
                           fontName=LATIN_FONT, fontSize=8, alignment=1)),
            ])

        emp_tbl = Table(emp_data, colWidths=[2*cm,3*cm,2*cm,2*cm,2*cm,3*cm,5.5*cm])
        emp_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), HDR_COLOR),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [ROW1_COLOR, colors.white]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.grey),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        elements.append(emp_tbl)

        # ملخص المشرف
        s_present = sum(sum(1 for st in att_map.get(e.id,{}).values() if st=="present") for e in s_emps)
        s_evals   = [ev for e in s_emps for ev in eval_map.get(e.id, [])]
        s_avg     = round(sum(s_evals)/len(s_evals),1) if s_evals else 0
        s_rate    = round(s_present/len(s_emps)*100,1) if s_emps else 0

        elements.append(Spacer(1, 0.3*cm))
        footer_data = [[
            Paragraph(f"{s_rate}%", small_r),
            Paragraph(str(s_avg), small_r),
            Paragraph(str(s_present), small_r),
            Paragraph(str(len(s_emps)), small_r),
            Paragraph(ar("الإجمالي"), normal_r),
        ]]
        footer_tbl = Table(footer_data, colWidths=[3*cm,4*cm,3*cm,3*cm,6.5*cm])
        footer_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), colors.HexColor("#ecf0f1")),
            ("GRID",         (0,0), (-1,-1), 0.4, colors.grey),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ]))
        elements.append(footer_tbl)

    doc.build(elements)
    return send_file(fpath, mimetype="application/pdf",
                     as_attachment=False,
                     download_name=fname)

# ─────────────────────────────────────────────
#  الحضور — جلب (التطبيق)
# ─────────────────────────────────────────────
@app.get("/api/attendance")
@api_login_required
def api_attendance_get():
    u     = get_api_user()
    d_str = freq.args.get("date") or datetime.now(RIYADH_TZ).date().isoformat()
    try:
        d_val = parse_date(d_str)
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    if u.role == "admin":
        emps = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    else:
        emps = Employee.query.filter_by(user_id=u.id, is_active=True).order_by(Employee.name).all()

    emp_ids = [e.id for e in emps]
    atts    = {a.employee_id: a for a in
               Attendance.query.filter(
                   Attendance.employee_id.in_(emp_ids),
                   Attendance.date == d_val).all()}

    return jsonify([{
        "emp_id":     e.id,
        "emp_number": e.emp_number,
        "name":       e.name,
        "emp_name":   e.name,
        "department": e.department,
        "site":       e.site,
        "status":     atts[e.id].status  if e.id in atts else None,
        "remarks":    atts[e.id].remarks if e.id in atts else "",
    } for e in emps])


# ─────────────────────────────────────────────
#  الحضور — حفظ (التطبيق)
# ─────────────────────────────────────────────
@app.post("/api/attendance/save")
@api_login_required
def api_attendance_save():
    u        = get_api_user()
    data     = freq.get_json(force=True) or {}
    att_date = data.get("date") or datetime.now(RIYADH_TZ).date().isoformat()
    records  = data.get("records") or []

    try:
        att_date = parse_date(att_date)
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    changed = 0
    for rec in records:
        emp_id  = rec.get("emp_id")
        status  = rec.get("status")
        remarks = rec.get("remarks") or ""
        if not emp_id or not status:
            continue
        emp = db.session.get(Employee, emp_id)
        if not emp:
            continue
        if u.role != "admin" and emp.user_id != u.id:
            continue
        existing = Attendance.query.filter_by(employee_id=emp_id, date=att_date).first()
        if existing:
            existing.status  = status
            existing.remarks = remarks
        else:
            db.session.add(Attendance(
                employee_id=emp_id, supervisor_id=u.id,
                date=att_date, status=status, remarks=remarks))
        changed += 1

    db.session.commit()
    return jsonify({"message": f"Saved {changed} records"})


# ─────────────────────────────────────────────
#  إضافة موظف (التطبيق)
# ─────────────────────────────────────────────
@app.post("/api/employees/add")
@api_login_required
def api_employee_add():
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    emp_number = (data.get("emp_number") or "").strip()
    name       = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    site       = (data.get("site") or "").strip()

    if not emp_number or not name:
        return jsonify({"error": "emp_number and name are required"}), 400
    if not emp_number.isdigit():
        return jsonify({"error": "emp_number must be digits only"}), 400
    # التحقق إن الاسم والقسم والموقع إنجليزي فقط
    for field, val in [("name", name), ("department", department), ("site", site)]:
        if val and not all(ord(c) < 128 for c in val):
            return jsonify({"error": f"{field} must be in English only"}), 400

    emp = Employee.query.filter_by(emp_number=emp_number).first()
    if emp:
        if not emp.is_active:
            emp.is_active = True
            emp.user_id   = u.id
            db.session.commit()
            return jsonify({"message": "Employee reactivated", "id": emp.id})
        return jsonify({"error": "Employee already exists"}), 409

    emp = Employee(emp_number=emp_number, name=name,
                   department=department, site=site,
                   user_id=u.id, is_active=True)
    db.session.add(emp)
    db.session.commit()
    return jsonify({"message": "Employee added", "id": emp.id}), 201


# ─────────────────────────────────────────────
#  تعطيل موظف (التطبيق)
# ─────────────────────────────────────────────
@app.post("/api/employees/<int:emp_id>/deactivate")
@api_login_required
def api_employee_deactivate(emp_id):
    u   = get_api_user()
    emp = Employee.query.get_or_404(emp_id)
    if u.role != "admin" and emp.user_id != u.id:
        return jsonify({"error": "Forbidden"}), 403
    emp.is_active = False
    db.session.commit()
    return jsonify({"message": "Employee deactivated"})


# ─────────────────────────────────────────────
#  إدارة حالة الموظف (supervisor + site_supervisor)
# ─────────────────────────────────────────────

@app.post("/api/employees/<int:emp_id>/resign")
@api_login_required
def api_employee_resign(emp_id):
    """تعيين موظف كمستقيل — supervisor يملك الموظف أو admin"""
    u   = get_api_user()
    emp = Employee.query.get_or_404(emp_id)
    if u.role not in ("admin", "supervisor") or (u.role == "supervisor" and emp.user_id != u.id):
        return jsonify(error="Forbidden"), 403
    emp.status      = "resigned"
    emp.is_active   = False
    emp.resigned_at = datetime.now(RIYADH_TZ).date()
    db.session.commit()
    return jsonify(message="تم تسجيل الاستقالة")


@app.post("/api/employees/<int:emp_id>/unassign")
@api_login_required
def api_employee_unassign(emp_id):
    """فصل موظف عن مشرفه (يصبح unassigned) — supervisor أو admin"""
    u   = get_api_user()
    emp = Employee.query.get_or_404(emp_id)
    if u.role not in ("admin", "supervisor") or (u.role == "supervisor" and emp.user_id != u.id):
        return jsonify(error="Forbidden"), 403
    emp.status  = "unassigned"
    emp.user_id = None
    db.session.commit()
    return jsonify(message="تم فصل الموظف عن المشرف")


@app.post("/api/employees/<int:emp_id>/reactivate")
@api_login_required
def api_employee_reactivate(emp_id):
    """إعادة تفعيل موظف مستقيل أو غير معيّن — admin فقط"""
    u   = get_api_user()
    if u.role != "admin":
        return jsonify(error="Forbidden"), 403
    emp = Employee.query.get_or_404(emp_id)
    emp.status      = "active"
    emp.is_active   = True
    emp.resigned_at = None
    db.session.commit()
    return jsonify(message="تم إعادة التفعيل")


@app.post("/api/employees/<int:emp_id>/assign")
@api_login_required
def api_employee_assign(emp_id):
    """تعيين موظف unassigned لمشرف — site_supervisor أو admin"""
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    supervisor_id = data.get("supervisor_id")
    if not supervisor_id:
        return jsonify(error="supervisor_id مطلوب"), 400

    emp = Employee.query.get_or_404(emp_id)
    if emp.status != "unassigned":
        return jsonify(error="الموظف ليس في حالة unassigned"), 400

    # site_supervisor: يتحقق إن المشرف المختار تحت إشرافه
    if u.role == "site_supervisor":
        link = SiteSupervisorMap.query.filter_by(
            site_sup_id=u.id, supervisor_id=supervisor_id).first()
        if not link:
            return jsonify(error="المشرف ليس تحت إشرافك"), 403
    elif u.role != "admin":
        return jsonify(error="Forbidden"), 403

    sup = User.query.get_or_404(supervisor_id)
    if sup.role != "supervisor":
        return jsonify(error="المستخدم المختار ليس مشرفاً"), 400

    emp.user_id   = supervisor_id
    emp.status    = "active"
    emp.is_active = True
    db.session.commit()
    return jsonify(message=f"تم تعيين الموظف للمشرف {sup.name}")


@app.get("/api/employees/unassigned")
@api_login_required
def api_employees_unassigned():
    """الموظفون غير المعيّنين — site_supervisor يرى نطاقه، admin يرى الكل"""
    u = get_api_user()
    if u.role not in ("site_supervisor", "admin"):
        return jsonify(error="Forbidden"), 403

    emps = Employee.query.filter_by(status="unassigned").order_by(Employee.name).all()
    return jsonify([{
        "id":         e.id,
        "name":       e.name,
        "emp_number": e.emp_number,
        "department": e.department or "",
        "site":       e.site or "",
    } for e in emps])


@app.get("/api/supervisor/employees")
@api_login_required
def api_supervisor_employees_full():
    """موظفو المشرف بكل الحالات — active + unassigned + resigned"""
    u = get_api_user()
    if u.role not in ("supervisor", "admin"):
        return jsonify(error="Forbidden"), 403

    uid = u.id
    # المشرف: موظفوه الحاليون (active) + من كانوا تحته وأصبحوا مستقيلين
    actives   = Employee.query.filter_by(user_id=uid, status="active").all()
    resigned  = Employee.query.filter_by(user_id=uid, status="resigned").all()

    def row(e):
        return {
            "id":          e.id,
            "name":        e.name,
            "emp_number":  e.emp_number,
            "department":  e.department or "",
            "site":        e.site or "",
            "status":      e.status,
            "resigned_at": str(e.resigned_at) if e.resigned_at else None,
        }

    return jsonify({
        "active":   [row(e) for e in actives],
        "resigned": [row(e) for e in resigned],
    })


# ─────────────────────────────────────────────
#  التقييم اليومي — حفظ (التطبيق)
# ─────────────────────────────────────────────
@app.post("/api/daily-eval/save")
@api_login_required
def api_daily_eval_save():
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    emp_id    = data.get("employee_id")
    eval_date = data.get("eval_date") or date.today().isoformat()

    emp = (db.session.get(Employee, emp_id) if u.role == "admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first())
    if not emp:
        return jsonify({"error": "Employee not found"}), 404

    try:
        eval_date = parse_date(eval_date)
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    if DailyEvaluation.query.filter_by(employee_id=emp.id, eval_date=eval_date).first():
        return jsonify({"error": "Evaluation already exists for this date"}), 409

    de = DailyEvaluation(
        employee_id=emp.id,
        evaluator_id=u.id,
        eval_date=eval_date,
        t1_text=data.get("t1_text",""),    t1_percent=data.get("t1_percent",0),
        t2_text=data.get("t2_text",""),    t2_percent=data.get("t2_percent",0),
        t3_text=data.get("t3_text",""),    t3_percent=data.get("t3_percent",0),
        t4_text=data.get("t4_text",""),    t4_percent=data.get("t4_percent",0),
        p_punctuality=data.get("p_punctuality",0),
        p_quality=data.get("p_quality",0),
        p_productivity=data.get("p_productivity",0),
        p_communication=data.get("p_communication",0),
        p_problemsolving=data.get("p_problemsolving",0),
        p_compliance=data.get("p_compliance",0),
        strengths=data.get("strengths",""),
        improvements=data.get("improvements",""),
        training_needed=data.get("training_needed",""),
        company_id=api_cid(),
    )
    compute_daily_scores(de)
    db.session.add(de)
    db.session.commit()
    return jsonify({"message": "Saved", "total": de.total_score, "band": de.overall_band}), 201


# ─────────────────────────────────────────────
#  التقييم الأسبوعي — حفظ (التطبيق)
# ─────────────────────────────────────────────
@app.post("/api/weekly-eval/save")
@api_login_required
def api_weekly_eval_save():
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    emp_id     = data.get("employee_id")
    week_start = data.get("week_start")
    week_end   = data.get("week_end")

    emp = (db.session.get(Employee, emp_id) if u.role == "admin"
           else Employee.query.filter_by(id=emp_id, user_id=u.id).first())
    if not emp:
        return jsonify({"error": "Employee not found"}), 404

    try:
        ws = parse_date(week_start)
        we = parse_date(week_end)
    except Exception:
        return jsonify({"error": "Invalid dates"}), 400

    ok, msg = validate_week_sun_to_thu(ws, we)
    if not ok:
        return jsonify({"error": msg}), 400

    if Evaluation.query.filter_by(employee_id=emp.id, week_start=ws, week_end=we).first():
        return jsonify({"error": "Evaluation already exists for this week"}), 409

    ev = Evaluation(
        employee_id=emp.id,
        evaluator_id=u.id,
        week_start=ws,
        week_end=we,
        t1_text=data.get("t1_text",""),    t1_percent=data.get("t1_percent",0),    t1_remarks=data.get("t1_remarks",""),
        t2_text=data.get("t2_text",""),    t2_percent=data.get("t2_percent",0),    t2_remarks=data.get("t2_remarks",""),
        t3_text=data.get("t3_text",""),    t3_percent=data.get("t3_percent",0),    t3_remarks=data.get("t3_remarks",""),
        t4_text=data.get("t4_text",""),    t4_percent=data.get("t4_percent",0),    t4_remarks=data.get("t4_remarks",""),
        p_punctuality=data.get("p_punctuality",0),   c_punctuality=data.get("c_punctuality",""),
        p_quality=data.get("p_quality",0),           c_quality=data.get("c_quality",""),
        p_productivity=data.get("p_productivity",0), c_productivity=data.get("c_productivity",""),
        p_communication=data.get("p_communication",0),c_communication=data.get("c_communication",""),
        p_problemsolving=data.get("p_problemsolving",0),c_problemsolving=data.get("c_problemsolving",""),
        p_compliance=data.get("p_compliance",0),     c_compliance=data.get("c_compliance",""),
        strengths=data.get("strengths",""),
        improvements=data.get("improvements",""),
        training_needed=data.get("training_needed",""),
        company_id=api_cid(),
    )
    compute_scores(ev)
    db.session.add(ev)
    db.session.commit()
    return jsonify({"message": "Saved", "total": ev.total_score, "band": ev.overall_band}), 201


# ─────────────────────────────────────────────
#  تقارير أسبوعية (التطبيق)
# ─────────────────────────────────────────────
@app.get("/api/reports/weekly")
@api_login_required
def api_weekly_reports():
    u      = get_api_user()
    ws_str = freq.args.get("week_start")
    we_str = freq.args.get("week_end")

    if ws_str and we_str:
        try:
            ws = parse_date(ws_str)
            we = parse_date(we_str)
        except Exception:
            ws, we = default_week_today()
    else:
        ws, we = default_week_today()

    q = (db.session.query(Evaluation, Employee)
         .join(Employee, Evaluation.employee_id == Employee.id)
         .filter(Evaluation.week_start == ws, Evaluation.week_end == we))
    if u.role != "admin":
        q = q.filter(Employee.user_id == u.id)

    return jsonify([{
        "eval_id":    ev.id,
        "emp_id":     emp.id,
        "emp_name":   emp.name,
        "emp_number": emp.emp_number,
        "department": emp.department,
        "targets":    ev.targets_score,
        "perf":       ev.perf_score,
        "total":      ev.total_score,
        "band":       ev.overall_band,
        "week_start": str(ev.week_start),
        "week_end":   str(ev.week_end),
    } for ev, emp in q.order_by(Employee.name).all()])


# ─────────────────────────────────────────────
#  تقارير يومية (التطبيق)
# ─────────────────────────────────────────────
@app.get("/api/reports/daily")
@api_login_required
def api_daily_reports():
    u     = get_api_user()
    d_str = freq.args.get("date") or date.today().isoformat()
    try:
        d_val = parse_date(d_str)
    except Exception:
        return jsonify({"error": "Invalid date"}), 400

    q = (db.session.query(DailyEvaluation, Employee)
         .join(Employee, DailyEvaluation.employee_id == Employee.id)
         .filter(DailyEvaluation.eval_date == d_val))
    if u.role != "admin":
        q = q.filter(Employee.user_id == u.id)

    return jsonify([{
        "eval_id":    de.id,
        "emp_id":     emp.id,
        "emp_name":   emp.name,
        "emp_number": emp.emp_number,
        "department": emp.department,
        "targets":    de.targets_score,
        "perf":       de.perf_score,
        "total":      de.total_score,
        "band":       de.overall_band,
        "eval_date":  str(de.eval_date),
    } for de, emp in q.order_by(Employee.name).all()])


# ─────────────────────────────────────────────
#  كل التقارير الأسبوعية لموظف معين
# ─────────────────────────────────────────────
@app.get("/api/employees/<int:emp_id>/reports/weekly")
@api_login_required
def api_employee_weekly_reports(emp_id):
    u   = get_api_user()
    emp = Employee.query.get_or_404(emp_id)
    if u.role == "supervisor" and emp.user_id != u.id:
        return jsonify(error="Forbidden"), 403
    if u.role == "site_supervisor":
        link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=emp.user_id).first()
        if not link:
            return jsonify(error="Forbidden"), 403

    evals = (Evaluation.query
             .filter_by(employee_id=emp.id)
             .order_by(Evaluation.week_start.desc())
             .all())

    return jsonify([{
        "eval_id":    ev.id,
        "emp_id":     emp.id,
        "emp_name":   emp.name,
        "emp_number": emp.emp_number,
        "department": emp.department,
        "targets":    ev.targets_score,
        "perf":       ev.perf_score,
        "total":      ev.total_score,
        "band":       ev.overall_band,
        "week_start": str(ev.week_start),
        "week_end":   str(ev.week_end),
    } for ev in evals])


# ─────────────────────────────────────────────
#  كل التقارير اليومية لموظف معين
# ─────────────────────────────────────────────
@app.get("/api/employees/<int:emp_id>/reports/daily")
@api_login_required
def api_employee_daily_reports(emp_id):
    u   = get_api_user()
    emp = Employee.query.get_or_404(emp_id)
    if u.role == "supervisor" and emp.user_id != u.id:
        return jsonify(error="Forbidden"), 403
    if u.role == "site_supervisor":
        link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=emp.user_id).first()
        if not link:
            return jsonify(error="Forbidden"), 403

    evals = (DailyEvaluation.query
             .filter_by(employee_id=emp.id)
             .order_by(DailyEvaluation.eval_date.desc())
             .all())

    return jsonify([{
        "eval_id":    de.id,
        "emp_id":     emp.id,
        "emp_name":   emp.name,
        "emp_number": emp.emp_number,
        "department": emp.department,
        "targets":    de.targets_score,
        "perf":       de.perf_score,
        "total":      de.total_score,
        "band":       de.overall_band,
        "eval_date":  str(de.eval_date),
    } for de in evals])


# ══════════════════════════════════════════════════════════
#  site_supervisor  API
# ══════════════════════════════════════════════════════════

@app.get("/api/site/supervisors")
@api_login_required
def api_site_supervisors():
    """قائمة المشرفين التابعين لهذا site_supervisor"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    ws_str = request.args.get("week_start")
    we_str = request.args.get("week_end")
    ws, we = (parse_date(ws_str), parse_date(we_str)) if ws_str and we_str else default_week_today()
    rows = (db.session.query(SiteSupervisorMap, User)
            .join(User, SiteSupervisorMap.supervisor_id == User.id)
            .filter(SiteSupervisorMap.site_sup_id == u.id)
            .all())

    result = []
    for link, sup in rows:
        # عدد موظفيه
        emp_ids = [e.id for e in Employee.query.filter_by(user_id=sup.id, is_active=True).all()]
        emp_count = len(emp_ids)

        # كم منهم تم تقييمه هذا الأسبوع
        evaluated_count = 0
        if emp_ids:
            evaluated_count = Evaluation.query.filter(
                Evaluation.employee_id.in_(emp_ids),
                Evaluation.week_start == ws,
                Evaluation.week_end   == we
            ).count()
        evaluated = evaluated_count > 0
        progress = round((evaluated_count / emp_count * 100)) if emp_count > 0 else 0

        result.append({
            "id":              sup.id,
            "name":            sup.name,
            "code":            sup.supervisor_code,
            "emp_count":       emp_count,
            "evaluated":       evaluated,
            "evaluated_count": evaluated_count,
            "progress":        progress,
            "week_start":      str(ws),
            "week_end":        str(we),
        })

    return jsonify(result)


@app.get("/api/site/supervisors/<int:sup_id>/report")
@api_login_required
def api_site_supervisor_report(sup_id):
    """تقرير مشرف محدد — يعيد قائمة تقييماته"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_id).first()
    if not link:
        return jsonify(error="forbidden"), 403

    ws_str = request.args.get("week_start")
    we_str = request.args.get("week_end")
    if ws_str and we_str:
        ws = parse_date(ws_str); we = parse_date(we_str)
    else:
        ws, we = default_week_today()

    se = SupervisorEvaluation.query.filter_by(
        supervisor_id=sup_id, week_start=ws, week_end=we
    ).first()

    sup = User.query.get_or_404(sup_id)

    return jsonify({
        "supervisor": {"id": sup.id, "name": sup.name, "code": sup.supervisor_code},
        "week_start": str(ws),
        "week_end":   str(we),
        "evaluation": {
            "id":            se.id,
            "targets_score": se.targets_score,
            "perf_score":    se.perf_score,
            "total_score":   se.total_score,
            "overall_band":  se.overall_band,
            "targets": [
                {"text": se.t1_text, "percent": se.t1_percent, "remarks": se.t1_remarks},
                {"text": se.t2_text, "percent": se.t2_percent, "remarks": se.t2_remarks},
                {"text": se.t3_text, "percent": se.t3_percent, "remarks": se.t3_remarks},
                {"text": se.t4_text, "percent": se.t4_percent, "remarks": se.t4_remarks},
            ],
            "performance": {
                "punctuality":   {"score": se.p_punctuality,   "comment": se.c_punctuality},
                "quality":       {"score": se.p_quality,       "comment": se.c_quality},
                "productivity":  {"score": se.p_productivity,  "comment": se.c_productivity},
                "communication": {"score": se.p_communication, "comment": se.c_communication},
                "problemsolving":{"score": se.p_problemsolving,"comment": se.c_problemsolving},
                "compliance":    {"score": se.p_compliance,    "comment": se.c_compliance},
            },
            "strengths":        se.strengths,
            "improvements":     se.improvements,
            "training_needed":  se.training_needed,
        } if se else None
    })


@app.post("/api/site/evaluate/<int:sup_id>")
@api_login_required
def api_site_evaluate(sup_id):
    """تقييم مشرف من site_supervisor"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_id).first()
    if not link:
        return jsonify(error="forbidden"), 403

    data = freq.get_json(force=True)
    ws = parse_date(data.get("week_start")); we = parse_date(data.get("week_end"))
    if not (ws and we):
        return jsonify(error="week_start and week_end required"), 400

    ok, msg = validate_week_sun_to_thu(ws, we)
    if not ok:
        return jsonify(error=msg), 400

    if SupervisorEvaluation.query.filter_by(supervisor_id=sup_id, week_start=ws, week_end=we).first():
        return jsonify(error="Evaluation already exists for this week"), 409

    se = SupervisorEvaluation(
        supervisor_id=sup_id, evaluator_id=u.id,
        week_start=ws, week_end=we,
        t1_text=data.get("t1_text",""), t1_percent=float(data.get("t1_percent") or 0), t1_remarks=data.get("t1_remarks",""),
        t2_text=data.get("t2_text",""), t2_percent=float(data.get("t2_percent") or 0), t2_remarks=data.get("t2_remarks",""),
        t3_text=data.get("t3_text",""), t3_percent=float(data.get("t3_percent") or 0), t3_remarks=data.get("t3_remarks",""),
        t4_text=data.get("t4_text",""), t4_percent=float(data.get("t4_percent") or 0), t4_remarks=data.get("t4_remarks",""),
        p_punctuality=int(data.get("p_punctuality") or 0), c_punctuality=data.get("c_punctuality",""),
        p_quality=int(data.get("p_quality") or 0), c_quality=data.get("c_quality",""),
        p_productivity=int(data.get("p_productivity") or 0), c_productivity=data.get("c_productivity",""),
        p_communication=int(data.get("p_communication") or 0), c_communication=data.get("c_communication",""),
        p_problemsolving=int(data.get("p_problemsolving") or 0), c_problemsolving=data.get("c_problemsolving",""),
        p_compliance=int(data.get("p_compliance") or 0), c_compliance=data.get("c_compliance",""),
        strengths=data.get("strengths",""),
        improvements=data.get("improvements",""),
        training_needed=data.get("training_needed",""),
        company_id=api_cid(),
    )
    compute_scores(se)
    db.session.add(se)
    db.session.commit()
    return jsonify(id=se.id, total_score=se.total_score, overall_band=se.overall_band), 201


@app.get("/api/site/daily-reports")
@api_login_required
def api_site_daily_reports():
    """التقارير اليومية لموظفي المشرفين التابعين"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    d_str = request.args.get("d")
    d_val = parse_date(d_str) if d_str else datetime.now(RIYADH_TZ).date()

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    sup_ids = [su.id for _, su in links]

    q = (db.session.query(DailyEvaluation, Employee)
         .join(Employee, DailyEvaluation.employee_id == Employee.id)
         .filter(DailyEvaluation.eval_date == d_val))
    if sup_ids:
        q = q.filter(Employee.user_id.in_(sup_ids))

    rows = q.order_by(Employee.name.asc()).all()
    return jsonify([{
        "eval_id":    de.id,
        "emp_id":     emp.id,
        "emp_name":   emp.name,
        "emp_number": emp.emp_number,
        "total":      de.total_score,
        "band":       de.overall_band,
        "eval_date":  str(de.eval_date),
    } for de, emp in rows])


@app.get("/api/site/attendance")
@api_login_required
def api_site_attendance():
    """حضور موظفي المشرفين التابعين"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    d_str = request.args.get("d")
    d_val = parse_date(d_str) if d_str else datetime.now(RIYADH_TZ).date()

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    sup_ids = [su.id for _, su in links]

    q = (db.session.query(Attendance, Employee, User)
         .join(Employee, Attendance.employee_id == Employee.id)
         .join(User, Employee.user_id == User.id)
         .filter(Attendance.date == d_val))
    if sup_ids:
        q = q.filter(Employee.user_id.in_(sup_ids))

    rows = q.order_by(User.name.asc(), Employee.name.asc()).all()
    return jsonify([{
        "att_id":       att.id,
        "emp_id":       emp.id,
        "emp_name":     emp.name,
        "supervisor":   sup.name,
        "status":       att.status,
        "remarks":      att.remarks,
        "date":         str(att.date),
    } for att, emp, sup in rows])


@app.get("/api/site/supervisors/<int:sup_id>/employees")
@api_login_required
def api_site_supervisor_employees(sup_id):
    """موظفو مشرف محدد مع حالة تقييمهم الأسبوعي"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_id).first()
    if not link:
        return jsonify(error="forbidden"), 403

    ws, we = default_week_today()
    employees = Employee.query.filter_by(user_id=sup_id, is_active=True).order_by(Employee.name).all()

    result = []
    for emp in employees:
        eval_this_week = Evaluation.query.filter_by(
            employee_id=emp.id, week_start=ws, week_end=we
        ).first()
        result.append({
            "id":         emp.id,
            "name":       emp.name,
            "emp_number": emp.emp_number,
            "department": emp.department or "",
            "evaluated":  eval_this_week is not None,
            "eval_id":    eval_this_week.id if eval_this_week else None,
            "total_score": eval_this_week.total_score if eval_this_week else None,
            "overall_band": eval_this_week.overall_band if eval_this_week else None,
            "week_start": str(ws),
            "week_end":   str(we),
        })
    return jsonify(result)


@app.get("/api/site/reports/summary")
@api_login_required
def api_site_reports_summary():
    """ملخص تقارير جميع موظفي المشرفين التابعين"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    ws_str = request.args.get("week_start")
    we_str = request.args.get("week_end")
    ws, we = (parse_date(ws_str), parse_date(we_str)) if ws_str and we_str else default_week_today()

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    sup_ids = [su.id for _, su in links]

    all_emps = Employee.query.filter(Employee.user_id.in_(sup_ids), Employee.is_active == True).all()
    emp_ids  = [e.id for e in all_emps]

    evals = Evaluation.query.filter(
        Evaluation.employee_id.in_(emp_ids),
        Evaluation.week_start == ws,
        Evaluation.week_end   == we
    ).all()
    eval_map = {ev.employee_id: ev for ev in evals}

    evaluated     = [e for e in all_emps if e.id in eval_map]
    not_evaluated = [e for e in all_emps if e.id not in eval_map]

    def emp_row(emp, ev=None):
        sup = next((su for _, su in links if su.id == emp.user_id), None)
        return {
            "id":           emp.id,
            "name":         emp.name,
            "emp_number":   emp.emp_number,
            "department":   emp.department or "",
            "supervisor":   sup.name if sup else "",
            "evaluated":    ev is not None,
            "total_score":  ev.total_score if ev else None,
            "overall_band": ev.overall_band if ev else None,
        }

    return jsonify({
        "week_start":     str(ws),
        "week_end":       str(we),
        "total":          len(all_emps),
        "evaluated_count": len(evaluated),
        "not_evaluated_count": len(not_evaluated),
        "not_evaluated":  [emp_row(e) for e in not_evaluated],
        "evaluated":      [emp_row(e, eval_map[e.id]) for e in evaluated],
    })


@app.get("/api/site/attendance/supervisors")
@api_login_required
def api_site_attendance_supervisors():
    """المشرفون — هل سجلوا حضور موظفيهم اليوم"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    d_str = request.args.get("d")
    d_val = parse_date(d_str) if d_str else datetime.now(RIYADH_TZ).date()

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())

    result = []
    for link, sup in links:
        emp_ids = [e.id for e in Employee.query.filter_by(user_id=sup.id, is_active=True).all()]
        recorded = Attendance.query.filter(
            Attendance.employee_id.in_(emp_ids),
            Attendance.date == d_val
        ).count() if emp_ids else 0

        result.append({
            "id":          sup.id,
            "name":        sup.name,
            "code":        sup.supervisor_code,
            "emp_count":   len(emp_ids),
            "recorded":    recorded,
            "did_attend":  recorded > 0,
            "date":        str(d_val),
        })
    return jsonify(result)


@app.get("/api/site/attendance/supervisor/<int:sup_id>")
@api_login_required
def api_site_attendance_supervisor_detail(sup_id):
    """حضور موظفي مشرف محدد"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    link = SiteSupervisorMap.query.filter_by(site_sup_id=u.id, supervisor_id=sup_id).first()
    if not link:
        return jsonify(error="forbidden"), 403

    d_str = request.args.get("d")
    d_val = parse_date(d_str) if d_str else datetime.now(RIYADH_TZ).date()

    rows = (db.session.query(Attendance, Employee)
            .join(Employee, Attendance.employee_id == Employee.id)
            .filter(Employee.user_id == sup_id, Attendance.date == d_val)
            .order_by(Employee.name).all())

    present = sum(1 for a, _ in rows if a.status == "present")
    absent  = sum(1 for a, _ in rows if a.status == "absent")
    leave   = sum(1 for a, _ in rows if a.status == "leave")

    return jsonify({
        "date":    str(d_val),
        "summary": {"present": present, "absent": absent, "leave": leave},
        "rows": [{
            "att_id":   att.id,
            "emp_id":   emp.id,
            "emp_name": emp.name,
            "emp_number": emp.emp_number,
            "status":   att.status,
            "remarks":  att.remarks or "",
        } for att, emp in rows]
    })


@app.get("/api/site/requests")
@api_login_required
def api_site_requests():
    """طلبات المشرفين التابعين"""
    u = get_api_user()
    if not u or u.role != "site_supervisor":
        return jsonify(error="forbidden"), 403

    links = (db.session.query(SiteSupervisorMap, User)
             .join(User, SiteSupervisorMap.supervisor_id == User.id)
             .filter(SiteSupervisorMap.site_sup_id == u.id).all())
    sup_ids = [su.id for _, su in links]

    items = (Request.query
             .filter(Request.supervisor_id.in_(sup_ids))
             .order_by(Request.created_at.desc()).all())

    sup_map = {su.id: su.name for _, su in links}

    return jsonify([{
        "id":           r.id,
        "type":         r.type,
        "status":       r.status,
        "supervisor_id": r.supervisor_id,
        "supervisor_name": sup_map.get(r.supervisor_id, ""),
        "created_at":   str(r.created_at)[:10],
        "notes":        r.reason or "",
    } for r in items])


# ─────────────────────────────────────────────
#  إشعار مخصص من الأدمن للمشرفين
# ─────────────────────────────────────────────

@app.post("/api/admin/notify")
@api_admin_required
def api_admin_notify():
    """الأدمن يرسل إشعار لمشرف/مشرفين أو للكل"""
    data       = freq.get_json(force=True) or {}
    title      = (data.get("title") or "").strip()
    body       = (data.get("body")  or "").strip()
    target     = data.get("target", "all")   # "all" | "supervisors" | "site_supervisors" | list of user_ids
    user_ids   = data.get("user_ids", [])    # قائمة ids إذا target == "custom"

    if not title or not body:
        return jsonify(error="العنوان والرسالة مطلوبان"), 400

    # تحديد المستلمين
    if target == "all":
        users = User.query.filter(
            User.role.in_(["supervisor", "site_supervisor"]),
            User.is_active == True,
            User.is_hidden == False
        ).all()
    elif target == "supervisors":
        users = User.query.filter(User.role=="supervisor", User.is_active==True, User.is_hidden==False).all()
    elif target == "site_supervisors":
        users = User.query.filter(User.role=="site_supervisor", User.is_active==True, User.is_hidden==False).all()
    elif target == "custom" and user_ids:
        users = User.query.filter(User.id.in_(user_ids), User.is_active == True).all()
    else:
        return jsonify(error="target غير صالح"), 400

    sent = 0
    for u in users:
        tokens = [t.device_token for t in MobileToken.query.filter_by(user_id=u.id).all() if t.device_token]
        if tokens:
            threading.Thread(
                target=_send_push_to_user,
                args=(u.id, title, body),
                daemon=True
            ).start()
            sent += 1

    return jsonify(message=f"تم إرسال الإشعار لـ {sent} مستخدم", sent=sent, total=len(users))


@app.get("/api/admin/notify/users")
@api_admin_required
def api_admin_notify_users():
    """قائمة المشرفين للاختيار منهم عند الإرسال المخصص"""
    users = User.query.filter(
        User.role.in_(["supervisor", "site_supervisor"]),
        User.is_active == True,
        User.is_hidden == False
    ).order_by(User.role, User.name).all()
    return jsonify([{
        "id":   u.id,
        "name": u.name,
        "role": u.role,
        "code": u.supervisor_code,
    } for u in users])



# ══════════════════════════════════════════════════════════
#  نظام الإنذارات
# ══════════════════════════════════════════════════════════

DEFAULT_WARNING_REASONS = [
    "عدم اتباع أوامر المشرف",
    "التغيب بدون إذن",
    "التأخر المتكرر",
    "الإهمال في العمل",
    "سوء السلوك",
]

def seed_warning_reasons():
    """إضافة الأسباب الافتراضية إذا لم تكن موجودة"""
    for text in DEFAULT_WARNING_REASONS:
        if not WarningReason.query.filter_by(text=text).first():
            db.session.add(WarningReason(text=text, is_default=True))
    db.session.commit()

@app.get("/api/warning/reasons")
@api_login_required
def api_warning_reasons():
    """قائمة أسباب الإنذار"""
    reasons = WarningReason.query.order_by(WarningReason.is_default.desc(), WarningReason.id).all()
    return jsonify([{"id": r.id, "text": r.text, "is_default": r.is_default} for r in reasons])

@app.post("/api/warning/reasons")
@api_login_required
def api_warning_reasons_add():
    """إضافة سبب إنذار مخصص"""
    u    = get_api_user()
    data = freq.get_json(force=True) or {}
    text = (data.get("text") or "").strip()
    if not text:
        return jsonify(error="النص مطلوب"), 400
    if WarningReason.query.filter_by(text=text).first():
        return jsonify(error="السبب موجود مسبقاً"), 409
    r = WarningReason(text=text, is_default=False, created_by=u.id)
    db.session.add(r)
    db.session.commit()
    return jsonify({"id": r.id, "text": r.text}), 201

@app.get("/api/warning/form/<int:request_id>")
@api_login_required
def api_warning_form(request_id):
    """بيانات نموذج الإنذار للمشرف"""
    u   = get_api_user()
    req = Request.query.get_or_404(request_id)
    if u.role not in ("admin",) and req.supervisor_id != u.id:
        return jsonify(error="Forbidden"), 403

    sig = WarningSignature.query.filter_by(request_id=request_id).first()

    from zoneinfo import ZoneInfo
    now_riyadh = datetime.now(ZoneInfo("Asia/Riyadh"))

    task_item = HRTask.query.filter_by(request_id=request_id).first()
    try:
        official_reason = (task_item.official_reason if task_item and task_item.official_reason else req.reason) or "—"
    except Exception:
        official_reason = req.reason or "—"

    return jsonify({
        "request_id":    req.id,
        "employee_name": req.employee.name,
        "emp_number":    req.employee.emp_number,
        "department":    req.employee.department,
        "reason":        official_reason,
        "date":          now_riyadh.strftime("%Y/%m/%d"),
        "signed":        sig is not None,
        "signed_at":     str(sig.signed_at) if sig else None,
        "signer_name":   sig.employee_name if sig else None,
    })

@app.post("/api/warning/sign/<int:request_id>")
@api_login_required
def api_warning_sign(request_id):
    """حفظ توقيع الموظف"""
    u   = get_api_user()
    req = Request.query.get_or_404(request_id)
    if req.supervisor_id != u.id and u.role != "admin":
        return jsonify(error="Forbidden"), 403

    data      = freq.get_json(force=True) or {}
    signature = (data.get("signature") or "").strip()
    emp_name  = (data.get("employee_name") or req.employee.name).strip()

    if not signature:
        return jsonify(error="التوقيع مطلوب"), 400

    sig = WarningSignature.query.filter_by(request_id=request_id).first()
    if sig:
        sig.signature     = signature
        sig.employee_name = emp_name
        from zoneinfo import ZoneInfo
        sig.signed_at     = datetime.now(ZoneInfo("Asia/Riyadh"))
    else:
        from zoneinfo import ZoneInfo
        sig = WarningSignature(
            request_id    = request_id,
            signature     = signature,
            employee_name = emp_name,
            signed_at     = datetime.now(ZoneInfo("Asia/Riyadh")),
        )
        db.session.add(sig)
    db.session.commit()

    # توليد PDF الإنذار
    try:
        pdf_name = generate_warning_pdf(req, sig)
        if pdf_name:
            # تحديث HRTask بمسار الـ PDF
            task = HRTask.query.filter_by(request_id=req.id).first()
            if task:
                task.warning_pdf = pdf_name
                db.session.commit()
    except Exception as e:
        app.logger.error("warning PDF error: %s", e)

    return jsonify(message="تم حفظ التوقيع"), 200


# ─────────────────────────────────────────────
#  تحميل PDF الإنذار
# ─────────────────────────────────────────────
@app.get("/api/warning/<int:request_id>/download")
def api_warning_download(request_id):
    # يقبل token من header أو query string (لأن Safari لا يرسل headers)
    from flask import request as freq2
    token = freq2.headers.get("Authorization", "").replace("Bearer ", "").strip()
    if not token:
        token = freq2.args.get("token", "")
    if not token:
        return jsonify(error="Unauthorized"), 401
    mt = MobileToken.query.filter_by(token=token).first()
    if not mt:
        return jsonify(error="Unauthorized"), 401

    task = HRTask.query.filter_by(request_id=request_id).first_or_404()
    if not task.warning_pdf:
        return jsonify(error="PDF غير موجود"), 404
    fpath = os.path.join(UPLOAD_FOLDER, task.warning_pdf)
    if not os.path.exists(fpath):
        return jsonify(error="الملف غير موجود على السيرفر"), 404
    return send_file(fpath,
                     mimetype="application/pdf",
                     as_attachment=False,
                     download_name=task.warning_pdf)

# ─────────────────────────────────────────────────────────────────
#  توقيع الموظف على نموذج الإجازة
# ─────────────────────────────────────────────────────────────────
@app.post("/api/leave/sign/<int:request_id>")
@api_login_required
def api_leave_sign(request_id):
    u   = get_api_user()
    req = Request.query.get_or_404(request_id)
    if req.supervisor_id != u.id and u.role not in ("admin", "hr"):
        return jsonify(error="Forbidden"), 403

    data      = freq.get_json(force=True) or {}
    signature = (data.get("signature") or "").strip()
    emp_name  = (data.get("employee_name") or (req.employee.name if req.employee else "")).strip()

    if not signature:
        return jsonify(error="التوقيع مطلوب"), 400

    from zoneinfo import ZoneInfo
    now = datetime.now(ZoneInfo("Asia/Riyadh"))
    sig = LeaveSignature.query.filter_by(request_id=request_id).first()
    if sig:
        sig.signature = signature; sig.employee_name = emp_name; sig.signed_at = now
    else:
        sig = LeaveSignature(request_id=request_id, signature=signature,
                             employee_name=emp_name, signed_at=now)
        db.session.add(sig)

    # تحديث حالة نموذج الإجازة
    lf = LeaveForm.query.filter_by(request_id=request_id).first()
    if lf:
        lf.signed    = True
        lf.signed_at = now
        # أعد توليد PDF مع التوقيع
        try:
            pdf_name = generate_leave_pdf(req, sig)
            if pdf_name:
                lf.filename = pdf_name
        except Exception as e:
            app.logger.error("leave PDF regen error: %s", e)

    db.session.commit()

    # إشعار HR بعد التوقيع
    try:
        hr_users = User.query.filter_by(role="hr", company_id=req.company_id, is_active=True).all()
        for hr in hr_users:
            _send_push_to_user(hr.id, title="توقيع نموذج إجازة",
                               body=f"وقّع {emp_name} على نموذج إجازة {req.employee.name if req.employee else ''}")
    except Exception:
        pass

    return jsonify(message="تم حفظ التوقيع", signed=True), 200


@app.get("/api/leave/form/<int:request_id>")
@api_login_required
def api_leave_form_info(request_id):
    """معلومات نموذج الإجازة + حالة التوقيع"""
    req = Request.query.get_or_404(request_id)
    lf  = LeaveForm.query.filter_by(request_id=request_id).first()
    sig = LeaveSignature.query.filter_by(request_id=request_id).first()
    emp = req.employee
    sup = req.supervisor
    return jsonify({
        "request_id":    request_id,
        "employee_name": emp.name       if emp else "",
        "emp_number":    emp.emp_number if emp else "",
        "department":    emp.department if emp else "",
        "site":          emp.site       if emp else "",
        "supervisor":    sup.name       if sup else "",
        "start_date":    str(req.start_date) if req.start_date else "",
        "end_date":      str(req.end_date)   if req.end_date   else "",
        "days":          str((req.end_date - req.start_date).days + 1) if req.start_date and req.end_date else "",
        "reason":        req.reason or "",
        "approved_at":   str(req.decided_at.date()) if req.decided_at else "",
        "has_form":      lf is not None,
        "pdf_url":       f"/api/leave-form/{request_id}/download" if lf else None,
        "signed":        lf.signed if lf else False,
        "signed_at":     str(lf.signed_at) if lf and lf.signed_at else None,
        "signer_name":   sig.employee_name if sig else None,
    })


# ─────────────────────────────────────────────────────────────────
#  توقيع الموظف على نموذج الاستئذان
# ─────────────────────────────────────────────────────────────────
@app.post("/api/permission/sign/<int:request_id>")
@api_login_required
def api_permission_sign(request_id):
    u   = get_api_user()
    req = Request.query.get_or_404(request_id)
    if req.type != "permission":
        return jsonify(error="ليس طلب استئذان"), 400
    if req.supervisor_id != u.id and u.role not in ("admin", "hr"):
        return jsonify(error="Forbidden"), 403

    data      = freq.get_json(force=True) or {}
    signature = (data.get("signature") or "").strip()
    emp_name  = (data.get("employee_name") or (req.employee.name if req.employee else "")).strip()

    if not signature:
        return jsonify(error="التوقيع مطلوب"), 400

    from zoneinfo import ZoneInfo
    now = datetime.now(ZoneInfo("Asia/Riyadh"))
    sig = PermissionSignature.query.filter_by(request_id=request_id).first()
    if sig:
        sig.signature = signature; sig.employee_name = emp_name; sig.signed_at = now
    else:
        sig = PermissionSignature(request_id=request_id, signature=signature,
                                  employee_name=emp_name, signed_at=now)
        db.session.add(sig)
    db.session.commit()

    # توليد PDF الاستئذان بعد التوقيع
    try:
        pdf_name = generate_permission_pdf(req, sig)
        if pdf_name:
            task = HRTask.query.filter_by(request_id=req.id).first()
            if task:
                task.warning_pdf = pdf_name  # نعيد استخدام حقل warning_pdf
                db.session.commit()
    except Exception as e:
        app.logger.error("permission PDF error: %s", e)

    return jsonify(message="تم حفظ التوقيع", signed=True), 200


@app.get("/api/permission/form/<int:request_id>")
@api_login_required
def api_permission_form_info(request_id):
    req = Request.query.get_or_404(request_id)
    sig = PermissionSignature.query.filter_by(request_id=request_id).first()
    emp = req.employee
    sup = req.supervisor
    task = HRTask.query.filter_by(request_id=request_id).first()
    return jsonify({
        "request_id":    request_id,
        "employee_name": emp.name       if emp else "",
        "emp_number":    emp.emp_number if emp else "",
        "department":    emp.department if emp else "",
        "supervisor":    sup.name       if sup else "",
        "date":          str(req.start_date or req.created_at.date()),
        "reason":        req.reason or "",
        "status":        req.status,
        "signed":        sig is not None,
        "signed_at":     str(sig.signed_at) if sig else None,
        "signer_name":   sig.employee_name  if sig else None,
        "pdf_url":       f"/api/warning/{request_id}/download" if (task and task.warning_pdf) else None,
    })


# =====================================================================
# WEEKLY REPORT EXPORTS
# =====================================================================

def _weekly_report_data(ws, we):
    """جمع بيانات التقرير الأسبوعي: قائمة الموظفين + تقييماتهم + حضورهم."""
    rows = (db.session.query(Evaluation, Employee, User)
            .join(Employee, Evaluation.employee_id == Employee.id)
            .join(User, Employee.user_id == User.id)
            .filter(Evaluation.week_start == ws, Evaluation.week_end == we)
            .order_by(User.supervisor_code.asc(), Employee.name.asc())
            .all())

    # الحضور لهذا الأسبوع
    emp_ids = [e.id for _, e, _ in rows]
    att_map = {}
    if emp_ids:
        att_records = (Attendance.query
                       .filter(Attendance.employee_id.in_(emp_ids),
                               Attendance.date >= ws, Attendance.date <= we)
                       .all())
        for a in att_records:
            att_map.setdefault(a.employee_id, []).append(a)

    result = []
    for ev, emp, sup in rows:
        att_list = att_map.get(emp.id, [])
        present  = sum(1 for a in att_list if a.status == "present")
        absent   = sum(1 for a in att_list if a.status == "absent")
        leave    = sum(1 for a in att_list if a.status == "leave")
        result.append({
            "sup_code":    sup.supervisor_code,
            "sup_name":    sup.name,
            "emp_number":  emp.emp_number,
            "emp_name":    emp.name,
            "department":  emp.department or "",
            "site":        emp.site or "",
            "band":        ev.overall_band or "—",
            "total_score": ev.total_score or 0,
            "targets":     ev.targets_score or 0,
            "perf":        ev.perf_score or 0,
            "present":     present,
            "absent":      absent,
            "leave":       leave,
        })
    return result


@app.get("/admin/report/weekly/print")
@admin_required
def admin_report_weekly_print():
    ws = parse_date(request.args.get("week_start", ""))
    we = parse_date(request.args.get("week_end", ""))
    if not ws or not we:
        ws, we = default_week_today()
    data = _weekly_report_data(ws, we)
    from datetime import datetime as _dt
    return render_template("report_weekly_print.html", rows=data, ws=ws, we=we,
                           now=_dt.now(RIYADH_TZ))


@app.get("/admin/report/weekly/excel")
@admin_required
def admin_report_weekly_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response

    ws_date = parse_date(request.args.get("week_start", ""))
    we_date = parse_date(request.args.get("week_end", ""))
    if not ws_date or not we_date:
        ws_date, we_date = default_week_today()

    data = _weekly_report_data(ws_date, we_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Report"

    # Header row
    headers = ["#", "Supervisor", "Sup ID", "Employee", "Emp #",
               "Department", "Site", "Band", "Score", "Targets", "Perf",
               "Present", "Absent", "Leave"]
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    band_colors = {
        "Excellent":        "C8E6C9",
        "Good":             "BBDEFB",
        "Satisfactory":     "FFE0B2",
        "Needs Improvement":"FFCDD2",
    }

    for i, r in enumerate(data, 2):
        row_vals = [i-1, r["sup_name"], r["sup_code"], r["emp_name"],
                    r["emp_number"], r["department"], r["site"],
                    r["band"], r["total_score"], r["targets"], r["perf"],
                    r["present"], r["absent"], r["leave"]]
        fill_color = band_colors.get(r["band"], "FFFFFF")
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col == 8:  # Band column
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(bold=True)
            if col in (9, 10, 11):  # Score columns
                cell.number_format = "0.0"

    # Column widths
    widths = [5, 22, 12, 24, 12, 18, 14, 16, 8, 8, 8, 9, 9, 9]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"weekly_{ws_date.isoformat()}_{we_date.isoformat()}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


@app.get("/admin/report/weekly/pdf")
@admin_required
def admin_report_weekly_pdf():
    import io
    from flask import make_response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    ws_date = parse_date(request.args.get("week_start", ""))
    we_date = parse_date(request.args.get("week_end", ""))
    if not ws_date or not we_date:
        ws_date, we_date = default_week_today()

    data = _weekly_report_data(ws_date, we_date)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=14, spaceAfter=4)
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.grey, spaceAfter=10)

    BAND_COLORS = {
        "Excellent":        colors.Color(0.78, 0.93, 0.78),
        "Good":             colors.Color(0.73, 0.87, 0.98),
        "Satisfactory":     colors.Color(1.0,  0.88, 0.7),
        "Needs Improvement":colors.Color(1.0,  0.80, 0.80),
    }

    # Table data
    header = ["#", "Supervisor", "Employee", "Emp #", "Dept",
              "Band", "Score", "T", "P", "Pres", "Abs", "Lv"]
    table_data = [header]
    row_styles = []

    for i, r in enumerate(data, 1):
        table_data.append([
            str(i), r["sup_name"][:20], r["emp_name"][:22],
            r["emp_number"], r["department"][:14],
            r["band"], f"{r['total_score']:.1f}",
            f"{r['targets']:.1f}", f"{r['perf']:.1f}",
            str(r["present"]), str(r["absent"]), str(r["leave"]),
        ])
        bg = BAND_COLORS.get(r["band"])
        if bg:
            row_styles.append(("BACKGROUND", (5, i+1), (5, i+1), bg))

    col_widths = [8*mm, 38*mm, 40*mm, 22*mm, 25*mm,
                  26*mm, 14*mm, 11*mm, 11*mm, 12*mm, 10*mm, 10*mm]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    base_style = [
        ("BACKGROUND",  (0,0), (-1,0), colors.Color(0.06, 0.09, 0.16)),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,-1), 7.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.Color(0.97,0.97,0.97)]),
        ("GRID",        (0,0), (-1,-1), 0.4, colors.Color(0.82,0.82,0.82)),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",       (6,0), (-1,-1), "CENTER"),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ] + row_styles
    tbl.setStyle(TableStyle(base_style))

    story = [
        Paragraph(f"Weekly Employee Evaluation Report", title_style),
        Paragraph(f"Week: {ws_date.strftime('%d %b %Y')} — {we_date.strftime('%d %b %Y')}  |  Total: {len(data)} evaluations", sub_style),
        tbl,
    ]
    doc.build(story)
    buf.seek(0)

    fname = f"weekly_{ws_date.isoformat()}.pdf"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


# =====================================================================
# HSE MODULE — Safety Officer Tracking
# =====================================================================

ALLOWED_IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "gif", "webp"}
HSE_UPLOAD_DIR = os.path.join(UPLOAD_FOLDER, "hse")


def _parse_date(s, default=None):
    s = (s or "").strip()
    if not s:
        return default
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default
os.makedirs(HSE_UPLOAD_DIR, exist_ok=True)
OBS_CATEGORIES = [
    "PPE", "Excavation", "Housekeeping", "Barrier Management",
    "Vehicle & Equipment", "Work at Height", "Lifting Operations",
    "Line of Fire", "Control of Chemicals", "High Risk Situation", "Other",
]


def _allowed_image(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


def _save_hse_photo(file, prefix, company_id=None):
    if not file or not file.filename:
        return None
    if not _allowed_image(file.filename):
        return None
    ext = file.filename.rsplit(".", 1)[1].lower()
    fname = f"hse_{prefix}_{uuid.uuid4().hex[:10]}.{ext}"
    if company_id:
        subfolder = os.path.join(HSE_UPLOAD_DIR, str(company_id))
        os.makedirs(subfolder, exist_ok=True)
        file.save(os.path.join(subfolder, fname))
        return f"{company_id}/{fname}"
    file.save(os.path.join(HSE_UPLOAD_DIR, fname))
    return fname


def hse_officer_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        u = cur_user()
        if not u:
            return redirect(url_for("login"))
        if u.role != "safety_officer":
            flash("هذه الصفحة خاصة بضباط السلامة فقط.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return inner


HSE_SUPERVISOR_CODE = os.environ.get("HSE_SUPERVISOR_CODE", "39468")

def hse_supervisor_required(f):
    @wraps(f)
    def inner(*args, **kwargs):
        u = cur_user()
        if not u:
            return redirect(url_for("login"))
        if not is_hse_supervisor(u):
            flash("هذه الصفحة للمشرفين فقط.", "danger")
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return inner


# ── HSE Models ──────────────────────────────────────────────────────

class HseCheckin(db.Model):
    __tablename__ = "hse_checkin"
    id            = db.Column(db.Integer, primary_key=True)
    officer_id    = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date          = db.Column(db.Date, nullable=False)
    location      = db.Column(db.String(255), nullable=False)
    supervisor_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    company_id    = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("officer_id", "date", name="uq_hse_checkin_od"),)


class HseObservation(db.Model):
    __tablename__ = "hse_observation"
    id             = db.Column(db.Integer, primary_key=True)
    officer_id     = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date           = db.Column(db.Date, nullable=False)
    location       = db.Column(db.String(255))
    obs_type       = db.Column(db.Enum("unsafe_act", "unsafe_condition", "positive"), nullable=False)
    category       = db.Column(db.String(100))
    risk_level     = db.Column(db.Enum("L", "M", "H"))
    description    = db.Column(db.Text)
    action_taken   = db.Column(db.Text)
    status         = db.Column(db.Enum("open", "closed"), default="open")
    closed_at      = db.Column(db.Date, nullable=True)
    closure_action = db.Column(db.Text, nullable=True)
    company_id     = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    photos         = db.relationship("HseObservationPhoto", backref="observation", lazy="dynamic")


class HseObservationPhoto(db.Model):
    __tablename__ = "hse_observation_photos"
    id             = db.Column(db.Integer, primary_key=True)
    observation_id = db.Column(db.Integer, db.ForeignKey("hse_observation.id"), nullable=False)
    photo_path     = db.Column(db.String(500))
    photo_type     = db.Column(db.Enum("before", "after"))
    uploaded_at    = db.Column(db.DateTime, default=datetime.utcnow)
    company_id     = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)


class HseJsoClosure(db.Model):
    __tablename__ = "hse_jso_closure"
    id           = db.Column(db.Integer, primary_key=True)
    officer_id   = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    jso_number   = db.Column(db.String(50), nullable=False)
    date         = db.Column(db.Date, nullable=False)
    location     = db.Column(db.String(255))
    action_taken = db.Column(db.Text)
    photo_path   = db.Column(db.String(500), nullable=True)
    company_id   = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)


class HseTbt(db.Model):
    __tablename__ = "hse_tbt"
    id              = db.Column(db.Integer, primary_key=True)
    officer_id      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date            = db.Column(db.Date, nullable=False)
    topic           = db.Column(db.String(255))
    location        = db.Column(db.String(255))
    supervisor_id   = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    sign_photo_path = db.Column(db.String(500), nullable=True)
    company_id      = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at      = db.Column(db.DateTime, default=datetime.utcnow)
    attendance      = db.relationship("HseTbtAttendance", backref="tbt", lazy="dynamic")


class HseTbtAttendance(db.Model):
    __tablename__ = "hse_tbt_attendance"
    id         = db.Column(db.Integer, primary_key=True)
    tbt_id     = db.Column(db.Integer, db.ForeignKey("hse_tbt.id"), nullable=False)
    emp_number = db.Column(db.String(50))
    emp_name   = db.Column(db.String(120))
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)


class HseNearMiss(db.Model):
    __tablename__ = "hse_near_miss"
    id              = db.Column(db.Integer, primary_key=True)
    officer_id      = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date            = db.Column(db.Date, nullable=False)
    location        = db.Column(db.String(255))
    description     = db.Column(db.Text)
    immediate_cause = db.Column(db.Text)
    action_taken    = db.Column(db.Text)
    reported_to     = db.Column(db.String(255))
    photo_path      = db.Column(db.String(500), nullable=True)
    company_id      = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at      = db.Column(db.DateTime, default=datetime.utcnow)


class HseBbs(db.Model):
    __tablename__ = "hse_bbs"
    id          = db.Column(db.Integer, primary_key=True)
    officer_id  = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    date        = db.Column(db.Date, nullable=False)
    card_count  = db.Column(db.Integer, nullable=False)
    notes       = db.Column(db.Text, nullable=True)
    company_id  = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("officer_id", "date", name="uq_hse_bbs_od"),)


class HseSupervisorAccess(db.Model):
    """مستخدمون منحهم الأدمن صلاحية لوحة تحكم HSE"""
    __tablename__ = "hse_supervisor_access"
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, unique=True)
    granted_at = db.Column(db.DateTime, default=lambda: datetime.now(RIYADH_TZ))
    granted_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)


# ── Phase 2 Models ────────────────────────────────────────────────────

class HsePtw(db.Model):
    __tablename__ = "hse_ptw"
    id             = db.Column(db.Integer, primary_key=True)
    officer_id     = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    company_id     = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    permit_number  = db.Column(db.String(50), nullable=False)
    permit_type    = db.Column(db.String(50), nullable=False)
    description    = db.Column(db.Text)
    location       = db.Column(db.String(255))
    week_start     = db.Column(db.Date, nullable=False)
    week_end       = db.Column(db.Date, nullable=False)
    status         = db.Column(db.Enum("active", "suspended", "closed"), default="active")
    attached_to_id = db.Column(db.Integer, db.ForeignKey("hse_ptw.id"), nullable=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)


class HseManpower(db.Model):
    __tablename__ = "hse_manpower"
    id          = db.Column(db.Integer, primary_key=True)
    officer_id  = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    company_id  = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    date        = db.Column(db.Date, nullable=False)
    location    = db.Column(db.String(255))
    total_count = db.Column(db.Integer, nullable=False)
    breakdown   = db.Column(db.Text, nullable=True)
    notes       = db.Column(db.Text, nullable=True)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("officer_id", "date", name="uq_hse_manpower_od"),)


class HseCorrectiveAction(db.Model):
    __tablename__ = "hse_corrective_action"
    id               = db.Column(db.Integer, primary_key=True)
    observation_id   = db.Column(db.Integer, db.ForeignKey("hse_observation.id"), nullable=False)
    company_id       = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    assigned_to      = db.Column(db.String(255))
    due_date         = db.Column(db.Date, nullable=False)
    action_required  = db.Column(db.Text, nullable=False)
    status           = db.Column(db.Enum("open", "in_progress", "completed"), default="open")
    completed_at     = db.Column(db.Date, nullable=True)
    completion_notes = db.Column(db.Text, nullable=True)
    created_by       = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at       = db.Column(db.DateTime, default=datetime.utcnow)
    observation      = db.relationship("HseObservation", backref="corrective_actions")


INSPECTION_ITEMS = [
    "PPE Compliance (Helmet, Vest, Boots)",
    "Housekeeping & Work Area Clean",
    "Barricading & Warning Signs in Place",
    "Fire Extinguisher Accessible",
    "PTW Displayed at Work Site",
    "First Aid Kit Available",
    "No Unauthorized Personnel in Area",
    "Electrical Safety (No Exposed Wires)",
    "Scaffolding Tagged & Inspected",
    "Emergency Exits Clear",
]


class HseInspection(db.Model):
    __tablename__ = "hse_inspection"
    id            = db.Column(db.Integer, primary_key=True)
    officer_id    = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    company_id    = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=True)
    date          = db.Column(db.Date, nullable=False)
    location      = db.Column(db.String(255))
    checklist     = db.Column(db.Text, nullable=False)
    overall_score = db.Column(db.Float, nullable=True)
    notes         = db.Column(db.Text, nullable=True)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint("officer_id", "date", name="uq_hse_inspection_od"),)


def is_hse_supervisor(u):
    """صحيح إذا كان المستخدم لديه صلاحية HSE supervisor"""
    if not u or not getattr(u, "is_active", False):
        return False
    if getattr(u, "supervisor_code", None) == HSE_SUPERVISOR_CODE:
        return True
    return HseSupervisorAccess.query.filter_by(user_id=u.id).first() is not None


@app.context_processor
def inject_hse_flag():
    from flask import g as _g
    u = getattr(_g, "user", None)
    try:
        flag = is_hse_supervisor(u)
    except Exception:
        flag = False
    return {"g_is_hse_supervisor": flag}


# ── HSE Helpers ──────────────────────────────────────────────────────

def _hse_today_location(officer_id):
    today = datetime.now(RIYADH_TZ).date()
    ci = HseCheckin.query.filter_by(officer_id=officer_id, date=today).first()
    return ci.location if ci else ""


def _hse_locations():
    q = db.session.query(HseCheckin.location,
                         func.count(HseCheckin.location).label("cnt"))
    _cid = cid()
    if _cid:
        q = q.filter(HseCheckin.company_id == _cid)
    rows = q.group_by(HseCheckin.location).order_by(func.count(HseCheckin.location).desc()).all()
    return [r.location for r in rows]


def _notify_hse_supervisors(title, body):
    """يرسل push notification لكل من لديه صلاحية HSE supervisor"""
    import threading as _thr
    # المشرف الأساسي
    primary = User.query.filter_by(supervisor_code=HSE_SUPERVISOR_CODE, is_active=True).first()
    recipients = {primary.id} if primary else set()
    # المستخدمون الآخرون الممنوحة لهم الصلاحية
    for acc in HseSupervisorAccess.query.all():
        u = db.session.get(User, acc.user_id)
        if u and u.is_active:
            recipients.add(u.id)
    for uid in recipients:
        _thr.Thread(target=_send_push_to_user, args=(uid, title, body), daemon=True).start()


# ── HSE: Serve photos ────────────────────────────────────────────────

@app.get("/hse/photo/<path:filename>")
@login_required
def hse_photo(filename):
    from flask import send_from_directory
    return send_from_directory(HSE_UPLOAD_DIR, filename)


# ── HSE API ──────────────────────────────────────────────────────────

@app.get("/hse/api/locations")
@login_required
def hse_api_locations():
    return jsonify(_hse_locations())


@app.get("/hse/api/supervisor_lookup")
@login_required
def hse_supervisor_lookup():
    code = request.args.get("code", "").strip()
    if not code:
        return jsonify({"found": False})
    sup = User.query.filter(User.supervisor_code == code, User.is_active == True).first()
    if sup:
        return jsonify({"found": True, "name": sup.name})
    return jsonify({"found": False})


@app.get("/hse/api/employee_lookup")
@login_required
def hse_employee_lookup():
    emp_num = request.args.get("emp_number", "").strip()
    if not emp_num:
        return jsonify({"found": False})
    emp = Employee.query.filter_by(emp_number=emp_num, status="active").first()
    if emp:
        return jsonify({"found": True, "name": emp.name})
    for prefix in ("NSH-", "GA-"):
        if emp_num.upper().startswith(prefix):
            bare = emp_num[len(prefix):]
            emp = Employee.query.filter_by(emp_number=bare, status="active").first()
            if emp:
                return jsonify({"found": True, "name": emp.name})
        cand = Employee.query.filter_by(emp_number=f"{prefix}{emp_num}", status="active").first()
        if cand:
            return jsonify({"found": True, "name": cand.name})
    return jsonify({"found": False})


# ── Admin: HSE Supervisor Access ─────────────────────────────────────

@app.route("/admin/hse-access", methods=["GET", "POST"])
@login_required
@admin_required
def admin_hse_access():
    if request.method == "POST":
        action = request.form.get("action")
        code   = (request.form.get("code") or "").strip()
        user_id = request.form.get("user_id", type=int)

        if action == "add" and code:
            target = User.query.filter_by(supervisor_code=code, is_active=True).first()
            if not target:
                flash("User not found or inactive.", "warning")
            elif target.supervisor_code == HSE_SUPERVISOR_CODE:
                flash("Primary supervisor — always has access.", "info")
            elif HseSupervisorAccess.query.filter_by(user_id=target.id).first():
                flash(f"{target.name} already has access.", "info")
            else:
                db.session.add(HseSupervisorAccess(
                    user_id=target.id, granted_by=cur_user().id))
                db.session.commit()
                flash(f"Access granted to {target.name}.", "success")

        elif action == "remove" and user_id:
            acc = HseSupervisorAccess.query.filter_by(user_id=user_id).first()
            if acc:
                name = db.session.get(User, user_id).name if db.session.get(User, user_id) else "?"
                db.session.delete(acc)
                db.session.commit()
                flash(f"Access removed from {name}.", "success")

        return redirect(url_for("admin_hse_access"))

    primary = User.query.filter_by(supervisor_code=HSE_SUPERVISOR_CODE, is_active=True).first()
    access_list = (db.session.query(HseSupervisorAccess, User)
                   .join(User, HseSupervisorAccess.user_id == User.id)
                   .all())
    return render_template("admin_hse_access.html",
                           primary=primary, access_list=access_list)


# ── HSE: Daily Check-in ───────────────────────────────────────────────

@app.route("/hse/checkin", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_checkin():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    existing = HseCheckin.query.filter_by(officer_id=u.id, date=today).first()
    locations = _hse_locations()

    if request.method == "POST":
        location = request.form.get("location", "").strip()
        sup_code = request.form.get("supervisor_code", "").strip()

        if not location:
            flash("Location is required.", "warning")
            return redirect(url_for("hse_checkin"))

        supervisor = None
        if sup_code:
            supervisor = User.query.filter(
                User.supervisor_code == sup_code, User.is_active == True
            ).first()
            if not supervisor:
                flash("Supervisor code not found.", "warning")
                return redirect(url_for("hse_checkin"))

        if existing:
            existing.location = location
            existing.supervisor_id = supervisor.id if supervisor else None
            flash("Check-in updated.", "success")
        else:
            ci = HseCheckin(
                officer_id=u.id, date=today, location=location,
                supervisor_id=supervisor.id if supervisor else None,
                company_id=cid(),
            )
            db.session.add(ci)
            flash("Checked in successfully.", "success")

        db.session.commit()
        return redirect(url_for("hse_checkin"))

    checkin_supervisor = None
    if existing and existing.supervisor_id:
        checkin_supervisor = db.session.get(User, existing.supervisor_id)

    return render_template("hse_checkin.html",
                           existing=existing, today=today,
                           locations=locations, checkin_supervisor=checkin_supervisor)


# ── HSE: Daily Observations ───────────────────────────────────────────

@app.route("/hse/observations", methods=["GET"])
@login_required
@hse_officer_required
def hse_observations():
    u = cur_user()
    status_filter = request.args.get("status", "open")
    page = request.args.get("page", 1, type=int)
    q = HseObservation.query.filter_by(officer_id=u.id)
    if status_filter in ("open", "closed"):
        q = q.filter_by(status=status_filter)
    pagination = q.order_by(HseObservation.date.desc()).paginate(page=page, per_page=15, error_out=False)
    obs = pagination.items
    obs_photos = {o.id: list(o.photos) for o in obs}
    return render_template("hse_observations.html",
                           obs=obs, obs_photos=obs_photos, status_filter=status_filter,
                           pagination=pagination)


@app.route("/hse/observation/new", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_observation_new():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    today_location = _hse_today_location(u.id)
    locations = _hse_locations()

    if request.method == "POST":
        obs_date = _safe_date(request.form.get("date")) or today
        location = request.form.get("location", "").strip()
        obs_type = request.form.get("obs_type", "").strip()
        category = request.form.get("category", "").strip()
        risk_level = request.form.get("risk_level", "").strip() or None
        description = request.form.get("description", "").strip()
        action_taken = request.form.get("action_taken", "").strip()

        if not obs_type:
            flash("Observation type is required.", "warning")
            return redirect(url_for("hse_observation_new"))

        obs = HseObservation(
            officer_id=u.id, date=obs_date, location=location,
            obs_type=obs_type, category=category, risk_level=risk_level,
            description=description, action_taken=action_taken,
            company_id=cid(),
        )
        db.session.add(obs)
        db.session.flush()

        for i in range(1, 4):
            f = request.files.get(f"photo_{i}")
            path = _save_hse_photo(f, "obs", company_id=cid())
            if path:
                db.session.add(HseObservationPhoto(
                    observation_id=obs.id, photo_path=path, photo_type="before"
                ))

        db.session.commit()
        flash("Observation saved.", "success")
        if risk_level == "H":
            _notify_hse_supervisors("⚠ High-Risk Observation",
                                    f"{u.name}: {category or 'No category'} at {location}")
        return redirect(url_for("hse_observations"))

    return render_template("hse_observation_new.html",
                           today=today, today_location=today_location,
                           locations=locations, categories=OBS_CATEGORIES)


@app.route("/hse/observation/<int:obs_id>/close", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_observation_close(obs_id):
    u = cur_user()
    obs = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()

    if obs.status == "closed":
        flash("This observation is already closed.", "info")
        return redirect(url_for("hse_observations", status="closed"))

    if request.method == "POST":
        closure_action = request.form.get("closure_action", "").strip()
        if not closure_action:
            flash("Closure action is required.", "warning")
            return redirect(url_for("hse_observation_close", obs_id=obs_id))

        f = request.files.get("closure_photo")
        path = _save_hse_photo(f, "obs_close", company_id=cid())
        if path:
            db.session.add(HseObservationPhoto(
                observation_id=obs.id, photo_path=path, photo_type="after"
            ))

        obs.status = "closed"
        obs.closed_at = datetime.now(RIYADH_TZ).date()
        obs.closure_action = closure_action
        db.session.commit()
        flash("Observation closed successfully.", "success")
        return redirect(url_for("hse_observations", status="closed"))

    return render_template("hse_observation_close.html", obs=obs)


# ── HSE: JSO Closure ──────────────────────────────────────────────────

@app.route("/hse/jso", methods=["GET"])
@login_required
@hse_officer_required
def hse_jso_list():
    u = cur_user()
    page = request.args.get("page", 1, type=int)
    pagination = (HseJsoClosure.query.filter_by(officer_id=u.id)
                  .order_by(HseJsoClosure.date.desc())
                  .paginate(page=page, per_page=15, error_out=False))
    return render_template("hse_jso.html", items=pagination.items, pagination=pagination)


@app.route("/hse/jso/new", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_jso_new():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    today_location = _hse_today_location(u.id)
    locations = _hse_locations()

    if request.method == "POST":
        jso_number = request.form.get("jso_number", "").strip()
        jso_date = _safe_date(request.form.get("date")) or today
        location = request.form.get("location", "").strip()
        action_taken = request.form.get("action_taken", "").strip()

        if not jso_number:
            flash("JSO number is required.", "warning")
            return redirect(url_for("hse_jso_new"))

        photo_path = _save_hse_photo(request.files.get("photo"), "jso", company_id=cid())

        db.session.add(HseJsoClosure(
            officer_id=u.id, jso_number=jso_number, date=jso_date,
            location=location, action_taken=action_taken, photo_path=photo_path,
            company_id=cid(),
        ))
        db.session.commit()
        flash("JSO closure recorded.", "success")
        return redirect(url_for("hse_jso_list"))

    return render_template("hse_jso_new.html",
                           today=today, today_location=today_location, locations=locations)


# ── HSE: TBT ──────────────────────────────────────────────────────────

@app.route("/hse/tbt", methods=["GET"])
@login_required
@hse_officer_required
def hse_tbt_list():
    u = cur_user()
    page = request.args.get("page", 1, type=int)
    pagination = (HseTbt.query.filter_by(officer_id=u.id)
                  .order_by(HseTbt.date.desc())
                  .paginate(page=page, per_page=15, error_out=False))
    tbts = pagination.items
    tbt_ids = [t.id for t in tbts]
    tbt_counts = dict(
        db.session.query(HseTbtAttendance.tbt_id, func.count())
        .filter(HseTbtAttendance.tbt_id.in_(tbt_ids))
        .group_by(HseTbtAttendance.tbt_id).all()
    ) if tbt_ids else {}
    sup_ids = {t.supervisor_id for t in tbts if t.supervisor_id}
    supervisors = {u.id: u.name for u in User.query.filter(User.id.in_(sup_ids)).all()} if sup_ids else {}
    return render_template("hse_tbt.html", tbts=tbts, tbt_counts=tbt_counts,
                           pagination=pagination, supervisors=supervisors)


@app.route("/hse/tbt/new", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_tbt_new():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    today_location = _hse_today_location(u.id)
    locations = _hse_locations()

    if request.method == "POST":
        tbt_date = _safe_date(request.form.get("date")) or today
        topic = request.form.get("topic", "").strip()
        location = request.form.get("location", "").strip()
        sup_code = request.form.get("supervisor_code", "").strip()

        supervisor = None
        if sup_code:
            supervisor = User.query.filter(
                User.supervisor_code == sup_code, User.is_active == True
            ).first()

        sign_path = _save_hse_photo(request.files.get("sign_photo"), "tbt", company_id=cid())

        tbt = HseTbt(
            officer_id=u.id, date=tbt_date, topic=topic, location=location,
            supervisor_id=supervisor.id if supervisor else None,
            sign_photo_path=sign_path,
            company_id=cid(),
        )
        db.session.add(tbt)
        db.session.flush()

        emp_numbers = request.form.getlist("emp_number[]")
        emp_names = request.form.getlist("emp_name[]")
        for num, name in zip(emp_numbers, emp_names):
            num = num.strip(); name = name.strip()
            if num and name:
                db.session.add(HseTbtAttendance(tbt_id=tbt.id, emp_number=num, emp_name=name))

        db.session.commit()
        flash("TBT session saved.", "success")
        return redirect(url_for("hse_tbt_list"))

    return render_template("hse_tbt_new.html",
                           today=today, today_location=today_location, locations=locations)


# ── HSE: Near Miss ────────────────────────────────────────────────────

@app.route("/hse/nearmiss", methods=["GET"])
@login_required
@hse_officer_required
def hse_nearmiss_list():
    u = cur_user()
    page = request.args.get("page", 1, type=int)
    pagination = (HseNearMiss.query.filter_by(officer_id=u.id)
                  .order_by(HseNearMiss.date.desc())
                  .paginate(page=page, per_page=15, error_out=False))
    return render_template("hse_nearmiss.html", items=pagination.items, pagination=pagination)


@app.route("/hse/nearmiss/new", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_nearmiss_new():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    today_location = _hse_today_location(u.id)
    locations = _hse_locations()

    if request.method == "POST":
        nm_date = _safe_date(request.form.get("date")) or today
        location = request.form.get("location", "").strip()
        description = request.form.get("description", "").strip()
        immediate_cause = request.form.get("immediate_cause", "").strip()
        action_taken = request.form.get("action_taken", "").strip()
        reported_to = request.form.get("reported_to", "").strip()

        if not all([location, description, immediate_cause, action_taken, reported_to]):
            flash("All fields except photo are required.", "warning")
            return redirect(url_for("hse_nearmiss_new"))

        photo_path = _save_hse_photo(request.files.get("photo"), "nm", company_id=cid())

        db.session.add(HseNearMiss(
            officer_id=u.id, date=nm_date, location=location,
            description=description, immediate_cause=immediate_cause,
            action_taken=action_taken, reported_to=reported_to, photo_path=photo_path,
            company_id=cid(),
        ))
        db.session.commit()
        flash("Near miss recorded.", "success")
        return redirect(url_for("hse_nearmiss_list"))

    return render_template("hse_nearmiss_new.html",
                           today=today, today_location=today_location, locations=locations)


# ── HSE: BBS Daily Count ──────────────────────────────────────────────

@app.route("/hse/bbs", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_bbs():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()

    if request.method == "POST":
        bbs_date = _safe_date(request.form.get("date")) or today
        try:
            card_count = int(request.form.get("card_count", 0))
        except ValueError:
            card_count = 0
        notes = request.form.get("notes", "").strip()

        rec = HseBbs.query.filter_by(officer_id=u.id, date=bbs_date).first()
        if rec:
            rec.card_count = card_count
            rec.notes = notes
            flash("BBS count updated.", "success")
        else:
            db.session.add(HseBbs(officer_id=u.id, date=bbs_date,
                                  card_count=card_count, notes=notes,
                                  company_id=cid()))
            flash("BBS count saved.", "success")
        db.session.commit()
        return redirect(url_for("hse_bbs"))

    existing = HseBbs.query.filter_by(officer_id=u.id, date=today).first()
    history = (HseBbs.query.filter_by(officer_id=u.id)
               .order_by(HseBbs.date.desc()).limit(30).all())
    return render_template("hse_bbs.html", existing=existing, today=today, history=history)


# ── HSE: Supervisor Dashboard ─────────────────────────────────────────

@app.route("/hse/dashboard", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_dashboard():
    today = datetime.now(RIYADH_TZ).date()
    # Saudi work week starts Sunday; isoweekday: Mon=1 … Sun=7
    days_since_sunday = (today.isoweekday() % 7)
    week_start = today - timedelta(days=days_since_sunday)

    _cid = cid()
    officers_q = User.query.filter_by(role="safety_officer", is_active=True)
    if _cid:
        officers_q = officers_q.filter(User.company_id == _cid)
    officers = officers_q.all()

    checkin_q = HseCheckin.query.filter_by(date=today)
    if _cid:
        checkin_q = checkin_q.filter(HseCheckin.company_id == _cid)
    today_checkins = {ci.officer_id: ci for ci in checkin_q.all()}

    officer_ids = [o.id for o in officers]

    # ── استعلامات مجمّعة GROUP BY بدلاً من N+1 حلقة ──────────────────
    def _grp(model, col, filt_date_col, extra=None):
        q = (db.session.query(col, func.count())
             .filter(col.in_(officer_ids),
                     filt_date_col.between(week_start, today))
             .group_by(col))
        if extra is not None:
            q = q.filter(extra)
        return dict(q.all())

    obs_map = _grp(HseObservation, HseObservation.officer_id, HseObservation.date)
    jso_map = _grp(HseJsoClosure,  HseJsoClosure.officer_id,  HseJsoClosure.date)
    tbt_map = _grp(HseTbt,         HseTbt.officer_id,         HseTbt.date)
    nm_map  = _grp(HseNearMiss,    HseNearMiss.officer_id,    HseNearMiss.date)

    bbs_map = dict(
        db.session.query(HseBbs.officer_id,
                         func.coalesce(func.sum(HseBbs.card_count), 0))
        .filter(HseBbs.officer_id.in_(officer_ids),
                HseBbs.date.between(week_start, today))
        .group_by(HseBbs.officer_id).all()
    )

    ci_map = dict(
        db.session.query(HseCheckin.officer_id, func.count())
        .filter(HseCheckin.officer_id.in_(officer_ids),
                HseCheckin.date.between(week_start, today))
        .group_by(HseCheckin.officer_id).all()
    )

    ptw_map = dict(
        db.session.query(HsePtw.officer_id, func.count())
        .filter(HsePtw.officer_id.in_(officer_ids),
                HsePtw.status == "active",
                HsePtw.week_end >= today)
        .group_by(HsePtw.officer_id).all()
    )

    insp_map = dict(
        db.session.query(HseInspection.officer_id, func.count())
        .filter(HseInspection.officer_id.in_(officer_ids),
                HseInspection.date.between(week_start, today))
        .group_by(HseInspection.officer_id).all()
    )

    mp_map = {
        mp.officer_id: mp.total_count
        for mp in HseManpower.query.filter(
            HseManpower.officer_id.in_(officer_ids),
            HseManpower.date == today
        ).all()
    }

    insp_score_map = {
        ins.officer_id: ins.overall_score
        for ins in HseInspection.query.filter(
            HseInspection.officer_id.in_(officer_ids),
            HseInspection.date == today
        ).all()
    }

    weekly = {}
    for o in officers:
        oid = o.id
        obs = obs_map.get(oid, 0)
        jso = jso_map.get(oid, 0)
        tbt = tbt_map.get(oid, 0)
        nm  = nm_map.get(oid, 0)
        bbs = int(bbs_map.get(oid, 0))
        ci  = ci_map.get(oid, 0)
        ptw = ptw_map.get(oid, 0)
        insp = insp_map.get(oid, 0)
        score = round(ci*1 + obs*2 + tbt*3 + nm*5 + bbs*0.1 + insp*2 + ptw*1, 1)
        weekly[oid] = {
            "obs": obs, "jso": jso, "tbt": tbt, "nm": nm, "bbs": bbs,
            "total": obs + jso + tbt + nm,
            "ptw": ptw, "manpower": mp_map.get(oid, 0),
            "insp_score": insp_score_map.get(oid),
            "score": score,
        }
    # ─────────────────────────────────────────────────────────────────

    high_risk_q = HseObservation.query.filter_by(status="open", risk_level="H")
    if _cid:
        high_risk_q = high_risk_q.filter(HseObservation.company_id == _cid)
    high_risk = high_risk_q.order_by(HseObservation.date.asc()).all()
    officer_map = {o.id: o for o in officers}

    # Week totals for KPI cards
    total_obs_week  = sum(weekly[o.id]["obs"] for o in officers)
    total_jso_week  = sum(weekly[o.id]["jso"] for o in officers)
    total_tbt_week  = sum(weekly[o.id]["tbt"] for o in officers)
    total_nm_week   = sum(weekly[o.id]["nm"]  for o in officers)
    total_bbs_week  = sum(weekly[o.id]["bbs"] for o in officers)

    obs_base = HseObservation.query.filter(HseObservation.date.between(week_start, today))
    if _cid:
        obs_base = obs_base.filter(HseObservation.company_id == _cid)
    obs_unsafe_act  = obs_base.filter(HseObservation.obs_type == "unsafe_act").count()
    obs_unsafe_cond = obs_base.filter(HseObservation.obs_type == "unsafe_condition").count()
    obs_positive    = obs_base.filter(HseObservation.obs_type == "positive").count()

    max_activity = max((weekly[o.id]["total"] for o in officers), default=1) or 1

    # Corrective actions
    ca_q = HseCorrectiveAction.query
    if _cid: ca_q = ca_q.filter(HseCorrectiveAction.company_id == _cid)
    overdue_ca = ca_q.filter(
        HseCorrectiveAction.status != "completed",
        HseCorrectiveAction.due_date < today
    ).order_by(HseCorrectiveAction.due_date.asc()).all()
    open_ca = ca_q.filter(HseCorrectiveAction.status != "completed").count()

    # PTW expiring soon
    ptw_expiring_q = HsePtw.query.filter(
        HsePtw.status == "active",
        HsePtw.week_end >= today,
        HsePtw.week_end <= today + timedelta(days=2)
    )
    if _cid: ptw_expiring_q = ptw_expiring_q.filter(HsePtw.company_id == _cid)
    ptw_expiring = ptw_expiring_q.order_by(HsePtw.week_end.asc()).all()

    # Total man power today
    mp_today_q = db.session.query(func.coalesce(func.sum(HseManpower.total_count), 0))\
                   .filter(HseManpower.date == today)
    if _cid: mp_today_q = mp_today_q.filter(HseManpower.company_id == _cid)
    total_manpower_today = int(mp_today_q.scalar() or 0)

    # Total active PTW
    ptw_active_q = HsePtw.query.filter(HsePtw.status == "active", HsePtw.week_end >= today)
    if _cid: ptw_active_q = ptw_active_q.filter(HsePtw.company_id == _cid)
    total_ptw_active = ptw_active_q.count()

    # Inspection avg score this week
    insp_q = HseInspection.query.filter(HseInspection.date.between(week_start, today))
    if _cid: insp_q = insp_q.filter(HseInspection.company_id == _cid)
    insp_scores = [i.overall_score for i in insp_q.all() if i.overall_score is not None]
    avg_insp_score = round(sum(insp_scores) / len(insp_scores), 1) if insp_scores else None

    # Officers with no activity today
    inactive_officers = [o for o in officers if o.id not in today_checkins
                         and weekly[o.id]["total"] == 0]

    # Flash alerts for dashboard visitor
    if overdue_ca:
        flash(f"⚠ {len(overdue_ca)} Corrective Action(s) are overdue — please follow up.", "warning")
    if ptw_expiring:
        flash(f"📋 {len(ptw_expiring)} PTW permit(s) expire within 48 hours.", "warning")
    if inactive_officers:
        names = ", ".join(o.name for o in inactive_officers[:3])
        extra = f" (+{len(inactive_officers)-3} more)" if len(inactive_officers) > 3 else ""
        flash(f"🔴 No activity today: {names}{extra}", "warning")

    return render_template("hse_dashboard.html",
                           officers=officers, today=today, week_start=week_start,
                           today_checkins=today_checkins, weekly=weekly,
                           high_risk=high_risk, officer_map=officer_map,
                           total_obs_week=total_obs_week, total_jso_week=total_jso_week,
                           total_tbt_week=total_tbt_week, total_nm_week=total_nm_week,
                           total_bbs_week=total_bbs_week,
                           obs_unsafe_act=obs_unsafe_act, obs_unsafe_cond=obs_unsafe_cond,
                           obs_positive=obs_positive, max_activity=max_activity,
                           overdue_ca=overdue_ca, open_ca=open_ca,
                           ptw_expiring=ptw_expiring, total_ptw_active=total_ptw_active,
                           total_manpower_today=total_manpower_today,
                           avg_insp_score=avg_insp_score,
                           inactive_officers=inactive_officers)


@app.route("/hse/dashboard/officer/<int:officer_id>", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_officer_detail(officer_id):
    _cid = cid()
    officer_q = User.query.filter_by(id=officer_id, role="safety_officer")
    if _cid is not None:
        officer_q = officer_q.filter(User.company_id == _cid)
    officer = officer_q.first_or_404()

    date_from = _safe_date(request.args.get("from"))
    date_to = _safe_date(request.args.get("to"))

    def apply_dates(q, col):
        if date_from:
            q = q.filter(col >= date_from)
        if date_to:
            q = q.filter(col <= date_to)
        return q

    checkins = apply_dates(
        HseCheckin.query.filter_by(officer_id=officer_id), HseCheckin.date
    ).order_by(HseCheckin.date.desc()).all()

    observations = apply_dates(
        HseObservation.query.filter_by(officer_id=officer_id), HseObservation.date
    ).order_by(HseObservation.date.desc()).all()

    jsos = apply_dates(
        HseJsoClosure.query.filter_by(officer_id=officer_id), HseJsoClosure.date
    ).order_by(HseJsoClosure.date.desc()).all()

    tbts = apply_dates(
        HseTbt.query.filter_by(officer_id=officer_id), HseTbt.date
    ).order_by(HseTbt.date.desc()).all()

    nearmisses = apply_dates(
        HseNearMiss.query.filter_by(officer_id=officer_id), HseNearMiss.date
    ).order_by(HseNearMiss.date.desc()).all()

    bbs_records = apply_dates(
        HseBbs.query.filter_by(officer_id=officer_id), HseBbs.date
    ).order_by(HseBbs.date.desc()).all()

    obs_photos    = {o.id: list(o.photos) for o in observations}
    _tbt_ids      = [t.id for t in tbts]
    tbt_counts    = dict(
        db.session.query(HseTbtAttendance.tbt_id, func.count())
        .filter(HseTbtAttendance.tbt_id.in_(_tbt_ids))
        .group_by(HseTbtAttendance.tbt_id).all()
    ) if _tbt_ids else {}
    tbt_attendees = {t.id: list(t.attendance) for t in tbts}

    # Phase 2 data
    ptw_records = apply_dates(
        HsePtw.query.filter_by(officer_id=officer_id), HsePtw.week_start
    ).order_by(HsePtw.week_start.desc()).all()

    mp_records = apply_dates(
        HseManpower.query.filter_by(officer_id=officer_id), HseManpower.date
    ).order_by(HseManpower.date.desc()).all()

    insp_records = apply_dates(
        HseInspection.query.filter_by(officer_id=officer_id), HseInspection.date
    ).order_by(HseInspection.date.desc()).all()

    # Weekly scores — last 8 weeks
    import json as _json
    today = datetime.now(RIYADH_TZ).date()
    weekly_scores = []
    for i in range(7, -1, -1):
        days_since_sun = today.isoweekday() % 7
        ws_i = today - timedelta(days=days_since_sun + i * 7)
        we_i = ws_i + timedelta(days=4)
        weekly_scores.append({
            "label": ws_i.strftime("%d %b"),
            "score": _hse_officer_score(officer_id, ws_i, we_i),
        })

    return render_template("hse_officer_detail.html",
                           officer=officer,
                           checkins=checkins,
                           observations=observations, obs_photos=obs_photos,
                           jsos=jsos,
                           tbts=tbts, tbt_counts=tbt_counts, tbt_attendees=tbt_attendees,
                           nearmisses=nearmisses,
                           bbs_records=bbs_records,
                           ptw_records=ptw_records,
                           mp_records=mp_records,
                           insp_records=insp_records,
                           weekly_scores=weekly_scores,
                           date_from=request.args.get("from", ""),
                           date_to=request.args.get("to", ""))


# ===================== HSE Edit / Delete =====================

@app.route("/hse/observation/<int:obs_id>/edit", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_observation_edit(obs_id):
    u = cur_user()
    obs = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    locations = _hse_locations()
    if request.method == "POST":
        obs.date        = _safe_date(request.form.get("date")) or obs.date
        obs.location    = request.form.get("location", "").strip()
        obs.obs_type    = request.form.get("obs_type", "").strip()
        obs.category    = request.form.get("category", "").strip()
        obs.risk_level  = request.form.get("risk_level", "").strip() or None
        obs.description = request.form.get("description", "").strip()
        obs.action_taken= request.form.get("action_taken", "").strip()
        db.session.commit()
        flash("Observation updated.", "success")
        return redirect(url_for("hse_observations"))
    return render_template("hse_observation_edit.html",
                           obs=obs, locations=locations, categories=OBS_CATEGORIES)


@app.route("/hse/observation/<int:obs_id>/delete", methods=["POST"])
@login_required
@hse_officer_required
def hse_observation_delete(obs_id):
    u = cur_user()
    obs = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    for p in list(obs.photos):
        db.session.delete(p)
    db.session.delete(obs)
    db.session.commit()
    flash("Observation deleted.", "success")
    return redirect(url_for("hse_observations"))


@app.route("/hse/jso/<int:jso_id>/edit", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_jso_edit(jso_id):
    u = cur_user()
    jso = HseJsoClosure.query.filter_by(id=jso_id, officer_id=u.id).first_or_404()
    locations = _hse_locations()
    if request.method == "POST":
        jso.jso_number  = request.form.get("jso_number", "").strip()
        jso.date        = _safe_date(request.form.get("date")) or jso.date
        jso.location    = request.form.get("location", "").strip()
        jso.action_taken= request.form.get("action_taken", "").strip()
        new_photo = _save_hse_photo(request.files.get("photo"), "jso", company_id=cid())
        if new_photo:
            jso.photo_path = new_photo
        db.session.commit()
        flash("JSO record updated.", "success")
        return redirect(url_for("hse_jso_list"))
    return render_template("hse_jso_edit.html", jso=jso, locations=locations)


@app.route("/hse/jso/<int:jso_id>/delete", methods=["POST"])
@login_required
@hse_officer_required
def hse_jso_delete(jso_id):
    u = cur_user()
    jso = HseJsoClosure.query.filter_by(id=jso_id, officer_id=u.id).first_or_404()
    db.session.delete(jso)
    db.session.commit()
    flash("JSO record deleted.", "success")
    return redirect(url_for("hse_jso_list"))


@app.route("/hse/tbt/<int:tbt_id>/edit", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_tbt_edit(tbt_id):
    u = cur_user()
    tbt = HseTbt.query.filter_by(id=tbt_id, officer_id=u.id).first_or_404()
    locations = _hse_locations()
    attendees = HseTbtAttendance.query.filter_by(tbt_id=tbt_id).all()
    if request.method == "POST":
        tbt.date     = _safe_date(request.form.get("date")) or tbt.date
        tbt.topic    = request.form.get("topic", "").strip()
        tbt.location = request.form.get("location", "").strip()
        sup_code = request.form.get("supervisor_code", "").strip()
        if sup_code:
            sup = User.query.filter(User.supervisor_code == sup_code, User.is_active == True).first()
            if sup:
                tbt.supervisor_id = sup.id
        new_sign = _save_hse_photo(request.files.get("sign_photo"), "tbt", company_id=cid())
        if new_sign:
            tbt.sign_photo_path = new_sign
        HseTbtAttendance.query.filter_by(tbt_id=tbt.id).delete()
        for num, name in zip(request.form.getlist("emp_number[]"), request.form.getlist("emp_name[]")):
            num = num.strip(); name = name.strip()
            if num and name:
                db.session.add(HseTbtAttendance(tbt_id=tbt.id, emp_number=num, emp_name=name))
        db.session.commit()
        flash("TBT session updated.", "success")
        return redirect(url_for("hse_tbt_list"))
    sup_code_val = ""
    if tbt.supervisor_id:
        sup = db.session.get(User, tbt.supervisor_id)
        sup_code_val = sup.supervisor_code if sup else ""
    return render_template("hse_tbt_edit.html",
                           tbt=tbt, locations=locations,
                           attendees=attendees, sup_code_val=sup_code_val)


@app.route("/hse/tbt/<int:tbt_id>/delete", methods=["POST"])
@login_required
@hse_officer_required
def hse_tbt_delete(tbt_id):
    u = cur_user()
    tbt = HseTbt.query.filter_by(id=tbt_id, officer_id=u.id).first_or_404()
    HseTbtAttendance.query.filter_by(tbt_id=tbt.id).delete()
    db.session.delete(tbt)
    db.session.commit()
    flash("TBT session deleted.", "success")
    return redirect(url_for("hse_tbt_list"))


@app.route("/hse/nearmiss/<int:nm_id>/edit", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_nearmiss_edit(nm_id):
    u = cur_user()
    nm = HseNearMiss.query.filter_by(id=nm_id, officer_id=u.id).first_or_404()
    locations = _hse_locations()
    if request.method == "POST":
        nm.date            = _safe_date(request.form.get("date")) or nm.date
        nm.location        = request.form.get("location", "").strip()
        nm.description     = request.form.get("description", "").strip()
        nm.immediate_cause = request.form.get("immediate_cause", "").strip()
        nm.action_taken    = request.form.get("action_taken", "").strip()
        nm.reported_to     = request.form.get("reported_to", "").strip()
        new_photo = _save_hse_photo(request.files.get("photo"), "nm", company_id=cid())
        if new_photo:
            nm.photo_path = new_photo
        db.session.commit()
        flash("Near miss updated.", "success")
        return redirect(url_for("hse_nearmiss_list"))
    return render_template("hse_nearmiss_edit.html", nm=nm, locations=locations)


@app.route("/hse/nearmiss/<int:nm_id>/delete", methods=["POST"])
@login_required
@hse_officer_required
def hse_nearmiss_delete(nm_id):
    u = cur_user()
    nm = HseNearMiss.query.filter_by(id=nm_id, officer_id=u.id).first_or_404()
    db.session.delete(nm)
    db.session.commit()
    flash("Near miss deleted.", "success")
    return redirect(url_for("hse_nearmiss_list"))


@app.route("/hse/bbs/<int:bbs_id>/delete", methods=["POST"])
@login_required
@hse_officer_required
def hse_bbs_delete(bbs_id):
    u = cur_user()
    rec = HseBbs.query.filter_by(id=bbs_id, officer_id=u.id).first_or_404()
    db.session.delete(rec)
    db.session.commit()
    flash("BBS record deleted.", "success")
    return redirect(url_for("hse_bbs"))


# ===================== HSE Officer Export =====================

def _collect_officer_data(officer_id, date_from=None, date_to=None):
    def apply_dates(q, col):
        if date_from: q = q.filter(col >= date_from)
        if date_to:   q = q.filter(col <= date_to)
        return q
    checkins     = apply_dates(HseCheckin.query.filter_by(officer_id=officer_id),     HseCheckin.date).order_by(HseCheckin.date.desc()).all()
    observations = apply_dates(HseObservation.query.filter_by(officer_id=officer_id), HseObservation.date).order_by(HseObservation.date.desc()).all()
    jsos         = apply_dates(HseJsoClosure.query.filter_by(officer_id=officer_id),  HseJsoClosure.date).order_by(HseJsoClosure.date.desc()).all()
    tbts         = apply_dates(HseTbt.query.filter_by(officer_id=officer_id),         HseTbt.date).order_by(HseTbt.date.desc()).all()
    nearmisses   = apply_dates(HseNearMiss.query.filter_by(officer_id=officer_id),    HseNearMiss.date).order_by(HseNearMiss.date.desc()).all()
    bbs          = apply_dates(HseBbs.query.filter_by(officer_id=officer_id),         HseBbs.date).order_by(HseBbs.date.desc()).all()
    return checkins, observations, jsos, tbts, nearmisses, bbs


@app.get("/hse/dashboard/officer/<int:officer_id>/excel")
@login_required
@hse_supervisor_required
def hse_officer_detail_excel(officer_id):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response

    officer = User.query.filter_by(id=officer_id, role="safety_officer").first_or_404()
    date_from = _safe_date(request.args.get("from"))
    date_to   = _safe_date(request.args.get("to"))
    checkins, observations, jsos, tbts, nearmisses, bbs_list = _collect_officer_data(officer_id, date_from, date_to)
    _excel_tbt_ids = [t.id for t in tbts]
    _excel_att_map = dict(
        db.session.query(HseTbtAttendance.tbt_id, func.count())
        .filter(HseTbtAttendance.tbt_id.in_(_excel_tbt_ids))
        .group_by(HseTbtAttendance.tbt_id).all()
    ) if _excel_tbt_ids else {}

    wb = openpyxl.Workbook()
    thin  = Side(style="thin", color="D1D5DB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor="0F172A")
    hfont = Font(color="FFFFFF", bold=True)

    def make_sheet(title, headers, rows_data):
        ws = wb.create_sheet(title)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center"); c.border = bdr
        for ri, row in enumerate(rows_data, 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = bdr; c.alignment = Alignment(vertical="center")
        for ci in range(1, len(headers)+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

    make_sheet("Check-ins", ["Date", "Location"],
               [(c.date.isoformat(), c.location) for c in checkins])
    make_sheet("Observations", ["Date", "Type", "Category", "Risk", "Status", "Description", "Action"],
               [(o.date.isoformat(), o.obs_type, o.category, o.risk_level or "", o.status,
                 o.description or "", o.action_taken or "") for o in observations])
    make_sheet("JSO Closures", ["Date", "JSO #", "Location", "Action Taken"],
               [(j.date.isoformat(), j.jso_number, j.location or "", j.action_taken or "") for j in jsos])
    make_sheet("TBTs", ["Date", "Topic", "Location", "Attendees"],
               [(t.date.isoformat(), t.topic or "", t.location or "",
                 _excel_att_map.get(t.id, 0)) for t in tbts])
    make_sheet("Near Misses", ["Date", "Location", "Description", "Cause", "Action", "Reported To"],
               [(n.date.isoformat(), n.location or "", n.description or "",
                 n.immediate_cause or "", n.action_taken or "", n.reported_to or "") for n in nearmisses])
    make_sheet("BBS", ["Date", "Cards", "Notes"],
               [(b.date.isoformat(), b.card_count, b.notes or "") for b in bbs_list])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"officer_{officer.supervisor_code}_{officer_id}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


@app.get("/hse/dashboard/officer/<int:officer_id>/pdf")
@login_required
@hse_supervisor_required
def hse_officer_detail_pdf(officer_id):
    import io
    from flask import make_response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    officer = User.query.filter_by(id=officer_id, role="safety_officer").first_or_404()
    date_from = _safe_date(request.args.get("from"))
    date_to   = _safe_date(request.args.get("to"))
    checkins, observations, jsos, tbts, nearmisses, bbs_list = _collect_officer_data(officer_id, date_from, date_to)
    _pdf_tbt_ids = [t.id for t in tbts]
    _pdf_att_map = dict(
        db.session.query(HseTbtAttendance.tbt_id, func.count())
        .filter(HseTbtAttendance.tbt_id.in_(_pdf_tbt_ids))
        .group_by(HseTbtAttendance.tbt_id).all()
    ) if _pdf_tbt_ids else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, spaceAfter=10, textColor=colors.HexColor("#6B7280"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceBefore=10, spaceAfter=4)
    hdr_fill = colors.HexColor("#0F172A")
    alt_fill = colors.HexColor("#F9FAFB")

    def build_table(header, rows, col_widths):
        data = [header] + (rows if rows else [["—"] * len(header)])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), hdr_fill),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("ALIGN",       (0,0), (-1,-1), "LEFT"),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, alt_fill]),
            ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
            ("TOPPADDING",  (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ])
        t.setStyle(style)
        return t

    W = 186*mm
    story = [
        Paragraph(f"HSE Officer Report — {officer.name}", h1),
        Paragraph(f"Generated: {datetime.now(RIYADH_TZ).strftime('%d %b %Y %H:%M')}" +
                  (f"  |  From: {date_from}" if date_from else "") +
                  (f"  To: {date_to}" if date_to else ""), sub),
        Paragraph(f"Check-ins ({len(checkins)})", h2),
        build_table(["Date", "Location"],
                    [(c.date.isoformat(), c.location) for c in checkins],
                    [30*mm, W-30*mm]),
        Paragraph(f"Observations ({len(observations)})", h2),
        build_table(["Date", "Type", "Category", "Risk", "Status"],
                    [(o.date.isoformat(), o.obs_type, o.category or "", o.risk_level or "", o.status) for o in observations],
                    [22*mm, 28*mm, 38*mm, 16*mm, W-104*mm]),
        Paragraph(f"JSO Closures ({len(jsos)})", h2),
        build_table(["Date", "JSO #", "Action Taken"],
                    [(j.date.isoformat(), j.jso_number, (j.action_taken or "")[:80]) for j in jsos],
                    [22*mm, 30*mm, W-52*mm]),
        Paragraph(f"TBTs ({len(tbts)})", h2),
        build_table(["Date", "Topic", "Location", "Attendees"],
                    [(t.date.isoformat(), (t.topic or "")[:40], t.location or "", str(_pdf_att_map.get(t.id, 0))) for t in tbts],
                    [22*mm, 60*mm, 60*mm, W-142*mm]),
        Paragraph(f"Near Misses ({len(nearmisses)})", h2),
        build_table(["Date", "Location", "Description"],
                    [(n.date.isoformat(), n.location or "", (n.description or "")[:80]) for n in nearmisses],
                    [22*mm, 40*mm, W-62*mm]),
        Paragraph(f"BBS Cards ({len(bbs_list)})", h2),
        build_table(["Date", "Cards", "Notes"],
                    [(b.date.isoformat(), str(b.card_count), b.notes or "") for b in bbs_list],
                    [22*mm, 22*mm, W-44*mm]),
    ]

    doc.build(story)
    buf.seek(0)
    fname = f"officer_{officer_id}_report.pdf"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp



# ===================== HSE Monthly Report =====================

def _hse_monthly_report_data(year, month):
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])
    return _hse_weekly_report_data(first_day, last_day)


@app.route("/hse/reports/monthly", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_report_monthly_picker():
    today = datetime.now(RIYADH_TZ).date()
    return render_template("hse_report_monthly_picker.html", today=today)


@app.route("/hse/reports/monthly/view", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_report_monthly():
    today = datetime.now(RIYADH_TZ).date()
    try:
        year  = int(request.args.get("year",  today.year))
        month = int(request.args.get("month", today.month))
    except (TypeError, ValueError):
        return redirect(url_for("hse_report_monthly_picker"))
    from calendar import monthrange, month_name as _mn
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])
    rows = _hse_monthly_report_data(year, month)
    from datetime import datetime as _dt
    return render_template("hse_report_monthly.html",
                           rows=rows, year=year, month=month,
                           month_name=_mn[month], first_day=first_day, last_day=last_day,
                           now=_dt.now(RIYADH_TZ))


@app.get("/hse/reports/monthly/excel")
@login_required
@hse_supervisor_required
def hse_report_monthly_excel():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response
    from calendar import monthrange

    today = datetime.now(RIYADH_TZ).date()
    try:
        year  = int(request.args.get("year",  today.year))
        month = int(request.args.get("month", today.month))
    except (TypeError, ValueError):
        return redirect(url_for("hse_report_monthly_picker"))

    rows = _hse_monthly_report_data(year, month)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"{year}-{month:02d}"
    thin = Side(style="thin", color="D1D5DB")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor="0F172A")
    hfont = Font(color="FFFFFF", bold=True)
    headers = ["Officer","Days Check-in","Obs Total","Unsafe Act","Unsafe Cond","Positive",
               "High Risk","Open Obs","JSO","TBTs","TBT Attend.","Near Miss","BBS Cards",
               "PTW","Man Power","Insp Score %","Open CAs","Score"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center"); c.border = bdr
    for ri, r in enumerate(rows, 2):
        vals = [r["officer_name"], r["checkin_days"], r["obs_total"],
                r["obs_unsafe_act"], r["obs_unsafe_cond"], r["obs_positive"],
                r["obs_high"], r["obs_open"], r["jso"], r["tbt"],
                r["tbt_attendees"], r["nm"], r["bbs"],
                r.get("ptw", 0), r.get("mp_total", 0),
                r.get("avg_insp"), r.get("ca_open", 0), r.get("score", 0)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = bdr; c.alignment = Alignment(vertical="center")
    for ci, w in enumerate([24,14,10,11,12,10,11,10,9,8,13,12,11,8,11,12,10,8], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"hse_monthly_{year}_{month:02d}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


@app.get("/hse/reports/monthly/pdf")
@login_required
@hse_supervisor_required
def hse_report_monthly_pdf():
    import io
    from flask import make_response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from calendar import monthrange, month_name as _mn

    today = datetime.now(RIYADH_TZ).date()
    try:
        year  = int(request.args.get("year",  today.year))
        month = int(request.args.get("month", today.month))
    except (TypeError, ValueError):
        return redirect(url_for("hse_report_monthly_picker"))

    rows = _hse_monthly_report_data(year, month)
    first_day = date(year, month, 1)
    last_day  = date(year, month, monthrange(year, month)[1])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Heading1"], fontSize=13, spaceAfter=3)
    sub_s   = ParagraphStyle("s", parent=styles["Normal"],   fontSize=9, spaceAfter=8,
                             textColor=colors.HexColor("#6B7280"))
    story = [
        Paragraph(f"HSE Monthly Report — {_mn[month]} {year}", title_s),
        Paragraph(f"{first_day.strftime('%d %b')} — {last_day.strftime('%d %b %Y')}", sub_s),
        Spacer(1, 4*mm),
    ]
    col_headers = ["Officer","Days\nCheck-in","Obs\nTotal","Unsafe\nAct","Unsafe\nCond",
                   "Positive","High\nRisk","Open","JSO","TBTs","TBT\nAttend.","Near\nMiss","BBS",
                   "PTW","Man\nPower","Insp\n%","Open\nCAs","Score"]
    data = [col_headers] + [
        [r["officer_name"], str(r["checkin_days"]), str(r["obs_total"]),
         str(r["obs_unsafe_act"]), str(r["obs_unsafe_cond"]), str(r["obs_positive"]),
         str(r["obs_high"]), str(r["obs_open"]), str(r["jso"]),
         str(r["tbt"]), str(r["tbt_attendees"]), str(r["nm"]), str(r["bbs"]),
         str(r.get("ptw", 0)), str(r.get("mp_total", 0)),
         f"{r['avg_insp']}%" if r.get("avg_insp") is not None else "—",
         str(r.get("ca_open", 0)), str(r.get("score", 0))]
        for r in rows
    ]
    col_widths = [44*mm,14*mm,12*mm,12*mm,12*mm,12*mm,12*mm,10*mm,10*mm,10*mm,14*mm,12*mm,10*mm,
                  10*mm,12*mm,10*mm,10*mm,10*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTSIZE",   (0,0),(-1,-1), 8),
        ("ALIGN",      (1,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID",       (0,0),(-1,-1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)
    fname = f"hse_monthly_{year}_{month:02d}.pdf"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


# ===================== HSE Reports =====================

def _hse_weekly_report_data(ws, we):
    """Returns list of dicts, one per safety officer, for the given week."""
    officers = User.query.filter_by(role="safety_officer", is_active=True).order_by(User.name).all()
    rows = []
    for o in officers:
        checkin_days = HseCheckin.query.filter(
            HseCheckin.officer_id == o.id,
            HseCheckin.date.between(ws, we)
        ).count()

        obs_all = HseObservation.query.filter(
            HseObservation.officer_id == o.id,
            HseObservation.date.between(ws, we)
        ).all()
        obs_total     = len(obs_all)
        obs_unsafe_act  = sum(1 for x in obs_all if x.obs_type == "unsafe_act")
        obs_unsafe_cond = sum(1 for x in obs_all if x.obs_type == "unsafe_condition")
        obs_positive    = sum(1 for x in obs_all if x.obs_type == "positive")
        obs_high        = sum(1 for x in obs_all if x.risk_level == "H")
        obs_open        = sum(1 for x in obs_all if x.status == "open")

        jso_cnt = HseJsoClosure.query.filter(
            HseJsoClosure.officer_id == o.id,
            HseJsoClosure.date.between(ws, we)
        ).count()

        tbts = HseTbt.query.filter(
            HseTbt.officer_id == o.id,
            HseTbt.date.between(ws, we)
        ).all()
        tbt_cnt = len(tbts)
        tbt_attendees = sum(
            db.session.query(func.count(HseTbtAttendance.id)).filter_by(tbt_id=t.id).scalar() or 0
            for t in tbts
        )

        nm_cnt = HseNearMiss.query.filter(
            HseNearMiss.officer_id == o.id,
            HseNearMiss.date.between(ws, we)
        ).count()

        bbs_cards = db.session.query(
            func.coalesce(func.sum(HseBbs.card_count), 0)
        ).filter(
            HseBbs.officer_id == o.id,
            HseBbs.date.between(ws, we)
        ).scalar() or 0

        ptw_cnt = HsePtw.query.filter(
            HsePtw.officer_id == o.id,
            HsePtw.week_start <= we,
            HsePtw.week_end   >= ws,
        ).count()

        mp_total = db.session.query(
            func.coalesce(func.sum(HseManpower.total_count), 0)
        ).filter(
            HseManpower.officer_id == o.id,
            HseManpower.date.between(ws, we)
        ).scalar() or 0

        insp_list = HseInspection.query.filter(
            HseInspection.officer_id == o.id,
            HseInspection.date.between(ws, we)
        ).all()
        insp_scores = [i.overall_score for i in insp_list if i.overall_score is not None]
        avg_insp = round(sum(insp_scores) / len(insp_scores), 1) if insp_scores else None

        ca_open = HseCorrectiveAction.query.join(
            HseObservation, HseCorrectiveAction.observation_id == HseObservation.id
        ).filter(
            HseObservation.officer_id == o.id,
            HseCorrectiveAction.status != "completed"
        ).count()

        score = _hse_officer_score(o.id, ws, we)

        rows.append({
            "officer_id":     o.id,
            "officer_name":   o.name,
            "checkin_days":   checkin_days,
            "obs_total":      obs_total,
            "obs_unsafe_act": obs_unsafe_act,
            "obs_unsafe_cond": obs_unsafe_cond,
            "obs_positive":   obs_positive,
            "obs_high":       obs_high,
            "obs_open":       obs_open,
            "jso":            jso_cnt,
            "tbt":            tbt_cnt,
            "tbt_attendees":  int(tbt_attendees),
            "nm":             nm_cnt,
            "bbs":            int(bbs_cards),
            "ptw":            ptw_cnt,
            "mp_total":       int(mp_total),
            "avg_insp":       avg_insp,
            "ca_open":        ca_open,
            "score":          score,
            "total_activity": obs_total + jso_cnt + tbt_cnt + nm_cnt,
        })
    return rows


@app.route("/hse/reports", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_report_picker():
    today = datetime.now(RIYADH_TZ).date()
    days_since_sunday = today.isoweekday() % 7
    default_ws = today - timedelta(days=days_since_sunday)
    return render_template("hse_report_picker.html", default_ws=default_ws)


@app.route("/hse/reports/weekly", methods=["GET"])
@login_required
@hse_supervisor_required
def hse_report_weekly():
    ws = _safe_date(request.args.get("week_start"))
    if not ws:
        return redirect(url_for("hse_report_picker"))
    we = ws + timedelta(days=4)          # Sun → Thu
    rows = _hse_weekly_report_data(ws, we)
    from datetime import datetime as _dt
    return render_template("hse_report_weekly.html",
                           rows=rows, ws=ws, we=we,
                           now=_dt.now(RIYADH_TZ))


# ===================== HSE Phase 2 Routes =====================

def _ptw_current_week():
    today = datetime.now(RIYADH_TZ).date()
    days_since_sun = (today.weekday() - 6) % 7   # الأحد = weekday 6 في Python
    week_start = today - timedelta(days=days_since_sun)
    week_end   = week_start + timedelta(days=4)   # الأحد + 4 = الخميس
    return week_start, week_end


def _hse_officer_score(officer_id, week_start, week_end):
    ci  = HseCheckin.query.filter(HseCheckin.officer_id == officer_id,
                                   HseCheckin.date.between(week_start, week_end)).count()
    obs = HseObservation.query.filter(HseObservation.officer_id == officer_id,
                                       HseObservation.date.between(week_start, week_end)).count()
    tbt = HseTbt.query.filter(HseTbt.officer_id == officer_id,
                               HseTbt.date.between(week_start, week_end)).count()
    nm  = HseNearMiss.query.filter(HseNearMiss.officer_id == officer_id,
                                    HseNearMiss.date.between(week_start, week_end)).count()
    bbs = db.session.query(func.coalesce(func.sum(HseBbs.card_count), 0)).filter(
           HseBbs.officer_id == officer_id,
           HseBbs.date.between(week_start, week_end)).scalar() or 0
    insp = HseInspection.query.filter(HseInspection.officer_id == officer_id,
                                       HseInspection.date.between(week_start, week_end)).count()
    ptw  = HsePtw.query.filter(HsePtw.officer_id == officer_id,
                                 HsePtw.week_start == week_start).count()
    return round(ci*1 + obs*2 + tbt*3 + nm*5 + bbs*0.1 + insp*2 + ptw*1, 1)


# ── PTW ──────────────────────────────────────────────────────────────

@app.route("/hse/ptw", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_ptw():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    ws, we = _ptw_current_week()

    if request.method == "POST":
        permit_number = (request.form.get("permit_number") or "").strip()
        permit_type   = (request.form.get("permit_type") or "").strip()
        description   = (request.form.get("description") or "").strip()
        location      = (request.form.get("location") or "").strip() or _hse_today_location(u.id)
        w_start       = _safe_date(request.form.get("week_start")) or ws
        w_end         = w_start + timedelta(days=5)
        attached_to   = request.form.get("attached_to_id") or None

        if not permit_number or not permit_type:
            flash("رقم البيرمت والنوع مطلوبان.", "danger")
            return redirect(url_for("hse_ptw"))

        p = HsePtw(
            officer_id=u.id, company_id=u.company_id,
            permit_number=permit_number, permit_type=permit_type,
            description=description, location=location,
            week_start=w_start, week_end=w_end,
            attached_to_id=int(attached_to) if attached_to else None,
        )
        db.session.add(p)
        db.session.commit()
        flash("تم حفظ البيرمت.", "success")
        return redirect(url_for("hse_ptw"))

    ptw_list = HsePtw.query.filter_by(officer_id=u.id).order_by(HsePtw.week_start.desc()).all()
    active_ptw = [p for p in ptw_list if p.status == "active"]
    loc = _hse_today_location(u.id)
    locations = _hse_locations()
    permit_numbers = [r[0] for r in db.session.query(HsePtw.permit_number).filter_by(
        company_id=u.company_id).distinct().all()]
    return render_template("hse_ptw.html", ptw_list=ptw_list, active_ptw=active_ptw,
                           ws=ws, we=we, today=today, loc=loc,
                           locations=locations, permit_numbers=permit_numbers)


@app.post("/hse/ptw/<int:ptw_id>/status")
@login_required
@hse_officer_required
def hse_ptw_status(ptw_id):
    u = cur_user()
    p = HsePtw.query.filter_by(id=ptw_id, officer_id=u.id).first_or_404()
    new_status = request.form.get("status", "active")
    if new_status in ("active", "suspended", "closed"):
        p.status = new_status
        db.session.commit()
    return redirect(url_for("hse_ptw"))


@app.post("/hse/ptw/<int:ptw_id>/renew")
@login_required
@hse_officer_required
def hse_ptw_renew(ptw_id):
    u = cur_user()
    p = HsePtw.query.filter_by(id=ptw_id, officer_id=u.id).first_or_404()
    ws, we = _ptw_current_week()
    p.week_start = ws
    p.week_end   = we
    p.status     = "active"
    db.session.commit()
    flash("تم تجديد البيرمت للأسبوع الحالي.", "success")
    return redirect(url_for("hse_ptw"))


@app.get("/hse/ptw/suggest")
@login_required
def hse_ptw_suggest():
    u = cur_user()
    q = (request.args.get("q") or "").strip()
    rows = (HsePtw.query
            .filter(HsePtw.company_id == u.company_id,
                    HsePtw.permit_number.ilike(f"%{q}%"))
            .with_entities(HsePtw.permit_number, HsePtw.permit_type,
                           HsePtw.description, HsePtw.location)
            .distinct().limit(10).all())
    return __import__("flask").jsonify([
        {"number": r[0], "type": r[1], "desc": r[2] or "", "loc": r[3] or ""}
        for r in rows
    ])


# ── Man Power ─────────────────────────────────────────────────────────

@app.route("/hse/manpower", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_manpower():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()

    existing = HseManpower.query.filter_by(officer_id=u.id, date=today).first()

    if request.method == "POST":
        total = request.form.get("total_count", "0")
        notes = (request.form.get("notes") or "").strip()
        loc   = (request.form.get("location") or "").strip() or _hse_today_location(u.id)

        import json as _json
        trades  = request.form.getlist("trade[]")
        counts  = request.form.getlist("count[]")
        bkdown  = {t.strip(): int(c) for t, c in zip(trades, counts)
                   if t.strip() and c.isdigit() and int(c) > 0}

        try:
            total = max(0, int(total))
        except ValueError:
            total = sum(bkdown.values())

        if existing:
            existing.total_count = total
            existing.location    = loc
            existing.breakdown   = _json.dumps(bkdown, ensure_ascii=False) if bkdown else None
            existing.notes       = notes
        else:
            rec = HseManpower(officer_id=u.id, company_id=u.company_id,
                               date=today, total_count=total, location=loc,
                               breakdown=_json.dumps(bkdown, ensure_ascii=False) if bkdown else None,
                               notes=notes)
            db.session.add(rec)
        db.session.commit()
        flash("تم حفظ Man Power.", "success")
        return redirect(url_for("hse_manpower"))

    history = (HseManpower.query.filter_by(officer_id=u.id)
               .order_by(HseManpower.date.desc()).limit(14).all())
    loc = _hse_today_location(u.id)
    locations = _hse_locations()
    return render_template("hse_manpower.html", existing=existing, history=history,
                           today=today, loc=loc, locations=locations)


# ── Inspection Checklist ──────────────────────────────────────────────

@app.route("/hse/inspection", methods=["GET", "POST"])
@login_required
@hse_officer_required
def hse_inspection():
    u = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    existing = HseInspection.query.filter_by(officer_id=u.id, date=today).first()

    if request.method == "POST":
        import json as _json
        loc   = (request.form.get("location") or "").strip() or _hse_today_location(u.id)
        notes = (request.form.get("notes") or "").strip()
        items = []
        ok_count = 0
        for i, item_text in enumerate(INSPECTION_ITEMS):
            ok   = request.form.get(f"item_{i}") == "ok"
            note = (request.form.get(f"note_{i}") or "").strip()
            items.append({"item": item_text, "ok": ok, "note": note})
            if ok:
                ok_count += 1
        score = round(ok_count / len(INSPECTION_ITEMS) * 100, 1)

        if existing:
            existing.location      = loc
            existing.checklist     = _json.dumps(items, ensure_ascii=False)
            existing.overall_score = score
            existing.notes         = notes
        else:
            rec = HseInspection(officer_id=u.id, company_id=u.company_id,
                                 date=today, location=loc,
                                 checklist=_json.dumps(items, ensure_ascii=False),
                                 overall_score=score, notes=notes)
            db.session.add(rec)
        db.session.commit()
        flash(f"تم حفظ الـ Checklist — النتيجة: {score}%", "success")
        return redirect(url_for("hse_inspection"))

    import json as _json
    existing_items = None
    if existing and existing.checklist:
        try:
            existing_items = _json.loads(existing.checklist)
        except Exception:
            existing_items = None

    history = (HseInspection.query.filter_by(officer_id=u.id)
               .order_by(HseInspection.date.desc()).limit(10).all())
    loc = _hse_today_location(u.id)
    locations = _hse_locations()
    return render_template("hse_inspection.html",
                           existing=existing, existing_items=existing_items,
                           items=INSPECTION_ITEMS, history=history,
                           today=today, loc=loc, locations=locations)


# ── Corrective Actions ────────────────────────────────────────────────

@app.post("/hse/corrective/add")
@login_required
@hse_supervisor_required
def hse_corrective_add():
    obs_id         = request.form.get("observation_id")
    assigned_to    = (request.form.get("assigned_to") or "").strip()
    due_date       = _safe_date(request.form.get("due_date"))
    action_required = (request.form.get("action_required") or "").strip()

    if not obs_id or not due_date or not action_required:
        flash("جميع الحقول مطلوبة.", "danger")
        return redirect(url_for("hse_dashboard"))

    u = cur_user()
    obs = HseObservation.query.get_or_404(int(obs_id))
    ca = HseCorrectiveAction(
        observation_id=int(obs_id), company_id=u.company_id,
        assigned_to=assigned_to, due_date=due_date,
        action_required=action_required, created_by=u.id,
    )
    db.session.add(ca)
    db.session.commit()
    # notify supervisors
    _notify_hse_supervisors(
        "📋 Corrective Action Added",
        f"New CA assigned to {assigned_to or 'N/A'} — due {due_date.strftime('%d %b %Y')}"
    )
    # notify the officer who owns the observation
    import threading as _thr
    _thr.Thread(
        target=_send_push_to_user,
        args=(obs.officer_id,
              "📋 إجراء تصحيحي جديد",
              f"تم فتح CA على ملاحظتك — مطلوب: {action_required[:80]}"),
        daemon=True,
    ).start()
    flash("تم إضافة الإجراء التصحيحي.", "success")
    return redirect(url_for("hse_dashboard"))


@app.post("/hse/corrective/<int:ca_id>/update")
@login_required
@hse_supervisor_required
def hse_corrective_update(ca_id):
    u = cur_user()
    ca = HseCorrectiveAction.query.get_or_404(ca_id)
    status = request.form.get("status", "open")
    if status in ("open", "in_progress", "completed"):
        ca.status = status
        if status == "completed":
            ca.completed_at     = datetime.now(RIYADH_TZ).date()
            ca.completion_notes = (request.form.get("completion_notes") or "").strip()
        db.session.commit()
        flash("تم تحديث الإجراء.", "success")
    return redirect(url_for("hse_dashboard"))

# ── Officer: My Corrective Actions ───────────────────────────────────

@app.get("/hse/my-ca")
@login_required
@hse_officer_required
def hse_my_ca():
    u    = cur_user()
    today = datetime.now(RIYADH_TZ).date()
    cas  = (HseCorrectiveAction.query
            .join(HseObservation, HseCorrectiveAction.observation_id == HseObservation.id)
            .filter(HseObservation.officer_id == u.id)
            .order_by(HseCorrectiveAction.due_date.asc())
            .all())
    open_cas     = [c for c in cas if c.status != "completed"]
    closed_cas   = [c for c in cas if c.status == "completed"]
    overdue_cas  = [c for c in open_cas if c.due_date < today]
    return render_template("hse_my_ca.html",
                           open_cas=open_cas, closed_cas=closed_cas,
                           overdue_cas=overdue_cas, today=today)


# ── Charts API ────────────────────────────────────────────────────────

@app.get("/api/hse/charts")
@login_required
@hse_supervisor_required
def hse_charts_data():
    import json as _json
    _cid = cid()
    today = datetime.now(RIYADH_TZ).date()

    # آخر 4 أسابيع
    weeks = []
    for i in range(3, -1, -1):
        days_since_sun = today.isoweekday() % 7
        ws = today - timedelta(days=days_since_sun + i * 7)
        we = ws + timedelta(days=4)
        weeks.append((ws, we))

    obs_trend = {"labels": [], "unsafe_act": [], "unsafe_cond": [], "positive": []}
    for ws, we in weeks:
        obs_trend["labels"].append(ws.strftime("%d %b"))
        q = HseObservation.query.filter(HseObservation.date.between(ws, we))
        if _cid: q = q.filter(HseObservation.company_id == _cid)
        obs_trend["unsafe_act"].append(q.filter_by(obs_type="unsafe_act").count())
        obs_trend["unsafe_cond"].append(q.filter_by(obs_type="unsafe_condition").count())
        obs_trend["positive"].append(q.filter_by(obs_type="positive").count())

    # Category distribution this month
    first_of_month = today.replace(day=1)
    cat_q = db.session.query(HseObservation.category, func.count(HseObservation.id))\
              .filter(HseObservation.date >= first_of_month)
    if _cid: cat_q = cat_q.filter(HseObservation.company_id == _cid)
    cat_data = {r[0] or "Other": r[1] for r in cat_q.group_by(HseObservation.category).all()}

    # Officer scores this week
    days_since_sun = today.isoweekday() % 7
    ws_cur = today - timedelta(days=days_since_sun)
    we_cur = ws_cur + timedelta(days=4)
    officers_q = User.query.filter_by(role="safety_officer", is_active=True)
    if _cid: officers_q = officers_q.filter(User.company_id == _cid)
    officers = officers_q.all()
    officer_scores = {
        "labels":  [o.name for o in officers],
        "scores":  [_hse_officer_score(o.id, ws_cur, we_cur) for o in officers],
    }

    # Man power last 7 days
    mp_data = {"labels": [], "counts": []}
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        mp_data["labels"].append(d.strftime("%a %d"))
        total_q = db.session.query(func.coalesce(func.sum(HseManpower.total_count), 0))\
                    .filter(HseManpower.date == d)
        if _cid: total_q = total_q.filter(HseManpower.company_id == _cid)
        mp_data["counts"].append(int(total_q.scalar() or 0))

    return __import__("flask").jsonify({
        "obs_trend": obs_trend,
        "cat_data": cat_data,
        "officer_scores": officer_scores,
        "mp_data": mp_data,
    })


@app.get("/hse/reports/weekly/excel")
@login_required
@hse_supervisor_required
def hse_report_weekly_excel():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response

    ws_date = _safe_date(request.args.get("week_start"))
    if not ws_date:
        return redirect(url_for("hse_report_picker"))
    we_date = ws_date + timedelta(days=4)
    rows = _hse_weekly_report_data(ws_date, we_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HSE Weekly"

    headers = ["Officer", "Check-in Days",
               "Obs Total", "Unsafe Act", "Unsafe Cond", "Positive", "High-Risk", "Open",
               "JSO Closures", "TBTs", "TBT Attendees", "Near Misses", "BBS Cards"]
    hfill = PatternFill("solid", fgColor="0F172A")
    hfont = Font(color="FFFFFF", bold=True)
    thin  = Side(style="thin", color="D1D5DB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center"); c.border = bdr

    for ri, r in enumerate(rows, 2):
        vals = [r["officer_name"], r["checkin_days"],
                r["obs_total"], r["obs_unsafe_act"], r["obs_unsafe_cond"],
                r["obs_positive"], r["obs_high"], r["obs_open"],
                r["jso"], r["tbt"], r["tbt_attendees"], r["nm"], r["bbs"]]
        zero_fill = PatternFill("solid", fgColor="FEF2F2") if r["total_activity"] == 0 else None
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = bdr
            c.alignment = Alignment(vertical="center")
            if zero_fill:
                c.fill = zero_fill

    widths = [24, 14, 10, 11, 12, 10, 11, 8, 13, 8, 14, 13, 11]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"hse_weekly_{ws_date.isoformat()}.xlsx"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


@app.get("/hse/reports/weekly/pdf")
@login_required
@hse_supervisor_required
def hse_report_weekly_pdf():
    import io
    from flask import make_response
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    ws_date = _safe_date(request.args.get("week_start"))
    if not ws_date:
        return redirect(url_for("hse_report_picker"))
    we_date = ws_date + timedelta(days=4)
    rows = _hse_weekly_report_data(ws_date, we_date)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Heading1"], fontSize=13, spaceAfter=3)
    sub_s   = ParagraphStyle("s", parent=styles["Normal"],   fontSize=9,  spaceAfter=8, textColor=colors.HexColor("#6B7280"))

    story = [
        Paragraph("HSE Weekly Report", title_s),
        Paragraph(f"Week: {ws_date.strftime('%d %b %Y')} — {we_date.strftime('%d %b %Y')}", sub_s),
        Spacer(1, 4*mm),
    ]

    col_headers = ["Officer", "Days\nChecked In",
                   "Obs\nTotal", "Unsafe\nAct", "Unsafe\nCond", "Positive", "High\nRisk", "Open",
                   "JSO", "TBTs", "TBT\nAttend.", "Near\nMiss", "BBS\nCards"]
    data = [col_headers]
    for r in rows:
        data.append([
            r["officer_name"], str(r["checkin_days"]),
            str(r["obs_total"]), str(r["obs_unsafe_act"]), str(r["obs_unsafe_cond"]),
            str(r["obs_positive"]), str(r["obs_high"]), str(r["obs_open"]),
            str(r["jso"]), str(r["tbt"]), str(r["tbt_attendees"]),
            str(r["nm"]), str(r["bbs"]),
        ])

    col_widths = [50*mm, 16*mm, 14*mm, 14*mm, 14*mm, 14*mm, 14*mm, 12*mm,
                  14*mm, 12*mm, 16*mm, 14*mm, 14*mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("FONTSIZE",    (0,1), (-1,-1), 8),
        ("ALIGN",       (1,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
    ])
    for ri, r in enumerate(rows, 1):
        if r["total_activity"] == 0:
            style.add("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#FEF2F2"))
    t.setStyle(style)
    story.append(t)

    doc.build(story)
    buf.seek(0)
    fname = f"hse_weekly_{ws_date.isoformat()}.pdf"
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename={fname}"
    return resp


# ===================== HSE Mobile API =====================

def api_hse_officer_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = get_api_user()
        if not u or not u.is_active or u.role != "safety_officer":
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def api_hse_supervisor_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        u = get_api_user()
        if not u or not u.is_active or not is_hse_supervisor(u):
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper




# ── API: Corrective Action (iOS) ──────────────────────────────────────

@app.post("/api/hse/corrective_action")
@api_hse_supervisor_required
def api_hse_corrective_action_add():
    u = get_api_user()
    data = freq.get_json(force=True) or {}
    obs_id         = data.get("observation_id")
    assigned_to    = (data.get("assigned_to") or "").strip()
    due_date_str   = (data.get("due_date") or "").strip()
    action_required = (data.get("action_required") or "").strip()

    if not all([obs_id, due_date_str, action_required]):
        return jsonify({"error": "observation_id, due_date, action_required are required"}), 400

    try:
        due_date = parse_date(due_date_str)
    except Exception:
        return jsonify({"error": "invalid due_date format, use YYYY-MM-DD"}), 400

    obs = HseObservation.query.filter_by(id=int(obs_id), company_id=u.company_id).first()
    if not obs:
        return jsonify({"error": "observation not found"}), 404

    ca = HseCorrectiveAction(
        observation_id=obs.id,
        company_id=u.company_id,
        assigned_to=assigned_to or None,
        due_date=due_date,
        action_required=action_required,
        created_by=u.id,
    )
    db.session.add(ca)
    db.session.commit()

    import threading as _thr
    _thr.Thread(
        target=_send_push_to_user,
        args=(obs.officer_id,
              "📋 إجراء تصحيحي جديد",
              f"تم فتح CA على ملاحظتك — مطلوب: {action_required[:80]}"),
        daemon=True,
    ).start()

    return jsonify({"ok": True, "id": ca.id}), 201


@app.put("/api/hse/corrective_action/<int:ca_id>")
@api_hse_supervisor_required
def api_hse_corrective_action_update(ca_id):
    u = get_api_user()
    ca = HseCorrectiveAction.query.filter_by(id=ca_id, company_id=u.company_id).first_or_404()
    data = freq.get_json(force=True) or {}
    status = data.get("status", ca.status)
    if status not in ("open", "in_progress", "completed"):
        return jsonify({"error": "invalid status"}), 400
    ca.status = status
    if status == "completed":
        ca.completed_at     = datetime.now(RIYADH_TZ).date()
        ca.completion_notes = (data.get("completion_notes") or "").strip() or None
    db.session.commit()
    return jsonify({"ok": True})


@app.get("/api/hse/officer/<int:officer_id>/pdf")
@api_hse_supervisor_required
def api_hse_officer_pdf(officer_id):
    """PDF officer report via Bearer token (iOS)."""
    import io
    from flask import make_response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    officer = User.query.filter_by(id=officer_id, role="safety_officer").first_or_404()
    date_from = _safe_date(freq.args.get("from"))
    date_to   = _safe_date(freq.args.get("to"))
    checkins, observations, jsos, tbts, nearmisses, bbs_list = _collect_officer_data(officer_id, date_from, date_to)
    _att_ids = [t.id for t in tbts]
    _att_map = dict(
        db.session.query(HseTbtAttendance.tbt_id, func.count())
        .filter(HseTbtAttendance.tbt_id.in_(_att_ids))
        .group_by(HseTbtAttendance.tbt_id).all()
    ) if _att_ids else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    h1  = ParagraphStyle("h1",  parent=styles["Heading1"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"],   fontSize=9,  spaceAfter=10,
                         textColor=colors.HexColor("#6B7280"))
    h2  = ParagraphStyle("h2",  parent=styles["Heading2"], fontSize=11, spaceBefore=10, spaceAfter=4)
    hdr_fill = colors.HexColor("#0F172A")
    alt_fill = colors.HexColor("#F9FAFB")

    def _tbl(header, rows, col_widths):
        data = [header] + (rows if rows else [["-"] * len(header)])
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), hdr_fill),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTSIZE",    (0,0), (-1,0), 9),
            ("FONTSIZE",    (0,1), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, alt_fill]),
            ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#E5E7EB")),
            ("LEFTPADDING",  (0,0), (-1,-1), 5),
            ("RIGHTPADDING", (0,0), (-1,-1), 5),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ]))
        return t

    content = [
        Paragraph("HSE Officer Report - " + officer.name, h1),
        Paragraph("Period: " + str(date_from or "All time") + " to " + str(date_to or "Today"), sub),
        Spacer(1, 4*mm),
        Paragraph("Check-in History", h2),
        _tbl(["Date","Location","Time"],
             [[str(c.date), c.location or "-", str(c.created_at)[:10]] for c in checkins],
             [40*mm, 80*mm, 40*mm]),
        Spacer(1, 4*mm),
        Paragraph("Observations", h2),
        _tbl(["Date","Type","Category","Risk","Status"],
             [[str(o.date), o.obs_type or "-", o.category or "-", o.risk_level or "-", o.status]
              for o in observations],
             [28*mm, 32*mm, 36*mm, 20*mm, 24*mm]),
        Spacer(1, 4*mm),
        Paragraph("TBT Sessions", h2),
        _tbl(["Date","Topic","Location","Attendees"],
             [[str(t.date), (t.topic or "-")[:40], t.location or "-",
               str(_att_map.get(t.id, 0))] for t in tbts],
             [25*mm, 70*mm, 45*mm, 20*mm]),
        Spacer(1, 4*mm),
        Paragraph("Near Miss", h2),
        _tbl(["Date","Location","Description"],
             [[str(n.date), n.location or "-", (n.description or "-")[:60]] for n in nearmisses],
             [25*mm, 40*mm, 95*mm]),
    ]
    doc.build(content)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f"attachment; filename=officer_{officer_id}_report.pdf"
    return resp


# ── API: Weekly Report JSON (for iOS) ────────────────────────────────
@app.get("/api/hse/reports/weekly")
@api_hse_supervisor_required
def api_hse_report_weekly():
    ws = _safe_date(freq.args.get("week_start"))
    if not ws:
        today = datetime.now(RIYADH_TZ).date()
        days_since_sunday = today.isoweekday() % 7
        ws = today - timedelta(days=days_since_sunday)
    we = ws + timedelta(days=6)
    rows = _hse_weekly_report_data(ws, we)
    prev_ws = ws - timedelta(days=7)
    prev_we = prev_ws + timedelta(days=6)
    prev_rows = _hse_weekly_report_data(prev_ws, prev_we)
    return jsonify({"week_start": ws.isoformat(), "week_end": we.isoformat(),
                    "rows": rows, "prev_rows": prev_rows})


# ── API: Monthly Report JSON (for iOS) ───────────────────────────────
@app.get("/api/hse/reports/monthly")
@api_hse_supervisor_required
def api_hse_report_monthly_api():
    today = datetime.now(RIYADH_TZ).date()
    try:
        year  = int(freq.args.get("year",  today.year))
        month = int(freq.args.get("month", today.month))
    except (TypeError, ValueError):
        year, month = today.year, today.month
    rows = _hse_monthly_report_data(year, month)
    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    prev_rows = _hse_monthly_report_data(prev_year, prev_month)
    return jsonify({"year": year, "month": month, "rows": rows, "prev_rows": prev_rows})


@app.route("/api/hse/locations")
@api_login_required
def api_hse_locations():
    locs = db.session.query(HseCheckin.location).distinct().all()
    return jsonify([l[0] for l in locs if l[0]])

# ── Check-in ──────────────────────────────────────────
@app.route("/api/hse/checkin/today", methods=["GET"])
@api_hse_officer_required
def api_hse_checkin_today():
    u = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    ci = HseCheckin.query.filter_by(officer_id=u.id, date=today).first()
    if not ci:
        return jsonify(None)
    return jsonify({"id": ci.id, "date": ci.date.isoformat(), "location": ci.location})

@app.route("/api/hse/checkin", methods=["POST"])
@api_hse_officer_required
def api_hse_checkin_api():
    u = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    data = freq.get_json()
    location = (data.get("location") or "").strip()
    if not location:
        return jsonify({"error": "Location required"}), 400
    ci = HseCheckin.query.filter_by(officer_id=u.id, date=today).first()
    if ci:
        ci.location = location
    else:
        ci = HseCheckin(officer_id=u.id, date=today, location=location)
        db.session.add(ci)
    db.session.commit()
    return jsonify({"ok": True, "location": ci.location})

# ── Observations ──────────────────────────────────────
@app.route("/api/hse/observations", methods=["GET"])
@api_hse_officer_required
def api_hse_observations():
    u = get_api_user()
    page = int(freq.args.get("page", 1))
    status = freq.args.get("status", "")
    q = HseObservation.query.filter_by(officer_id=u.id)
    if status in ("open", "closed"):
        q = q.filter_by(status=status)
    pg = q.order_by(HseObservation.date.desc()).paginate(page=page, per_page=15, error_out=False)
    items = [{"id": o.id, "date": o.date.isoformat(), "location": o.location or "",
              "obs_type": o.obs_type or "", "category": o.category or "",
              "risk_level": o.risk_level or "", "description": o.description or "",
              "action_taken": o.action_taken or "", "status": o.status,
              "closed_at": o.closed_at.isoformat() if o.closed_at else None,
              "closure_action": o.closure_action or ""} for o in pg.items]
    return jsonify({"items": items, "page": pg.page, "pages": pg.pages, "total": pg.total})

@app.route("/api/hse/observation", methods=["POST"])
@api_hse_officer_required
def api_hse_observation_create():
    u = get_api_user()
    data = freq.get_json()
    today = datetime.now(RIYADH_TZ).date()
    date_val = _parse_date(data.get("date"), today)
    risk_level = data.get("risk_level", "L")
    location = (data.get("location") or "").strip()
    category = (data.get("category") or "").strip()
    o = HseObservation(
        officer_id=u.id,
        date=date_val,
        location=location,
        obs_type=data.get("obs_type", "unsafe_act"),
        category=category,
        risk_level=risk_level,
        description=(data.get("description") or "").strip(),
        action_taken=(data.get("action_taken") or "").strip(),
        company_id=api_cid(),
    )
    db.session.add(o)
    db.session.commit()
    if risk_level == "H":
        _notify_hse_supervisors("⚠ High-Risk Observation",
                                f"{u.name}: {category or 'No category'} at {location}")
    return jsonify({"ok": True, "id": o.id})

@app.route("/api/hse/observation/<int:obs_id>/close", methods=["POST"])
@api_hse_officer_required
def api_hse_observation_close(obs_id):
    u = get_api_user()
    o = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    data = freq.get_json()
    closure_action = (data.get("closure_action") or "").strip()
    if not closure_action:
        return jsonify({"error": "Closure action required"}), 400
    o.status = "closed"
    o.closed_at = datetime.now(RIYADH_TZ).date()
    o.closure_action = closure_action
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/hse/observation/<int:obs_id>", methods=["PUT"])
@api_hse_officer_required
def api_hse_observation_update(obs_id):
    u = get_api_user()
    o = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    data = freq.get_json()
    old_risk = o.risk_level
    o.location    = (data.get("location")    or o.location or "").strip()
    o.obs_type    = data.get("obs_type")    or o.obs_type
    o.category    = (data.get("category")    or o.category or "").strip()
    o.risk_level  = data.get("risk_level")  or o.risk_level
    o.description = (data.get("description") or o.description or "").strip()
    o.action_taken = (data.get("action_taken") or o.action_taken or "").strip()
    db.session.commit()
    if o.risk_level == "H" and old_risk != "H":
        _notify_hse_supervisors("⚠ High-Risk Observation (updated)",
                                f"{u.name}: {o.category or 'No category'} at {o.location}")
    return jsonify({"ok": True})

@app.route("/api/hse/observation/<int:obs_id>/photo", methods=["POST"])
@api_hse_officer_required
def api_hse_observation_photo(obs_id):
    u = get_api_user()
    o = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    f = freq.files.get("photo")
    path = _save_hse_photo(f, "obs", company_id=api_cid())
    if not path:
        return jsonify({"error": "Invalid or missing photo"}), 400
    photo_type = freq.form.get("photo_type", "before")
    db.session.add(HseObservationPhoto(observation_id=o.id, photo_path=path, photo_type=photo_type))
    db.session.commit()
    return jsonify({"ok": True, "path": path})

@app.route("/api/hse/observation/<int:obs_id>", methods=["DELETE"])
@api_hse_officer_required
def api_hse_observation_delete_api(obs_id):
    u = get_api_user()
    o = HseObservation.query.filter_by(id=obs_id, officer_id=u.id).first_or_404()
    HseObservationPhoto.query.filter_by(observation_id=o.id).delete()
    db.session.delete(o)
    db.session.commit()
    return jsonify({"ok": True})

# ── JSO Closure ────────────────────────────────────────
@app.route("/api/hse/jso", methods=["GET"])
@api_hse_officer_required
def api_hse_jso_list():
    u = get_api_user()
    page = int(freq.args.get("page", 1))
    pg = HseJsoClosure.query.filter_by(officer_id=u.id)\
        .order_by(HseJsoClosure.date.desc()).paginate(page=page, per_page=15, error_out=False)
    items = [{"id": j.id, "date": j.date.isoformat(), "jso_number": j.jso_number,
              "location": j.location or "", "action_taken": j.action_taken or ""} for j in pg.items]
    return jsonify({"items": items, "page": pg.page, "pages": pg.pages, "total": pg.total})

@app.route("/api/hse/jso", methods=["POST"])
@api_hse_officer_required
def api_hse_jso_create():
    u = get_api_user()
    data = freq.get_json()
    today = datetime.now(RIYADH_TZ).date()
    jso_number = (data.get("jso_number") or "").strip()
    if not jso_number:
        return jsonify({"error": "JSO number required"}), 400
    j = HseJsoClosure(
        officer_id=u.id,
        date=_parse_date(data.get("date"), today),
        jso_number=jso_number,
        location=(data.get("location") or "").strip(),
        action_taken=(data.get("action_taken") or "").strip(),
        company_id=api_cid(),
    )
    db.session.add(j)
    db.session.commit()
    return jsonify({"ok": True, "id": j.id})

@app.route("/api/hse/jso/<int:jso_id>", methods=["PUT"])
@api_hse_officer_required
def api_hse_jso_update(jso_id):
    u = get_api_user()
    j = HseJsoClosure.query.filter_by(id=jso_id, officer_id=u.id).first_or_404()
    data = freq.get_json()
    j.jso_number  = (data.get("jso_number")  or j.jso_number).strip()
    j.location    = (data.get("location")    or j.location or "").strip()
    j.action_taken = (data.get("action_taken") or j.action_taken or "").strip()
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/hse/jso/<int:jso_id>/photo", methods=["POST"])
@api_hse_officer_required
def api_hse_jso_photo(jso_id):
    u = get_api_user()
    j = HseJsoClosure.query.filter_by(id=jso_id, officer_id=u.id).first_or_404()
    f = freq.files.get("photo")
    path = _save_hse_photo(f, "jso", company_id=api_cid())
    if not path:
        return jsonify({"error": "Invalid or missing photo"}), 400
    if j.photo_path:
        try: os.remove(os.path.join(HSE_UPLOAD_DIR, j.photo_path))
        except Exception: pass
    j.photo_path = path
    db.session.commit()
    return jsonify({"ok": True, "path": path})

@app.route("/api/hse/jso/<int:jso_id>", methods=["DELETE"])
@api_hse_officer_required
def api_hse_jso_delete_api(jso_id):
    u = get_api_user()
    j = HseJsoClosure.query.filter_by(id=jso_id, officer_id=u.id).first_or_404()
    db.session.delete(j)
    db.session.commit()
    return jsonify({"ok": True})

# ── TBT ────────────────────────────────────────────────
@app.route("/api/hse/tbt", methods=["GET"])
@api_hse_officer_required
def api_hse_tbt_list():
    u = get_api_user()
    page = int(freq.args.get("page", 1))
    pg = HseTbt.query.filter_by(officer_id=u.id)\
        .order_by(HseTbt.date.desc()).paginate(page=page, per_page=15, error_out=False)
    items = []
    for t in pg.items:
        att = HseTbtAttendance.query.filter_by(tbt_id=t.id).all()
        sup = db.session.get(User, t.supervisor_id) if t.supervisor_id else None
        items.append({"id": t.id, "date": t.date.isoformat(), "topic": t.topic or "",
                      "location": t.location or "", "attendee_count": len(att),
                      "supervisor_name": sup.name if sup else "",
                      "attendees": [{"emp_number": a.emp_number, "emp_name": a.emp_name} for a in att]})
    return jsonify({"items": items, "page": pg.page, "pages": pg.pages, "total": pg.total})

@app.route("/api/hse/supervisor_lookup")
@api_login_required
def api_hse_supervisor_lookup():
    code = freq.args.get("code", "").strip()
    if not code:
        return jsonify({"found": False})
    sup = User.query.filter(User.supervisor_code == code, User.is_active == True).first()
    if sup:
        return jsonify({"found": True, "name": sup.name, "id": sup.id})
    return jsonify({"found": False})

@app.route("/api/hse/tbt", methods=["POST"])
@api_hse_officer_required
def api_hse_tbt_create():
    u = get_api_user()
    data = freq.get_json()
    today = datetime.now(RIYADH_TZ).date()
    topic = (data.get("topic") or "").strip()
    if not topic:
        return jsonify({"error": "Topic required"}), 400
    # Resolve supervisor by code if provided
    sup_id = None
    sup_code = (data.get("supervisor_code") or "").strip()
    if sup_code:
        sup = User.query.filter(User.supervisor_code == sup_code, User.is_active == True).first()
        if sup:
            sup_id = sup.id
    t = HseTbt(officer_id=u.id, date=_parse_date(data.get("date"), today),
               topic=topic, location=(data.get("location") or "").strip(),
               supervisor_id=sup_id, company_id=api_cid())
    db.session.add(t)
    db.session.flush()
    for att in (data.get("attendees") or []):
        db.session.add(HseTbtAttendance(tbt_id=t.id,
                                        emp_number=att.get("emp_number", ""),
                                        emp_name=att.get("emp_name", "")))
    db.session.commit()
    return jsonify({"ok": True, "id": t.id})

@app.route("/api/hse/tbt/<int:tbt_id>", methods=["PUT"])
@api_hse_officer_required
def api_hse_tbt_update(tbt_id):
    u = get_api_user()
    t = HseTbt.query.filter_by(id=tbt_id, officer_id=u.id).first_or_404()
    data = freq.get_json()
    topic = (data.get("topic") or "").strip()
    if not topic:
        return jsonify({"error": "Topic required"}), 400
    t.topic    = topic
    t.location = (data.get("location") or t.location or "").strip()
    HseTbtAttendance.query.filter_by(tbt_id=t.id).delete()
    for att in (data.get("attendees") or []):
        db.session.add(HseTbtAttendance(tbt_id=t.id,
                                        emp_number=att.get("emp_number", ""),
                                        emp_name=att.get("emp_name", "")))
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/hse/tbt/<int:tbt_id>/photo", methods=["POST"])
@api_hse_officer_required
def api_hse_tbt_photo(tbt_id):
    u = get_api_user()
    t = HseTbt.query.filter_by(id=tbt_id, officer_id=u.id).first_or_404()
    f = freq.files.get("photo")
    path = _save_hse_photo(f, "tbt", company_id=api_cid())
    if not path:
        return jsonify({"error": "Invalid or missing photo"}), 400
    if t.sign_photo_path:
        try: os.remove(os.path.join(HSE_UPLOAD_DIR, t.sign_photo_path))
        except Exception: pass
    t.sign_photo_path = path
    db.session.commit()
    return jsonify({"ok": True, "path": path})

@app.route("/api/hse/tbt/<int:tbt_id>", methods=["DELETE"])
@api_hse_officer_required
def api_hse_tbt_delete_api(tbt_id):
    u = get_api_user()
    t = HseTbt.query.filter_by(id=tbt_id, officer_id=u.id).first_or_404()
    HseTbtAttendance.query.filter_by(tbt_id=t.id).delete()
    db.session.delete(t)
    db.session.commit()
    return jsonify({"ok": True})

# ── Near Miss ──────────────────────────────────────────
@app.route("/api/hse/nearmiss", methods=["GET"])
@api_hse_officer_required
def api_hse_nearmiss_list():
    u = get_api_user()
    page = int(freq.args.get("page", 1))
    pg = HseNearMiss.query.filter_by(officer_id=u.id)\
        .order_by(HseNearMiss.date.desc()).paginate(page=page, per_page=15, error_out=False)
    items = [{"id": nm.id, "date": nm.date.isoformat(), "location": nm.location or "",
              "description": nm.description or "", "immediate_cause": nm.immediate_cause or "",
              "action_taken": nm.action_taken or "", "reported_to": nm.reported_to or ""} for nm in pg.items]
    return jsonify({"items": items, "page": pg.page, "pages": pg.pages, "total": pg.total})

@app.route("/api/hse/nearmiss", methods=["POST"])
@api_hse_officer_required
def api_hse_nearmiss_create():
    u = get_api_user()
    data = freq.get_json()
    today = datetime.now(RIYADH_TZ).date()
    description = (data.get("description") or "").strip()
    immediate_cause = (data.get("immediate_cause") or "").strip()
    action_taken = (data.get("action_taken") or "").strip()
    reported_to = (data.get("reported_to") or "").strip()
    if not all([description, immediate_cause, action_taken, reported_to]):
        return jsonify({"error": "All fields required"}), 400
    nm = HseNearMiss(officer_id=u.id, date=_parse_date(data.get("date"), today),
                     location=(data.get("location") or "").strip(),
                     description=description, immediate_cause=immediate_cause,
                     action_taken=action_taken, reported_to=reported_to,
                     company_id=api_cid())
    db.session.add(nm)
    db.session.commit()
    return jsonify({"ok": True, "id": nm.id})

@app.route("/api/hse/nearmiss/<int:nm_id>", methods=["PUT"])
@api_hse_officer_required
def api_hse_nearmiss_update(nm_id):
    u = get_api_user()
    nm = HseNearMiss.query.filter_by(id=nm_id, officer_id=u.id).first_or_404()
    data = freq.get_json()
    nm.location        = (data.get("location")        or nm.location or "").strip()
    nm.description     = (data.get("description")     or nm.description or "").strip()
    nm.immediate_cause = (data.get("immediate_cause") or nm.immediate_cause or "").strip()
    nm.action_taken    = (data.get("action_taken")    or nm.action_taken or "").strip()
    nm.reported_to     = (data.get("reported_to")     or nm.reported_to or "").strip()
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/hse/nearmiss/<int:nm_id>/photo", methods=["POST"])
@api_hse_officer_required
def api_hse_nearmiss_photo(nm_id):
    u = get_api_user()
    nm = HseNearMiss.query.filter_by(id=nm_id, officer_id=u.id).first_or_404()
    f = freq.files.get("photo")
    path = _save_hse_photo(f, "nm", company_id=api_cid())
    if not path:
        return jsonify({"error": "Invalid or missing photo"}), 400
    if nm.photo_path:
        try: os.remove(os.path.join(HSE_UPLOAD_DIR, nm.photo_path))
        except Exception: pass
    nm.photo_path = path
    db.session.commit()
    return jsonify({"ok": True, "path": path})

@app.route("/api/hse/nearmiss/<int:nm_id>", methods=["DELETE"])
@api_hse_officer_required
def api_hse_nearmiss_delete_api(nm_id):
    u = get_api_user()
    nm = HseNearMiss.query.filter_by(id=nm_id, officer_id=u.id).first_or_404()
    db.session.delete(nm)
    db.session.commit()
    return jsonify({"ok": True})

# ── BBS ────────────────────────────────────────────────
@app.route("/api/hse/bbs/today", methods=["GET"])
@api_hse_officer_required
def api_hse_bbs_today():
    u = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    rec = HseBbs.query.filter_by(officer_id=u.id, date=today).first()
    if not rec:
        return jsonify(None)
    return jsonify({"id": rec.id, "date": rec.date.isoformat(),
                    "card_count": rec.card_count, "notes": rec.notes or ""})

@app.route("/api/hse/bbs", methods=["GET"])
@api_hse_officer_required
def api_hse_bbs_list():
    u = get_api_user()
    recs = HseBbs.query.filter_by(officer_id=u.id).order_by(HseBbs.date.desc()).limit(30).all()
    return jsonify([{"id": r.id, "date": r.date.isoformat(),
                     "card_count": r.card_count, "notes": r.notes or ""} for r in recs])

@app.route("/api/hse/bbs", methods=["POST"])
@api_hse_officer_required
def api_hse_bbs_save():
    u = get_api_user()
    data = freq.get_json()
    today = datetime.now(RIYADH_TZ).date()
    card_count = int(data.get("card_count") or 0)
    notes = (data.get("notes") or "").strip()
    rec = HseBbs.query.filter_by(officer_id=u.id, date=today).first()
    if rec:
        rec.card_count = card_count
        rec.notes = notes
    else:
        rec = HseBbs(officer_id=u.id, date=today, card_count=card_count,
                     notes=notes, company_id=api_cid())
        db.session.add(rec)
    db.session.commit()
    return jsonify({"ok": True, "id": rec.id})

@app.route("/api/hse/bbs/<int:bbs_id>", methods=["DELETE"])
@api_hse_officer_required
def api_hse_bbs_delete_api(bbs_id):
    u = get_api_user()
    rec = HseBbs.query.filter_by(id=bbs_id, officer_id=u.id).first_or_404()
    db.session.delete(rec)
    db.session.commit()
    return jsonify({"ok": True})

# ── Employee Lookup (for TBT attendance) ──────────────
@app.route("/api/hse/employee_lookup", methods=["GET"])
@api_hse_officer_required
def api_hse_employee_lookup():
    emp_number = (freq.args.get("emp_number") or "").strip()
    if not emp_number:
        return jsonify({"found": False})
    emp = Employee.query.filter_by(emp_number=emp_number, status="active").first()
    if not emp:
        for prefix in ["NSH-", "GA-"]:
            if emp_number.startswith(prefix):
                emp = Employee.query.filter_by(emp_number=emp_number[len(prefix):], status="active").first()
                if emp:
                    break
        if not emp:
            for prefix in ["NSH-", "GA-"]:
                emp = Employee.query.filter_by(emp_number=prefix + emp_number, status="active").first()
                if emp:
                    break
    if emp:
        return jsonify({"found": True, "name": emp.name, "emp_number": emp.emp_number})
    return jsonify({"found": False})

# ── HSE Dashboard (supervisor 39468 only) ─────────────
@app.route("/api/hse/dashboard", methods=["GET"])
@api_hse_supervisor_required
def api_hse_dashboard():
    today = datetime.now(RIYADH_TZ).date()
    days_since_sunday = today.isoweekday() % 7
    week_start = today - timedelta(days=days_since_sunday)
    officers = User.query.filter(User.role == "safety_officer",
                                 User.is_active == True).all()
    data = []
    for o in officers:
        ci = HseCheckin.query.filter_by(officer_id=o.id, date=today).first()
        obs_w = HseObservation.query.filter(HseObservation.officer_id == o.id,
                                            HseObservation.date >= week_start).count()
        jso_w = HseJsoClosure.query.filter(HseJsoClosure.officer_id == o.id,
                                           HseJsoClosure.date >= week_start).count()
        tbt_w = HseTbt.query.filter(HseTbt.officer_id == o.id,
                                    HseTbt.date >= week_start).count()
        nm_w  = HseNearMiss.query.filter(HseNearMiss.officer_id == o.id,
                                         HseNearMiss.date >= week_start).count()
        bbs_w = db.session.query(db.func.sum(HseBbs.card_count)).filter(
            HseBbs.officer_id == o.id, HseBbs.date >= week_start).scalar() or 0
        data.append({"id": o.id, "name": o.name, "checked_in": ci is not None,
                     "location": ci.location if ci else None,
                     "obs_week": obs_w, "jso_week": jso_w, "tbt_week": tbt_w,
                     "nm_week": nm_w, "bbs_week": int(bbs_w),
                     "total_week": obs_w + jso_w + tbt_w + nm_w})
    high_risk = HseObservation.query.filter_by(status="open", risk_level="H")\
        .order_by(HseObservation.date).all()
    hr_items = []
    for o in high_risk:
        off = db.session.get(User, o.officer_id)
        hr_items.append({"id": o.id, "date": o.date.isoformat(),
                         "officer": off.name if off else "?",
                         "category": o.category or "", "location": o.location or "",
                         "description": o.description or ""})
    return jsonify({"today": today.isoformat(), "week_start": week_start.isoformat(),
                    "officers": data, "high_risk_open": hr_items})


@app.get("/api/hse/photo/<path:filename>")
@api_login_required
def api_hse_photo_serve(filename):
    from flask import send_from_directory
    return send_from_directory(HSE_UPLOAD_DIR, filename)


@app.route("/api/hse/officer/<int:officer_id>/detail", methods=["GET"])
@api_hse_supervisor_required
def api_hse_officer_detail(officer_id):
    days = min(int(freq.args.get("days", 30)), 180)
    officer = User.query.filter_by(id=officer_id, role="safety_officer",
                                   is_active=True).first_or_404()
    today = datetime.now(RIYADH_TZ).date()
    since = today - timedelta(days=days)
    ci = HseCheckin.query.filter_by(officer_id=officer_id, date=today).first()
    checkin_hist = HseCheckin.query.filter(
        HseCheckin.officer_id == officer_id,
        HseCheckin.date >= since
    ).order_by(HseCheckin.date.desc()).all()
    obs_list = HseObservation.query.filter(
        HseObservation.officer_id == officer_id,
        HseObservation.date >= since
    ).order_by(HseObservation.date.desc()).all()
    tbt_list = HseTbt.query.filter(
        HseTbt.officer_id == officer_id,
        HseTbt.date >= since
    ).order_by(HseTbt.date.desc()).all()
    jso_list = HseJsoClosure.query.filter(
        HseJsoClosure.officer_id == officer_id,
        HseJsoClosure.date >= since
    ).order_by(HseJsoClosure.date.desc()).all()
    nm_list = HseNearMiss.query.filter(
        HseNearMiss.officer_id == officer_id,
        HseNearMiss.date >= since
    ).order_by(HseNearMiss.date.desc()).all()
    bbs_list = HseBbs.query.filter(
        HseBbs.officer_id == officer_id,
        HseBbs.date >= since
    ).order_by(HseBbs.date.desc()).all()
    return jsonify({
        "officer": {"id": officer.id, "name": officer.name},
        "checkin_today": {"id": ci.id, "date": ci.date.isoformat(),
                          "location": ci.location} if ci else None,
        "checkin_history": [{"id": c.id, "date": c.date.isoformat(),
                              "location": c.location} for c in checkin_hist],
        "period_days": days,
        "observations": [{
            "id": o.id, "date": o.date.isoformat(), "location": o.location or "",
            "obs_type": o.obs_type or "", "category": o.category or "",
            "risk_level": o.risk_level or "", "description": o.description or "",
            "action_taken": o.action_taken or "", "status": o.status,
            "closed_at": o.closed_at.isoformat() if o.closed_at else None,
            "closure_action": o.closure_action or "",
            "photos": [{"path": p.photo_path, "photo_type": p.photo_type}
                       for p in HseObservationPhoto.query.filter_by(observation_id=o.id).all()]
        } for o in obs_list],
        "tbts": [{
            "id": t.id, "date": t.date.isoformat(), "topic": t.topic or "",
            "location": t.location or "",
            "sign_photo_path": t.sign_photo_path or "",
            "supervisor_name": (db.session.get(User, t.supervisor_id).name
                                if t.supervisor_id and db.session.get(User, t.supervisor_id) else ""),
            "attendee_count": HseTbtAttendance.query.filter_by(tbt_id=t.id).count(),
            "attendees": [{"emp_number": a.emp_number, "emp_name": a.emp_name}
                          for a in HseTbtAttendance.query.filter_by(tbt_id=t.id).all()]
        } for t in tbt_list],
        "jso_closures": [{
            "id": j.id, "date": j.date.isoformat(), "jso_number": j.jso_number,
            "location": j.location or "", "action_taken": j.action_taken or "",
            "photo_path": j.photo_path or ""
        } for j in jso_list],
        "near_misses": [{
            "id": n.id, "date": n.date.isoformat(), "location": n.location or "",
            "description": n.description or "", "immediate_cause": n.immediate_cause or "",
            "action_taken": n.action_taken or "", "reported_to": n.reported_to or "",
            "photo_path": n.photo_path or ""
        } for n in nm_list],
        "bbs": [{
            "id": b.id, "date": b.date.isoformat(),
            "card_count": b.card_count, "notes": b.notes or ""
        } for b in bbs_list]
    })


# ===================== HSE Phase 2 Mobile APIs =====================

@app.get("/api/hse/ptw")
@api_hse_officer_required
def api_hse_ptw_list():
    u = get_api_user()
    ptw_list = (HsePtw.query
                .filter_by(officer_id=u.id)
                .order_by(HsePtw.week_start.desc())
                .limit(50).all())
    today = datetime.now(RIYADH_TZ).date()
    return jsonify([{
        "id": p.id,
        "permit_number": p.permit_number,
        "permit_type": p.permit_type,
        "description": p.description or "",
        "location": p.location or "",
        "week_start": p.week_start.isoformat(),
        "week_end": p.week_end.isoformat(),
        "status": p.status,
        "attached_to_id": p.attached_to_id,
        "expired": p.week_end < today,
    } for p in ptw_list])


@app.post("/api/hse/ptw")
@api_hse_officer_required
def api_hse_ptw_create():
    u = get_api_user()
    data = freq.get_json(silent=True) or {}
    permit_number = (data.get("permit_number") or "").strip()
    permit_type   = (data.get("permit_type") or "").strip()
    if not permit_number or not permit_type:
        return jsonify({"error": "permit_number and permit_type required"}), 400
    ws, we = _ptw_current_week()
    if data.get("week_start"):
        try:
            ws = date.fromisoformat(data["week_start"])
            we = ws + timedelta(days=5)
        except ValueError:
            pass
    loc = (data.get("location") or "").strip() or _hse_today_location(u.id)
    p = HsePtw(
        officer_id=u.id, company_id=u.company_id,
        permit_number=permit_number, permit_type=permit_type,
        description=(data.get("description") or "").strip(),
        location=loc, week_start=ws, week_end=we,
        attached_to_id=data.get("attached_to_id"),
    )
    db.session.add(p)
    db.session.commit()
    return jsonify({"id": p.id, "status": "created"}), 201


@app.post("/api/hse/ptw/<int:ptw_id>/status")
@api_hse_officer_required
def api_hse_ptw_status(ptw_id):
    u = get_api_user()
    p = HsePtw.query.filter_by(id=ptw_id, officer_id=u.id).first_or_404()
    data = freq.get_json(silent=True) or {}
    new_status = data.get("status", "active")
    if new_status not in ("active", "suspended", "closed"):
        return jsonify({"error": "invalid status"}), 400
    if new_status == "active" and p.week_end < datetime.now(RIYADH_TZ).date():
        ws, we = _ptw_current_week()
        p.week_start = ws
        p.week_end   = we
    p.status = new_status
    db.session.commit()
    return jsonify({"id": p.id, "status": p.status})


@app.get("/api/hse/manpower/today")
@api_hse_officer_required
def api_hse_manpower_today():
    u = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    rec = HseManpower.query.filter_by(officer_id=u.id, date=today).first()
    if not rec:
        return jsonify(None)
    import json as _json
    return jsonify({
        "id": rec.id, "date": rec.date.isoformat(),
        "location": rec.location or "",
        "total_count": rec.total_count,
        "breakdown": _json.loads(rec.breakdown) if rec.breakdown else {},
        "notes": rec.notes or "",
    })


@app.post("/api/hse/manpower")
@api_hse_officer_required
def api_hse_manpower_save():
    u = get_api_user()
    data = freq.get_json(silent=True) or {}
    today = datetime.now(RIYADH_TZ).date()
    total = int(data.get("total_count", 0))
    loc   = (data.get("location") or "").strip() or _hse_today_location(u.id)
    notes = (data.get("notes") or "").strip()
    import json as _json
    breakdown = data.get("breakdown") or {}
    breakdown_json = _json.dumps(breakdown, ensure_ascii=False) if breakdown else None

    existing = HseManpower.query.filter_by(officer_id=u.id, date=today).first()
    if existing:
        existing.total_count = total
        existing.location    = loc
        existing.notes       = notes
        existing.breakdown   = breakdown_json
    else:
        existing = HseManpower(
            officer_id=u.id, company_id=u.company_id,
            date=today, total_count=total,
            location=loc, notes=notes, breakdown=breakdown_json,
        )
        db.session.add(existing)
    db.session.commit()
    return jsonify({"id": existing.id, "status": "saved"})


@app.get("/api/hse/manpower/history")
@api_hse_officer_required
def api_hse_manpower_history():
    u = get_api_user()
    import json as _json
    recs = (HseManpower.query.filter_by(officer_id=u.id)
            .order_by(HseManpower.date.desc()).limit(14).all())
    return jsonify([{
        "id": r.id, "date": r.date.isoformat(),
        "location": r.location or "",
        "total_count": r.total_count,
        "breakdown": _json.loads(r.breakdown) if r.breakdown else {},
        "notes": r.notes or "",
    } for r in recs])


@app.get("/api/hse/inspection/today")
@api_hse_officer_required
def api_hse_inspection_today():
    u = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    rec = HseInspection.query.filter_by(officer_id=u.id, date=today).first()
    if not rec:
        return jsonify(None)
    import json as _json
    return jsonify({
        "id": rec.id, "date": rec.date.isoformat(),
        "location": rec.location or "",
        "overall_score": rec.overall_score,
        "notes": rec.notes or "",
        "checklist": _json.loads(rec.checklist) if rec.checklist else [],
    })


@app.post("/api/hse/inspection")
@api_hse_officer_required
def api_hse_inspection_save():
    u = get_api_user()
    data = freq.get_json(silent=True) or {}
    today = datetime.now(RIYADH_TZ).date()
    import json as _json
    loc    = (data.get("location") or "").strip() or _hse_today_location(u.id)
    notes  = (data.get("notes") or "").strip()
    items  = data.get("checklist", [])
    if not items:
        items = [{"item": t, "ok": False, "note": ""} for t in INSPECTION_ITEMS]
    ok_count = sum(1 for i in items if i.get("ok"))
    score    = round(ok_count / len(items) * 100, 1) if items else 0.0

    existing = HseInspection.query.filter_by(officer_id=u.id, date=today).first()
    if existing:
        existing.location      = loc
        existing.checklist     = _json.dumps(items, ensure_ascii=False)
        existing.overall_score = score
        existing.notes         = notes
    else:
        existing = HseInspection(
            officer_id=u.id, company_id=u.company_id,
            date=today, location=loc,
            checklist=_json.dumps(items, ensure_ascii=False),
            overall_score=score, notes=notes,
        )
        db.session.add(existing)
    db.session.commit()
    return jsonify({"id": existing.id, "score": score, "status": "saved"})


@app.get("/api/hse/inspection/history")
@api_hse_officer_required
def api_hse_inspection_history():
    u = get_api_user()
    import json as _json
    recs = (HseInspection.query.filter_by(officer_id=u.id)
            .order_by(HseInspection.date.desc()).limit(10).all())
    return jsonify([{
        "id": r.id, "date": r.date.isoformat(),
        "location": r.location or "",
        "overall_score": r.overall_score,
        "notes": r.notes or "",
    } for r in recs])


@app.get("/api/hse/my-ca")
@api_hse_officer_required
def api_hse_my_ca():
    u     = get_api_user()
    today = datetime.now(RIYADH_TZ).date()
    cas   = (HseCorrectiveAction.query
             .join(HseObservation, HseCorrectiveAction.observation_id == HseObservation.id)
             .filter(HseObservation.officer_id == u.id)
             .order_by(HseCorrectiveAction.due_date.asc())
             .all())
    return jsonify([{
        "id":              c.id,
        "observation_id":  c.observation_id,
        "assigned_to":     c.assigned_to or "",
        "due_date":        c.due_date.isoformat(),
        "action_required": c.action_required,
        "status":          c.status,
        "overdue":         c.status != "completed" and c.due_date < today,
        "completed_at":    c.completed_at.isoformat() if c.completed_at else None,
        "completion_notes": c.completion_notes or "",
        "created_at":      c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
    } for c in cas])


# ===================== Bootstrapping =====================
with app.app_context():
    try:
        db.create_all()
        seed_warning_reasons()
    except Exception as e:
        app.logger.warning("seed_warning_reasons skipped: %s", e)

# على الاستضافة: Passenger يستورد الملف فقط، لذا ننادي الإنشاء هنا:
ensure_db_and_admin()

# ── Privacy Policy (public, no login required) ─────────────────
@app.route("/privacy")
def privacy_policy():
    return render_template("privacy.html")


# ═══════════════════════════════════════════════════════════════
#  Super Admin Routes  (supervisor_code = 39468, role = super_admin)
# ═══════════════════════════════════════════════════════════════

@app.route("/sa/")
@super_admin_required
def superadmin_dashboard():
    total_companies = Company.query.count()
    active_companies = Company.query.filter_by(is_active=True).count()
    pending_companies = Company.query.filter_by(is_active=False).count()
    total_users = User.query.filter(User.role != "super_admin").count()
    recent = Company.query.order_by(Company.created_at.desc()).limit(5).all()
    pending = Company.query.filter_by(is_active=False).order_by(Company.created_at.desc()).all()
    return render_template("superadmin_dashboard.html",
                           total_companies=total_companies,
                           active_companies=active_companies,
                           pending_companies=pending_companies,
                           total_users=total_users,
                           recent=recent,
                           pending=pending)


@app.route("/sa/companies")
@super_admin_required
def superadmin_companies():
    q = request.args.get("q", "").strip()
    query = Company.query
    if q:
        query = query.filter(Company.name.ilike(f"%{q}%"))
    companies = query.order_by(Company.created_at.desc()).all()
    # Attach user count
    for co in companies:
        co._user_count = User.query.filter_by(company_id=co.id).count()
    return render_template("superadmin_companies.html", companies=companies, q=q)


@app.route("/sa/companies/<int:company_id>")
@super_admin_required
def superadmin_company_detail(company_id):
    co = Company.query.get_or_404(company_id)
    users = User.query.filter_by(company_id=company_id).order_by(User.name).all()
    admin_user = next((u for u in users if u.role == "admin" and u.is_active), None) or \
                 next((u for u in users if u.is_active), None)
    return render_template("superadmin_company.html", co=co, users=users, admin_user=admin_user)


@app.post("/sa/companies/<int:company_id>/activate")
@super_admin_required
def superadmin_activate(company_id):
    co = Company.query.get_or_404(company_id)
    co.is_active = True
    db.session.commit()
    flash(f"Company '{co.name}' activated.", "success")
    return redirect(url_for("superadmin_company_detail", company_id=company_id))


@app.post("/sa/companies/<int:company_id>/deactivate")
@super_admin_required
def superadmin_deactivate(company_id):
    co = Company.query.get_or_404(company_id)
    co.is_active = False
    db.session.commit()
    flash(f"Company '{co.name}' deactivated.", "warning")
    return redirect(url_for("superadmin_company_detail", company_id=company_id))


@app.post("/sa/companies/create")
@super_admin_required
def superadmin_create_company():
    company_name = (request.form.get("company_name") or "").strip()
    admin_name   = (request.form.get("admin_name") or "").strip()
    admin_email  = (request.form.get("admin_email") or "").strip().lower()
    admin_pw     = (request.form.get("admin_password") or "").strip()

    if not all([company_name, admin_name, admin_email, admin_pw]):
        flash("All fields are required.", "danger")
        return redirect(url_for("superadmin_companies"))

    if User.query.filter_by(email=admin_email).first():
        flash("Email already in use.", "danger")
        return redirect(url_for("superadmin_companies"))

    import re, secrets as _sec
    slug = re.sub(r"[^a-z0-9]+", "-", company_name.lower()).strip("-")
    base_slug = slug
    counter = 1
    while Company.query.filter_by(slug=slug).first():
        slug = f"{base_slug}-{counter}"
        counter += 1

    co = Company(name=company_name, slug=slug, owner_email=admin_email, is_active=True)
    db.session.add(co)
    db.session.flush()

    sup_code = f"adm-{_sec.token_hex(4)}"
    while User.query.filter_by(supervisor_code=sup_code).first():
        sup_code = f"adm-{_sec.token_hex(4)}"

    u = User(supervisor_code=sup_code, name=admin_name, email=admin_email,
             role="admin", is_active=True, company_id=co.id)
    u.set_password(admin_pw)
    db.session.add(u)
    db.session.commit()

    flash(f"Company '{company_name}' created and activated.", "success")
    return redirect(url_for("superadmin_company_detail", company_id=co.id))


@app.post("/sa/companies/<int:company_id>/set-max-users")
@super_admin_required
def superadmin_set_max_users(company_id):
    co = Company.query.get_or_404(company_id)
    try:
        co.max_users = max(1, int(request.form.get("max_users", 50)))
        db.session.commit()
        flash(f"Max users updated to {co.max_users}.", "success")
    except (ValueError, TypeError):
        flash("Invalid value.", "danger")
    return redirect(url_for("superadmin_company_detail", company_id=company_id))


@app.post("/sa/companies/<int:company_id>/set-plan")
@super_admin_required
def superadmin_set_plan(company_id):
    co = Company.query.get_or_404(company_id)
    plan = request.form.get("plan", "free")
    if plan in ("free", "pro"):
        co.plan = plan
        db.session.commit()
        flash(f"Plan updated to {plan}.", "success")
    return redirect(url_for("superadmin_company_detail", company_id=company_id))


@app.route("/sa/impersonate/<int:user_id>")
@super_admin_required
def superadmin_impersonate(user_id):
    target = User.query.get_or_404(user_id)
    if target.role == "super_admin":
        flash("Cannot impersonate super admin.", "danger")
        return redirect(url_for("superadmin_dashboard"))
    session["original_user_id"] = session.get("user_id")
    session["user_id"] = target.id
    session["company_id"] = target.company_id
    session["impersonating"] = True
    flash(f"Now viewing as {target.name} ({target.role})", "info")
    return redirect(url_for("index"))


@app.route("/sa/stop-impersonate")
@login_required
def superadmin_stop_impersonate():
    if not session.get("impersonating"):
        return redirect(url_for("index"))
    original_id = session.get("original_user_id")
    session.clear()
    if original_id:
        session["user_id"] = original_id
    return redirect(url_for("superadmin_dashboard"))


# ── Super Admin API (for iOS app) ─────────────────────────────

def _sa_token_user():
    """Validates Bearer token and returns user if super_admin, else None."""
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        return None
    token_val = auth[7:]
    mt = MobileToken.query.filter_by(token=token_val).first()
    if not mt:
        return None
    u = db.session.get(User, mt.user_id)
    if not u or u.role != "super_admin":
        return None
    return u

@app.get("/api/sa/dashboard")
def api_sa_dashboard():
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    total     = Company.query.count()
    active    = Company.query.filter_by(is_active=True).count()
    pending   = Company.query.filter_by(is_active=False).count()
    users_cnt = User.query.filter(User.role != "super_admin").count()
    recent    = Company.query.order_by(Company.created_at.desc()).limit(5).all()
    pending_list = Company.query.filter_by(is_active=False).order_by(Company.created_at.desc()).all()
    def co_dict(co):
        return {
            "id": co.id, "name": co.name, "slug": co.slug,
            "plan": co.plan, "is_active": bool(co.is_active),
            "owner_email": co.owner_email,
            "user_count": User.query.filter_by(company_id=co.id).count(),
            "created_at": co.created_at.strftime("%Y-%m-%d") if co.created_at else None
        }
    return jsonify({
        "total_companies": total,
        "active_companies": active,
        "pending_companies": pending,
        "total_users": users_cnt,
        "recent": [co_dict(c) for c in recent],
        "pending": [co_dict(c) for c in pending_list]
    })

@app.get("/api/sa/companies")
def api_sa_companies():
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    q = request.args.get("q", "").strip()
    query = Company.query
    if q:
        query = query.filter(Company.name.ilike(f"%{q}%"))
    companies = query.order_by(Company.created_at.desc()).all()
    return jsonify([{
        "id": co.id, "name": co.name, "slug": co.slug,
        "plan": co.plan, "is_active": bool(co.is_active),
        "owner_email": co.owner_email,
        "user_count": User.query.filter_by(company_id=co.id).count(),
        "created_at": co.created_at.strftime("%Y-%m-%d") if co.created_at else None
    } for co in companies])

@app.get("/api/sa/companies/<int:company_id>")
def api_sa_company_detail(company_id):
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    co = Company.query.get_or_404(company_id)
    users = User.query.filter_by(company_id=company_id).order_by(User.name).all()
    return jsonify({
        "id": co.id, "name": co.name, "slug": co.slug,
        "plan": co.plan, "is_active": bool(co.is_active),
        "owner_email": co.owner_email, "max_users": co.max_users,
        "created_at": co.created_at.strftime("%Y-%m-%d") if co.created_at else None,
        "users": [{
            "id": u.id, "name": u.name, "role": u.role,
            "email": u.email, "supervisor_code": u.supervisor_code,
            "is_active": bool(u.is_active)
        } for u in users]
    })

@app.post("/api/sa/companies/<int:company_id>/activate")
def api_sa_activate(company_id):
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    co = Company.query.get_or_404(company_id)
    co.is_active = True
    db.session.commit()
    return jsonify({"ok": True})

@app.post("/api/sa/companies/<int:company_id>/deactivate")
def api_sa_deactivate(company_id):
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    co = Company.query.get_or_404(company_id)
    co.is_active = False
    db.session.commit()
    return jsonify({"ok": True})

@app.post("/api/sa/companies/create")
def api_sa_create_company():
    sa = _sa_token_user()
    if not sa:
        return jsonify({"error": "forbidden"}), 403
    data = request.get_json(force=True) or {}
    name  = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip().lower()
    plan  = data.get("plan", "free")
    if not name:
        return jsonify({"error": "name required"}), 400
    import re, unicodedata
    slug_base = re.sub(r"[^a-z0-9]+", "-", unicodedata.normalize("NFKD", name).encode("ascii","ignore").decode().lower()).strip("-") or "co"
    slug = slug_base
    i = 1
    while Company.query.filter_by(slug=slug).first():
        slug = f"{slug_base}-{i}"; i += 1
    co = Company(name=name, slug=slug, plan=plan, is_active=False, owner_email=email or None)
    db.session.add(co)
    db.session.flush()
    pwd = data.get("password") or ""
    admin_u = User(
        name=name + " Admin", supervisor_code=str(co.id) + "adm",
        role="admin", is_active=True, company_id=co.id,
        email=email or None
    )
    if pwd:
        admin_u.set_password(pwd)
    db.session.add(admin_u)
    db.session.commit()
    return jsonify({"ok": True, "company_id": co.id}), 201

@app.delete("/api/sa/companies/<int:company_id>")
def api_sa_delete_company(company_id):
    if not _sa_token_user():
        return jsonify({"error": "forbidden"}), 403
    co = db.session.get(Company, company_id)
    if not co:
        return jsonify({"error": "company not found"}), 404

    user_count = User.query.filter_by(company_id=company_id).count()
    emp_count  = Employee.query.filter_by(company_id=company_id).count()
    if user_count or emp_count:
        return jsonify({
            "error": "cannot delete company with existing data",
            "users": user_count,
            "employees": emp_count,
            "hint": "deactivate all users and employees first",
        }), 409

    db.session.delete(co)
    db.session.commit()
    return jsonify({"ok": True})


# ── Company Settings (admin only) ────────────────────────────

@app.route("/settings/company", methods=["GET", "POST"])
@admin_required
def settings_company():
    u = cur_user()
    co = db.session.get(Company, u.company_id)
    if not co:
        abort(404)
    if request.method == "POST":
        new_name = (request.form.get("name") or "").strip()
        if not new_name:
            flash("Company name cannot be empty.", "danger")
        else:
            co.name = new_name
            db.session.commit()
            flash("Company name updated.", "success")
        return redirect(url_for("settings_company"))
    return render_template("settings_company.html", co=co)


@app.route("/settings/users")
@admin_required
def settings_users():
    u = cur_user()
    users = (apply_company_filter(User.query, User)
             .filter(User.is_active == True)
             .order_by(User.role, User.name)
             .all())
    co = db.session.get(Company, u.company_id)
    return render_template("settings_users.html", users=users, co=co)


@app.post("/settings/users/<int:uid>/deactivate")
@admin_required
def settings_user_deactivate(uid):
    u = cur_user()
    if not u.company_id:
        abort(403)
    target = User.query.filter(
        User.id == uid, User.company_id == u.company_id
    ).first_or_404()
    if target.id == u.id:
        flash("You cannot deactivate yourself.", "danger")
    else:
        target.is_active = False
        db.session.commit()
        flash(f"{target.name} deactivated.", "success")
    return redirect(url_for("settings_users"))


@app.post("/settings/users/<int:uid>/activate")
@admin_required
def settings_user_activate(uid):
    u = cur_user()
    if not u.company_id:
        abort(403)
    target = User.query.filter(
        User.id == uid, User.company_id == u.company_id
    ).first_or_404()
    target.is_active = True
    db.session.commit()
    flash(f"{target.name} activated.", "success")
    return redirect(url_for("settings_users"))


# نقطة دخول WSGI لاسم "application"
application = app

# للتشغيل المحلي فقط
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080, debug=True)